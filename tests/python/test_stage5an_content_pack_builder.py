from __future__ import annotations

import json
from pathlib import Path

import yaml

from libreprimus.deep_research_export.content_pack import build_content_pack


def write_stage5an_fixture(tmp_path: Path) -> dict[str, Path]:
    ingest = tmp_path / "ingest"
    ingest.mkdir()
    research = tmp_path / "research-inputs" / "stage5fixture" / "bundle"
    research.mkdir(parents=True)
    (research / "deep_research_context.md").write_text("context without solve claim\n", encoding="utf-8")
    (research / "known_questions.md").write_text("what source is reliable?\n", encoding="utf-8")
    safe = tmp_path / "safe"
    safe.mkdir()
    (safe / "important_links.txt").write_text("https://example.org/source\n", encoding="utf-8")
    _write_xlsx(safe / "LP Excel.xlsx")
    metadata = tmp_path / "metadata-site"
    metadata.mkdir()
    (metadata / "index.html").write_text(
        '<html><head><meta name="robots" content="noindex,nofollow,noarchive"></head><body>metadata</body></html>',
        encoding="utf-8",
    )
    bundle = {
        "bundle_id": "fixture-bundle",
        "title": "Fixture Bundle",
        "publication_status": "private_deep_research_only",
        "review_status": "review_required",
    }
    source = {
        "source_id": "fixture-source",
        "title": "Fixture Source",
        "publication_status": "private_deep_research_only",
        "review_status": "review_required",
        "solve_claim": False,
    }
    content = {
        "content_id": "fixture-content",
        "bundle_id": "fixture-bundle",
        "source_id": "fixture-source",
        "relative_path_or_ref": str(research / "deep_research_context.md"),
        "publication_status": "generated_extract_review_required",
        "review_status": "review_required",
        "content_kind": "markdown",
        "solve_claim": False,
    }
    claim = {
        "claim_id": "fixture-claim",
        "source_id": "fixture-source",
        "publication_status": "blocked_private_or_sensitive",
        "review_status": "review_required",
        "solve_claim": False,
    }
    datasets = {
        "research-index.json": {"record_type": "fixture_research_index", "records": []},
        "research-bundles.json": {"record_count": 1, "records": [bundle]},
        "source-cards.json": {"record_count": 1, "records": [source]},
        "content-index.json": {"record_count": 1, "records": [content]},
        "community-claims.json": {"record_count": 1, "records": [claim]},
        "publication-gates.json": {"record_count": 1, "records": [{"gate_id": "fixture-gate"}]},
        "missing-sources.json": {"record_count": 0, "records": []},
        "deep-research-export.json": {"record_type": "fixture_export", "records": []},
    }
    for name, payload in datasets.items():
        (ingest / name).write_text(json.dumps(payload), encoding="utf-8")
    (ingest / "summary.yaml").write_text(yaml.safe_dump({"status": "complete"}), encoding="utf-8")
    return {"ingest": ingest, "research": research.parent, "safe": safe, "metadata": metadata}


def build_fixture_pack(tmp_path: Path) -> dict[str, Path]:
    fixture = write_stage5an_fixture(tmp_path)
    out = tmp_path / "pack"
    data = tmp_path / "data"
    build_content_pack(
        metadata_site_root=fixture["metadata"],
        website_ingest_dir=fixture["ingest"],
        research_input_roots=[fixture["research"]],
        safe_local_source_roots=[fixture["safe"]],
        out_root=out,
        policy_out=data / "policy.yaml",
        inputs_out=data / "inputs.yaml",
        manifest_summary_out=data / "manifest-summary.yaml",
        file_selection_summary_out=data / "file-selection.yaml",
        publication_gate_audit_out=data / "publication-audit.yaml",
    )
    return {**fixture, "pack": out, "data": data}


def test_content_pack_manifest_references_existing_hashed_files(tmp_path: Path) -> None:
    fixture = build_fixture_pack(tmp_path)
    manifest = json.loads((fixture["pack"] / "deep-research-content-pack-stage5an-manifest.json").read_text(encoding="utf-8"))
    assert manifest["raw_third_party_files_included"] is False
    assert manifest["included_file_count"] >= 4
    for record in manifest["included_files"]:
        assert (fixture["pack"] / record["relative_path"]).exists()
        assert len(record["sha256"]) == 64
        assert record["publication_status"]


def _write_xlsx(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "value"
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.save(path)
