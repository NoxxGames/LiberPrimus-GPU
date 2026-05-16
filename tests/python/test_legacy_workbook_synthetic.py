from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from libreprimus.cli import app
from libreprimus.legacy_workbook.export import extract_workbook, write_extraction


def _add_delta_block(
    ws,
    *,
    label_column: int,
    start_row: int,
    label: str,
    cipher_indices: list[int],
    message_indices: list[int],
    cipher_minus: list[int],
    message_minus: list[int],
) -> None:
    ws.cell(start_row, label_column, label)
    ws.cell(start_row + 1, label_column, "Cipher Rune")
    ws.cell(start_row + 2, label_column, "Gematria Position")
    ws.cell(start_row + 3, label_column, "Message Rune")
    ws.cell(start_row + 4, label_column, "Gematria Position")
    ws.cell(start_row + 5, label_column, "Cipher-Message")
    ws.cell(start_row + 6, label_column, "Message - Cipher")
    for index, cipher_index in enumerate(cipher_indices):
        column = label_column + 1 + index
        ws.cell(start_row + 1, column, f"r{index}")
        ws.cell(start_row + 2, column, cipher_index)
        ws.cell(start_row + 3, column, f"T{index}")
        ws.cell(start_row + 4, column, message_indices[index])
        ws.cell(start_row + 5, column, cipher_minus[index])
        ws.cell(start_row + 6, column, message_minus[index])


def _create_synthetic_workbook(path: Path, *, bad_delta: bool = False) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    readme = workbook.create_sheet("README")
    readme["A1"] = "Synthetic legacy workbook"
    readme["B1"] = "=1+1"

    prime = workbook.create_sheet("Prime Sums")
    prime["A4"] = "SUM"
    prime["B4"] = "Is Prime?"
    for column, token in enumerate(["A", "B", "TH"], start=3):
        prime.cell(5, column, token)
    prime["A6"] = 10
    prime["B6"] = True
    for column, value in enumerate([2, 3, 5], start=3):
        prime.cell(6, column, value)
    for column, token in enumerate(["C", "D"], start=3):
        prime.cell(7, column, token)
    prime["A8"] = 16
    prime["B8"] = "PRAWDA"
    for column, value in enumerate([7, 9], start=3):
        prime.cell(8, column, value)
    for column, token in enumerate(["E", "F"], start=3):
        prime.cell(9, column, token)
    prime["A10"] = 20
    prime["B10"] = "FAŁSZ"
    for column, value in enumerate([11, 13], start=3):
        prime.cell(10, column, value)

    welcome = workbook.create_sheet("Welcome")
    _add_delta_block(
        welcome,
        label_column=2,
        start_row=2,
        label="Line 1",
        cipher_indices=[5, 8],
        message_indices=[3, 6],
        cipher_minus=[99 if bad_delta else 2, 2],
        message_minus=[27, 27],
    )

    p56 = workbook.create_sheet("p56 An End")
    _add_delta_block(
        p56,
        label_column=2,
        start_row=2,
        label="Line 1",
        cipher_indices=[25, 11],
        message_indices=[24, 9],
        cipher_minus=[1, 2],
        message_minus=[28, 27],
    )

    workbook.save(path)
    return path


def test_synthetic_inventory_classifies_sheets(tmp_path: Path) -> None:
    path = _create_synthetic_workbook(tmp_path / "synthetic.xlsx")
    extraction = extract_workbook(path)

    classifications = {record.sheet_name: record.classification for record in extraction.sheet_records}
    assert classifications == {
        "README": "readme",
        "Prime Sums": "prime_sums",
        "Welcome": "solved_delta_sheet",
        "p56 An End": "solved_delta_sheet",
    }


def test_synthetic_delta_extraction_and_validation(tmp_path: Path) -> None:
    path = _create_synthetic_workbook(tmp_path / "synthetic.xlsx")
    extraction = extract_workbook(path)

    welcome = [record for record in extraction.delta_records if record.sheet_name == "Welcome"]
    assert len(welcome) == 2
    assert welcome[0].cipher_minus_message_mod29 == 2
    assert welcome[0].message_minus_cipher_mod29 == 27
    assert not extraction.warning_records


def test_synthetic_prime_sums_parse_booleans(tmp_path: Path) -> None:
    path = _create_synthetic_workbook(tmp_path / "synthetic.xlsx")
    extraction = extract_workbook(path)

    assert [record.is_prime for record in extraction.prime_sum_records] == [True, True, False]
    assert extraction.prime_sum_records[1].raw_is_prime == "PRAWDA"
    assert extraction.prime_sum_records[2].raw_is_prime == "FAŁSZ"


def test_synthetic_formula_inventory(tmp_path: Path) -> None:
    path = _create_synthetic_workbook(tmp_path / "synthetic.xlsx")
    extraction = extract_workbook(path)

    assert len(extraction.formula_records) == 1
    assert extraction.formula_records[0].cell == "B1"
    assert extraction.formula_records[0].formula == "=1+1"
    assert extraction.formula_records[0].cached_value is None


def test_synthetic_extract_output_is_deterministic(tmp_path: Path) -> None:
    path = _create_synthetic_workbook(tmp_path / "synthetic.xlsx")
    first = tmp_path / "first"
    second = tmp_path / "second"

    write_extraction(first, extract_workbook(path))
    write_extraction(second, extract_workbook(path))

    for name in [
        "sheet_inventory.json",
        "solved_delta_rows.jsonl",
        "prime_sum_rows.jsonl",
        "formula_cells.jsonl",
        "summary.json",
        "warnings.jsonl",
    ]:
        assert (first / name).read_text(encoding="utf-8") == (second / name).read_text(
            encoding="utf-8"
        )


def test_cli_summary_inventory_extract_and_validate(tmp_path: Path) -> None:
    path = _create_synthetic_workbook(tmp_path / "synthetic.xlsx")
    out_dir = tmp_path / "out"
    inventory_path = tmp_path / "inventory.json"
    runner = CliRunner()

    result = runner.invoke(app, ["legacy-workbook", "summary", "--workbook", str(path)])
    assert result.exit_code == 0
    assert "sheet_count" in result.output

    result = runner.invoke(
        app,
        ["legacy-workbook", "inventory", "--workbook", str(path), "--out", str(inventory_path)],
    )
    assert result.exit_code == 0
    assert inventory_path.is_file()

    result = runner.invoke(
        app,
        ["legacy-workbook", "extract", "--workbook", str(path), "--out-dir", str(out_dir)],
    )
    assert result.exit_code == 0
    for name in [
        "sheet_inventory.json",
        "solved_delta_rows.jsonl",
        "prime_sum_rows.jsonl",
        "formula_cells.jsonl",
        "summary.json",
        "warnings.jsonl",
    ]:
        assert (out_dir / name).is_file()

    result = runner.invoke(app, ["legacy-workbook", "validate", "--workbook", str(path)])
    assert result.exit_code == 0


def test_cli_validate_fails_on_bad_delta_unless_allowed(tmp_path: Path) -> None:
    path = _create_synthetic_workbook(tmp_path / "bad.xlsx", bad_delta=True)
    runner = CliRunner()

    result = runner.invoke(app, ["legacy-workbook", "validate", "--workbook", str(path)])
    assert result.exit_code == 1

    result = runner.invoke(
        app,
        ["legacy-workbook", "validate", "--workbook", str(path), "--allow-warnings"],
    )
    assert result.exit_code == 0
