from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from libreprimus.source_harvester.xlsx_extraction import extract_xlsx_metadata


def test_stage5aj_xlsx_metadata_preserves_formatting_counts(tmp_path: Path) -> None:
    source_root = tmp_path / "UsefulFilesAndIdeas"
    source_root.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LP Delimiters"
    sheet["A1"] = "ᚠᚢ"
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    sheet["B1"] = "=1+1"
    sheet["C1"] = "link"
    sheet["C1"].hyperlink = "https://example.test/source"
    workbook.create_sheet("Rune Counts")
    workbook.save(source_root / "LP Excel.xlsx")

    result = extract_xlsx_metadata(
        source_root=source_root,
        results_dir=tmp_path / "results",
        out=tmp_path / "xlsx.yaml",
    )

    record = result["records"][0]
    assert result["xlsx_workbooks_detected"] == 1
    assert record["lp_delimiters_sheet_present"] is True
    assert record["rune_counts_sheet_present"] is True
    assert record["formula_cell_count"] == 1
    assert record["highlighted_cell_count"] == 1
    assert record["hyperlinks_count"] == 1
    assert record["raw_workbook_committed"] is False
