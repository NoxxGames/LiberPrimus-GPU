"""Safe local-source extract builders for Stage 5AN."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from .file_selection import sanitize_text, slugify
from .inputs import repo_relative, resolve, write_json
from .models import SAFE_EXTRACT_SOURCE_NAMES


def generate_safe_extracts(source_roots: list[Path], out_root: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Generate deterministic private extracts from allowlisted local source files."""

    selected: list[dict[str, str]] = []
    excluded: list[dict[str, str]] = []
    target_root = resolve(out_root)
    target_root.mkdir(parents=True, exist_ok=True)
    seen: set[Path] = set()
    for source_root in source_roots:
        root = resolve(source_root)
        if not root.exists():
            excluded.append({"path": repo_relative(source_root), "reason": "safe_local_source_root_missing"})
            continue
        for path in sorted(root.rglob("*")):
            if not path.is_file() or path.name not in SAFE_EXTRACT_SOURCE_NAMES or path in seen:
                continue
            seen.add(path)
            if path.suffix.lower() == ".xlsx":
                output = target_root / f"{slugify(path.stem)}-workbook-metadata.json"
                payload = extract_xlsx_metadata(path)
                write_json(output, payload)
                kind = "xlsx_workbook_metadata"
            elif path.suffix.lower() in {".txt", ".md"}:
                output = target_root / f"{slugify(path.stem)}.txt"
                text, findings = sanitize_text(path.read_text(encoding="utf-8", errors="replace"))
                output.write_text(text, encoding="utf-8")
                kind = "private_text_copy"
                if findings:
                    excluded.append({"path": repo_relative(path), "reason": ",".join(findings)})
            else:
                excluded.append({"path": repo_relative(path), "reason": "safe_extract_type_unsupported"})
                continue
            selected.append(
                {
                    "source_path": repo_relative(output),
                    "original_source_name": path.name,
                    "path_kind": "generated_safe_extract",
                    "content_kind": kind,
                    "publication_status": "private_deep_research_only",
                    "review_status": "review_required",
                    "raw_source_origin": "safe_extract_from_allowlisted_local_source",
                }
            )
    return selected, excluded


def extract_xlsx_metadata(path: Path, *, max_cells_per_sheet: int = 20000) -> dict[str, Any]:
    """Extract deterministic workbook metadata without copying workbook bytes."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is present in CI
        return {
            "record_type": "stage5an_xlsx_safe_extract",
            "source_filename": path.name,
            "extract_status": "openpyxl_unavailable",
            "error": str(exc),
        }

    workbook = load_workbook(resolve(path), read_only=False, data_only=True)
    sheets: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        non_empty = 0
        highlighted = 0
        sampled_cells = 0
        for row in worksheet.iter_rows():
            for cell in row:
                sampled_cells += 1
                if cell.value not in (None, ""):
                    non_empty += 1
                fill = getattr(cell, "fill", None)
                if fill is not None and fill.fill_type:
                    highlighted += 1
                if sampled_cells >= max_cells_per_sheet:
                    break
            if sampled_cells >= max_cells_per_sheet:
                break
        sheets.append(
            {
                "sheet_name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "sampled_cells": sampled_cells,
                "non_empty_sampled_cells": non_empty,
                "highlighted_sampled_cells": highlighted,
                "truncated": sampled_cells >= max_cells_per_sheet,
            }
        )
    workbook.close()
    return {
        "record_type": "stage5an_xlsx_safe_extract",
        "source_filename": path.name,
        "extract_status": "metadata_only",
        "raw_workbook_copied": False,
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def write_redaction_log(path: Path, records: list[dict[str, str]]) -> None:
    """Write redaction/exclusion records as JSONL."""

    target = resolve(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("".join(json.dumps(record, sort_keys=True) + "\n" for record in records), encoding="utf-8")
