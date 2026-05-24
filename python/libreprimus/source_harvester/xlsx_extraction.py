"""Stage 5AJ deterministic XLSX metadata extraction."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from .export import repo_relative, resolve, write_json, write_jsonl, write_records
from .hashing import hash_file
from .models import STAGE5AJ_ID, STAGE5AJ_OUTPUT_DIR, STAGE5AJ_REPORTS, STAGE5AJ_SOURCE_ROOT, STAGE5AJ_SOURCE_STAGE_ID, STAGE5AJ_XLSX_SUMMARY_PATH
from .usefulfiles import source_id_for_filename


def extract_xlsx_metadata(
    *,
    source_root: Path = STAGE5AJ_SOURCE_ROOT,
    results_dir: Path = STAGE5AJ_OUTPUT_DIR,
    out: Path = STAGE5AJ_XLSX_SUMMARY_PATH,
) -> dict[str, Any]:
    """Summarize local XLSX files without committing workbook bodies."""

    root = resolve(source_root)
    workbook_records: list[dict[str, Any]] = []
    cell_records: list[dict[str, Any]] = []
    for path in sorted(root.glob("*.xlsx"), key=lambda item: item.name.lower()) if root.exists() else []:
        record, cells = _summarize_workbook(path)
        workbook_records.append(record)
        cell_records.extend(cells)
    summary = {
        "record_type": "stage5aj_xlsx_extraction_summary",
        "schema": "schemas/source-harvester/xlsx-extraction-summary-v0.schema.json",
        "stage_id": STAGE5AJ_ID,
        "source_stage_id": STAGE5AJ_SOURCE_STAGE_ID,
        "xlsx_workbooks_detected": len(workbook_records),
        "xlsx_workbooks_summarized": len(workbook_records),
        "lp_excel_detected": any(record["source_id"] == "lp_excel_workbook_local" for record in workbook_records),
        "translations_decryptions_detected": any(record["source_id"] == "translations_decryptions_xlsx_local" for record in workbook_records),
        "total_sheet_count": sum(int(record["sheet_count"]) for record in workbook_records),
        "total_highlighted_cell_count": sum(int(record["highlighted_cell_count"]) for record in workbook_records),
        "total_formula_cell_count": sum(int(record["formula_cell_count"]) for record in workbook_records),
        "total_images_count": sum(int(record["images_count"]) for record in workbook_records),
        "generated_cell_metadata_path": (results_dir / STAGE5AJ_REPORTS["xlsx_cells"]).as_posix(),
        "raw_workbook_committed": False,
        "generated_outputs_committed": False,
        "ocr_performed": False,
        "solve_claim": False,
    }
    write_records(out, workbook_records, **summary)
    write_json(results_dir / STAGE5AJ_REPORTS["xlsx_index"], {**summary, "records": workbook_records})
    write_jsonl(results_dir / STAGE5AJ_REPORTS["xlsx_cells"], cell_records)
    return {**summary, "records": workbook_records}


def _summarize_workbook(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    hashed = hash_file(path)
    workbook = load_workbook(path, data_only=False)
    source_id = source_id_for_filename(path.name) or f"stage5aj_xlsx_{path.stem.lower().replace(' ', '_')}"
    sheet_names = list(workbook.sheetnames)
    cell_records: list[dict[str, Any]] = []
    comments_count = 0
    hyperlinks_count = 0
    merged_range_count = 0
    formula_cell_count = 0
    highlighted_cell_count = 0
    non_empty_cell_count = 0
    rune_like_cell_count = 0
    numeric_cell_count = 0
    images_count = 0
    highlight_color_groups: dict[str, int] = {}
    table_like_region_count = 0
    for sheet in workbook.worksheets:
        merged_range_count += len(sheet.merged_cells.ranges)
        images_count += len(getattr(sheet, "_images", []))
        if sheet.max_row > 1 and sheet.max_column > 1:
            table_like_region_count += 1
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                value = cell.value
                if value is None:
                    continue
                non_empty_cell_count += 1
                if isinstance(value, (int, float)):
                    numeric_cell_count += 1
                text = str(value)
                if any("\u16a0" <= char <= "\u16ff" for char in text):
                    rune_like_cell_count += 1
                if cell.data_type == "f" or text.startswith("="):
                    formula_cell_count += 1
                if cell.hyperlink:
                    hyperlinks_count += 1
                if cell.comment:
                    comments_count += 1
                fill = _fill_color(cell)
                if fill:
                    highlighted_cell_count += 1
                    highlight_color_groups[fill] = highlight_color_groups.get(fill, 0) + 1
                if len(cell_records) < 5000:
                    cell_records.append(_cell_record(path, source_id, sheet.title, cell, fill))
    lower_names = {name.lower() for name in sheet_names}
    record = {
        "record_type": "stage5aj_xlsx_workbook_summary_record",
        "schema": "schemas/source-harvester/xlsx-extraction-summary-v0.schema.json",
        "stage_id": STAGE5AJ_ID,
        "source_stage_id": STAGE5AJ_SOURCE_STAGE_ID,
        "workbook_id": source_id,
        "source_id": source_id,
        "path": repo_relative(path),
        "sha256": hashed["sha256"],
        "size_bytes": hashed["size_bytes"],
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "defined_names_count": len(list(workbook.defined_names)),
        "images_count": images_count,
        "comments_count": comments_count,
        "hyperlinks_count": hyperlinks_count,
        "merged_range_count": merged_range_count,
        "formula_cell_count": formula_cell_count,
        "highlighted_cell_count": highlighted_cell_count,
        "non_empty_cell_count": non_empty_cell_count,
        "rune_like_cell_count": rune_like_cell_count,
        "numeric_cell_count": numeric_cell_count,
        "table_like_region_count": table_like_region_count,
        "delimiter_key_present": any("delimiter" in name for name in lower_names),
        "lp_delimiters_sheet_present": "lp delimiters" in lower_names,
        "lp_no_delimiters_sheet_present": "lp no delimiters" in lower_names,
        "rune_stream_by_section_sheet_present": "rune stream by section" in lower_names,
        "single_string_runes_sheet_present": "single string runes" in lower_names,
        "translations_sheet_present": "translations" in lower_names,
        "gematria_primus_sheet_present": "gematria primus" in lower_names,
        "rune_counts_sheet_present": "rune counts" in lower_names,
        "highlight_color_groups": dict(sorted(highlight_color_groups.items())),
        "repeat_highlight_candidate_count": sum(count for count in highlight_color_groups.values() if count > 1),
        "section_count": sum(1 for name in lower_names if "section" in name),
        "count_policy_records": sum(1 for name in lower_names if "count" in name),
        "raw_workbook_committed": False,
        "generated_outputs_committed": False,
        "ocr_performed": False,
        "solve_claim": False,
    }
    return record, cell_records


def _fill_color(cell: Any) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return None
    color = fill.fgColor
    if color is None:
        return None
    rgb = color.rgb or color.indexed or color.theme
    if rgb in {None, "00000000", "FFFFFFFF"}:
        return None
    return str(rgb)


def _cell_record(path: Path, source_id: str, sheet_name: str, cell: Any, fill: str | None) -> dict[str, Any]:
    value = cell.value
    return {
        "record_type": "stage5aj_xlsx_cell_metadata_record",
        "stage_id": STAGE5AJ_ID,
        "source_id": source_id,
        "path": repo_relative(path),
        "sheet_name": sheet_name,
        "cell_address": cell.coordinate,
        "value_repr": "" if value is None else str(value)[:200],
        "data_type": cell.data_type,
        "formula": bool(cell.data_type == "f" or str(value).startswith("=")),
        "number_format": cell.number_format,
        "fill_color": fill,
        "font_color": str(cell.font.color.rgb) if cell.font and cell.font.color and cell.font.color.type == "rgb" else None,
        "bold": bool(cell.font.bold) if cell.font else False,
        "italic": bool(cell.font.italic) if cell.font else False,
        "hyperlink": bool(cell.hyperlink),
        "comment_present": bool(cell.comment),
        "merged_cell": False,
        "raw_workbook_committed": False,
        "solve_claim": False,
    }
