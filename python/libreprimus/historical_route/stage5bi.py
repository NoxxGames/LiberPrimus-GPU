"""Stage 5BI Fandom source-lock triage and archive crosswalk metadata."""

from __future__ import annotations

import hashlib
import json
import subprocess
from collections import Counter
from pathlib import Path
from typing import Any

import yaml

from libreprimus.paths import repo_root

STAGE_ID = "stage-5bi"
STAGE_TITLE = "Stage 5BI - Fandom source-lock triage and original-archive crosswalk integration"
LOCAL_ARCHIVE = Path("third_party/CicadaSolversIddqd")
LOCAL_SPREADSHEET = Path("third_party/3N_3p_Bases_49-51.jpg.xlsx")
UPSTREAM_ARCHIVE_URL = "https://github.com/cicada-solvers/The-Complete-Cicada3301-Archive"
SOURCE_DEEP_RESEARCH_REPORT = (
    "08_LiberPrimus-GPU-Stage-5BH-DR-Fandom-Source-Lock-Triage-And-Archive-Crosswalk.md"
)

DATA_PATHS = {
    "page_triage": Path("data/historical-route/stage5bi-fandom-page-triage.yaml"),
    "item_candidates": Path("data/historical-route/stage5bi-fandom-item-source-lock-candidates.yaml"),
    "archive_crosswalk": Path("data/historical-route/stage5bi-original-archive-crosswalk-candidates.yaml"),
    "media_policy": Path("data/historical-route/stage5bi-fandom-media-non-original-policy.yaml"),
    "surface_context": Path("data/historical-route/stage5bi-2014-256-byte-surface-context.yaml"),
    "negative_controls": Path("data/historical-route/stage5bi-negative-control-quarantine.yaml"),
    "source_gaps": Path("data/historical-route/stage5bi-source-gap-register.yaml"),
    "guardrail": Path("data/historical-route/stage5bi-guardrail.yaml"),
    "token_block_context": Path("data/token-block/stage5bi-token-block-external-context.yaml"),
    "surface_token_block_context": Path("data/token-block/stage5bi-2014-surface-token-block-context.yaml"),
    "spreadsheet_reconciliation": Path("data/token-block/stage5bi-spreadsheet-stage5aw-reconciliation.yaml"),
    "spreadsheet_source_lock": Path("data/source-harvester/stage5bi-local-spreadsheet-source-lock.yaml"),
    "crosswalk_summary": Path("data/source-harvester/stage5bi-fandom-crosswalk-source-summary.yaml"),
    "summary": Path("data/project-state/stage5bi-summary.yaml"),
    "next_stage": Path("data/project-state/stage5bi-next-stage-decision.yaml"),
}

FANDOM_BASE = "https://uncovering-cicada.fandom.com/wiki/"


def _root() -> Path:
    return repo_root()


def _resolve(path: Path | str) -> Path:
    value = Path(path)
    if value.is_absolute():
        return value
    return _root() / value


def _repo_relative(path: Path) -> str:
    try:
        return path.resolve().relative_to(_root().resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def _write_yaml(path: Path | str, payload: dict[str, Any]) -> None:
    resolved = _resolve(path)
    resolved.parent.mkdir(parents=True, exist_ok=True)
    resolved.write_text(yaml.safe_dump(payload, sort_keys=False, allow_unicode=False), encoding="utf-8")


def _read_yaml(path: Path | str) -> dict[str, Any]:
    resolved = _resolve(path)
    if not resolved.is_file():
        return {}
    return yaml.safe_load(resolved.read_text(encoding="utf-8")) or {}


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _schema(path: str) -> str:
    return path


def _git_staged_paths() -> list[str]:
    try:
        result = subprocess.run(
            ["git", "diff", "--cached", "--name-only"],
            cwd=_root(),
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
        )
    except (OSError, subprocess.TimeoutExpired):
        return []
    if result.returncode != 0:
        return []
    return [line.strip().replace("\\", "/") for line in result.stdout.splitlines() if line.strip()]


def _base_record(record_type: str, schema: str) -> dict[str, Any]:
    return {
        "record_type": record_type,
        "schema": schema,
        "stage_id": STAGE_ID,
        "stage_title": STAGE_TITLE,
        "source_deep_research_stage": "stage-5bh-dr",
        "source_deep_research_report": SOURCE_DEEP_RESEARCH_REPORT,
        "source_previous_stage": "stage-5bf",
        "source_previous_deep_research_review": "stage-5bg",
    }


def _page(
    triage_id: str,
    title: str,
    slug: str,
    classification: str,
    tier: str,
    relevance: str,
    action: str,
    priority: str,
    rationale: str,
) -> dict[str, Any]:
    return {
        "triage_id": triage_id,
        "source_title": title,
        "source_url": FANDOM_BASE + slug,
        "source_lock_classification": classification,
        "source_tier": tier,
        "token_block_relevance": relevance,
        "recommended_action": action,
        "source_lock_priority": priority,
        "evidence_role": "route_map_or_review_context",
        "rationale": rationale,
        "fandom_page_body_committed": False,
        "fandom_images_committed": False,
        "execution_allowed": False,
        "solve_claim": False,
    }


def build_fandom_page_triage() -> dict[str, Any]:
    records = [
        _page(
            "stage5bi-page-what-happened-2014",
            "What Happened Part 1 (2014)",
            "What_Happened_Part_1_(2014)",
            "source_lock_after_original_archive_crosswalk",
            "tier2_archived_community_page_with_references",
            "token_block_constraint",
            "source_lock_original_artifact_in_archive",
            "highest",
            "Stage 5BH flags this page as a useful 2014 route map, but Stage 4K did not hash-lock the page body.",
        ),
        _page(
            "stage5bi-page-page-49-51",
            "Page 49-51",
            "Page_49-51",
            "source_lock_after_original_archive_crosswalk",
            "tier2_archived_community_page_with_references",
            "direct_token_block_relevance",
            "source_lock_original_artifact_in_archive",
            "highest",
            "Fandom page is context for the 256-position token surface, not source truth for token identity.",
        ),
        _page(
            "stage5bi-page-liber-primus-49-51-base-pack",
            "Liber Primus pp49-51 data interpreted as base 59, 60, 62 and 64",
            "Liber_Primus_pp49-51_data_interpreted_as_base_59,_60,_62_and_64.",
            "secondary_community_analysis",
            "tier3_reproducible_community_data",
            "direct_token_block_relevance",
            "source_lock_individually_now",
            "highest",
            "Community base-analysis context should be reconciled with Stage 5AW branch metadata, not executed.",
        ),
        _page(
            "stage5bi-page-what-happened-liber-primus",
            "What Happened Liber Primus (Post 2014)",
            "What_Happened_Liber_Primus_(Post_2014)",
            "reference_only_route_context",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "record_reference_only",
            "medium",
            "Useful post-2014 route context, but not an experiment input.",
        ),
        _page(
            "stage5bi-page-liber-primus",
            "Liber Primus",
            "Liber_Primus",
            "reference_only_route_context",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "record_reference_only",
            "medium",
            "Index-level Liber Primus context only.",
        ),
        _page(
            "stage5bi-page-solved-pages",
            "How the solved pages of the Liber Primus were solved",
            "How_the_solved_pages_of_the_Liber_Primus_were_solved",
            "secondary_community_analysis",
            "tier3_reproducible_community_data",
            "positive_control_only",
            "record_reference_only",
            "medium",
            "Can inform solved-fixture context, but not unsolved-page truth.",
        ),
        _page(
            "stage5bi-page-unsolved-pages",
            "Liber Primus Unsolved Pages",
            "Liber_Primus_Unsolved_Pages",
            "reference_only_route_context",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "record_reference_only",
            "medium",
            "Route context only; unsolved-page execution remains blocked.",
        ),
        _page(
            "stage5bi-page-frequency-analysis-unsolved",
            "Frequency Analysis Unsolved Pages",
            "Frequency_Analysis_Unsolved_Pages",
            "secondary_community_analysis",
            "tier3_reproducible_community_data",
            "historical_context_only",
            "record_reference_only",
            "medium",
            "Reference for community analysis; no scoring or experiment execution is authorized.",
        ),
        _page(
            "stage5bi-page-page-56",
            "PAGE 56",
            "PAGE_56",
            "secondary_community_analysis",
            "tier3_reproducible_community_data",
            "historical_context_only",
            "defer_until_archive_crosswalk",
            "medium",
            "p56 context remains blocked unless exact token buffers are separately source-locked.",
        ),
        _page(
            "stage5bi-page-page-57",
            "PAGE 57",
            "PAGE_57",
            "secondary_community_analysis",
            "tier3_reproducible_community_data",
            "historical_context_only",
            "defer_until_archive_crosswalk",
            "medium",
            "p57 route context remains reference-only.",
        ),
        _page(
            "stage5bi-page-liber-primus-keys",
            "Liber Primus keys",
            "Liber_Primus_keys",
            "source_lock_page_body",
            "tier3_reproducible_community_data",
            "positive_control_only",
            "source_lock_individually_now",
            "medium",
            "Exact key claims would need item-level source locks before verifier use.",
        ),
        _page(
            "stage5bi-page-onion7-numbers-page15",
            "Onion 7: numbers on page 15",
            "Onion_7:_numbers_on_page_15",
            "source_lock_page_body",
            "tier3_reproducible_community_data",
            "historical_context_only",
            "source_lock_individually_now",
            "medium",
            "Known bounded numeric context; no rerun is selected.",
        ),
        _page(
            "stage5bi-page-runes-transcriptions",
            "Runes Transcriptions",
            "Runes_Transcriptions",
            "source_lock_after_original_archive_crosswalk",
            "tier2_archived_community_page_with_references",
            "blocks_execution_until_source_locked",
            "defer_until_archive_crosswalk",
            "medium",
            "Transcript-like material must not override canonical corpus inactivity.",
        ),
        _page(
            "stage5bi-page-mentioned-primes",
            "Mentioned PRIMES",
            "Mentioned_PRIMES",
            "secondary_community_analysis",
            "tier3_reproducible_community_data",
            "historical_context_only",
            "record_reference_only",
            "low",
            "Prime references are context only unless exact source-backed transforms are separately scoped.",
        ),
        _page(
            "stage5bi-page-gematria-primus",
            "Gematria Primus",
            "Gematria_Primus",
            "reference_only_route_context",
            "tier3_reproducible_community_data",
            "historical_context_only",
            "record_reference_only",
            "low",
            "Reference to mappings, not a source of canonical token truth.",
        ),
        _page(
            "stage5bi-page-what-happened-2012",
            "What Happened (2012)",
            "What_Happened_(2012)",
            "reference_only_route_context",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "record_reference_only",
            "medium",
            "Route context for original archive crosswalk.",
        ),
        _page(
            "stage5bi-page-what-happened-2013",
            "What Happened Part 1 (2013)",
            "What_Happened_Part_1_(2013)",
            "reference_only_route_context",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "record_reference_only",
            "medium",
            "Route context for telnet, image, and signed-message artifacts.",
        ),
        _page(
            "stage5bi-page-telnet-prompt",
            "The Telnet Prompt",
            "The_Telnet_Prompt",
            "negative_control_or_false_path",
            "tier5_speculative_or_unverified",
            "historical_context_only",
            "record_negative_control",
            "high",
            "Community recreation must not be treated as exact original service behavior.",
        ),
        _page(
            "stage5bi-page-hidden-content-original-image-2013",
            "Hidden content of original image (January 4th 2013)",
            "Hidden_content_of_original_image_(January_4th_2013)",
            "source_lock_after_original_archive_crosswalk",
            "tier2_archived_community_page_with_references",
            "positive_control_only",
            "source_lock_original_artifact_in_archive",
            "high",
            "Potential stego-positive-control context only; no tool execution.",
        ),
        _page(
            "stage5bi-page-interconnectedness",
            "Interconnectedness",
            "Interconnectedness",
            "source_lock_after_original_archive_crosswalk",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "source_lock_original_artifact_in_archive",
            "high",
            "Fandom route pointer should crosswalk to local archive MP3 and page context.",
        ),
        _page(
            "stage5bi-page-instar-emergence",
            "Instar emergence (mp3 and hidden poem)",
            "Instar_emergence_(mp3_and_hidden_poem)",
            "source_lock_after_original_archive_crosswalk",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "source_lock_original_artifact_in_archive",
            "high",
            "Audio/stego context remains blocked pending expected-output policy.",
        ),
        _page(
            "stage5bi-page-pgp-live-cd",
            "PGP Signature from live cd",
            "PGP_Signature_from_live_cd",
            "negative_control_or_false_path",
            "tier5_speculative_or_unverified",
            "historical_context_only",
            "record_negative_control",
            "high",
            "Weak-provenance PGP candidate; no network key fetch or truth verification.",
        ),
        _page(
            "stage5bi-page-outguess",
            "OutGuess",
            "OutGuess",
            "tool_reference",
            "not_source_evidence",
            "positive_control_only",
            "record_reference_only",
            "medium",
            "Tool reference only; execution remains blocked outside explicit fixture stages.",
        ),
        _page(
            "stage5bi-page-outguessed-rune-table",
            "OutGuessed message from rune table",
            "OutGuessed_message_from_rune_table",
            "source_lock_after_original_archive_crosswalk",
            "tier1a_original_archive_crosswalk_candidate",
            "positive_control_only",
            "source_lock_original_artifact_in_archive",
            "high",
            "Expected-output source context should prefer local archive signed-message record.",
        ),
        _page(
            "stage5bi-page-2014-data-dump",
            "2014 CICADA 3301 DATA DUMP",
            "2014_CICADA_3301_DATA_DUMP",
            "source_lock_after_original_archive_crosswalk",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "defer_until_archive_crosswalk",
            "high",
            "Potential route-map index for 2014 artifacts, but not source truth.",
        ),
        _page(
            "stage5bi-page-2014-things-xored",
            "2014 THINGS THAT HAD BEEN XORED",
            "2014_THINGS_THAT_HAD_BEEN_XORED",
            "secondary_community_analysis",
            "tier5_speculative_or_unverified",
            "blocks_execution_until_source_locked",
            "record_source_gap",
            "high",
            "XOR-related community context is not an execution manifest.",
        ),
        _page(
            "stage5bi-page-main-menu-2014",
            "MAIN MENU OF PAGES FOR CICADA 2014",
            "MAIN_MENU_OF_PAGES_FOR_CICADA_2014",
            "reference_only_route_context",
            "tier2_archived_community_page_with_references",
            "historical_context_only",
            "record_reference_only",
            "medium",
            "Route index context only.",
        ),
        _page(
            "stage5bi-page-the-infamous-warning",
            "The Infamous Warning",
            "The_Infamous_Warning",
            "negative_control_or_false_path",
            "not_source_evidence",
            "no_token_block_relevance",
            "record_negative_control",
            "high",
            "Unconfirmed claims are quarantined as false-path controls.",
        ),
        _page(
            "stage5bi-page-fake-puzzles",
            "Fake puzzles",
            "Fake_puzzles",
            "negative_control_or_false_path",
            "not_source_evidence",
            "no_token_block_relevance",
            "record_negative_control",
            "high",
            "Fake or unverifiable puzzle material belongs in negative controls.",
        ),
        _page(
            "stage5bi-page-useful-links",
            "USEFUL LINKS",
            "USEFUL_LINKS",
            "reference_only_route_context",
            "not_source_evidence",
            "historical_context_only",
            "record_reference_only",
            "low",
            "Index page only.",
        ),
    ]
    return {
        **_base_record(
            "stage5bi_fandom_page_triage_record_set",
            _schema("schemas/historical-route/stage5bi-fandom-page-triage-v0.schema.json"),
        ),
        "page_count": len(records),
        "fandom_pages_are_secondary_route_context": True,
        "fandom_page_bodies_committed": False,
        "raw_html_committed": False,
        "records": records,
    }


def _candidate(
    candidate_id: str,
    title: str,
    source_url: str,
    candidate_kind: str,
    relevance: str,
    action: str,
    highest_priority: bool,
    crosswalk_required: bool = True,
    notes: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "candidate_id": candidate_id,
        "source_title": title,
        "source_url": source_url,
        "candidate_kind": candidate_kind,
        "source_lock_classification": "source_lock_after_original_archive_crosswalk"
        if crosswalk_required
        else "reference_only_route_context",
        "source_tier": "tier1a_original_archive_crosswalk_candidate"
        if crosswalk_required
        else "tier2_archived_community_page_with_references",
        "token_block_relevance": relevance,
        "recommended_action": action,
        "highest_priority_candidate": highest_priority,
        "original_archive_crosswalk_required": crosswalk_required,
        "execution_allowed": False,
        "raw_body_committed": False,
        "generated_output_committed": False,
        "solve_claim": False,
        "notes": notes or [],
    }


def build_item_source_lock_candidates() -> dict[str, Any]:
    records = [
        _candidate(
            "stage5bi-c01-2014-growing-hex-surface",
            "2014 growing string 256-byte surface",
            FANDOM_BASE + "What_Happened_Part_1_(2014)",
            "2014_256_byte_hex_surface",
            "token_block_constraint",
            "source_lock_original_artifact_in_archive",
            True,
            notes=["Context only; do not combine with the page 49-51 token surface in Stage 5BI."],
        ),
        _candidate(
            "stage5bi-c02-2014-1033-hex-surface",
            "2014 <!--1033--> 256-byte surface",
            FANDOM_BASE + "What_Happened_Part_1_(2014)",
            "2014_256_byte_hex_surface",
            "token_block_constraint",
            "source_lock_original_artifact_in_archive",
            True,
        ),
        _candidate(
            "stage5bi-c03-2014-3301-hex-surface",
            "2014 <!--3301--> 256-byte surface",
            FANDOM_BASE + "What_Happened_Part_1_(2014)",
            "2014_256_byte_hex_surface",
            "token_block_constraint",
            "source_lock_original_artifact_in_archive",
            True,
        ),
        _candidate(
            "stage5bi-c04-page49-51-256-token-surface",
            "Page 49-51 256-position token surface",
            FANDOM_BASE + "Page_49-51",
            "page49_51_256_token_surface",
            "direct_token_block_relevance",
            "source_lock_original_artifact_in_archive",
            True,
        ),
        _candidate(
            "stage5bi-c05-page49-51-base59-60-62-64-analysis",
            "Page 49-51 base 59/60/62/64 community analysis",
            FANDOM_BASE + "Liber_Primus_pp49-51_data_interpreted_as_base_59,_60,_62_and_64.",
            "community_token_block_analysis",
            "direct_token_block_relevance",
            "source_lock_individually_now",
            True,
        ),
        _candidate(
            "stage5bi-c06-boards-page49-51-256-byte-thread",
            "Boards.net page 49-51 256 byte strings thread",
            "https://cicada3301.boards.net/thread/41/pages-49-256-byte-strings",
            "community_context_thread",
            "blocks_execution_until_source_locked",
            "record_source_gap",
            True,
        ),
        _candidate(
            "stage5bi-c07-interconnectedness-mp3",
            "2014 Interconnectedness MP3",
            FANDOM_BASE + "Interconnectedness",
            "mp3_instar_or_interconnectedness_candidate",
            "historical_context_only",
            "source_lock_original_artifact_in_archive",
            True,
        ),
        _candidate(
            "stage5bi-c08-instar-emergence-audio-hidden-poem",
            "Instar emergence MP3 and hidden poem",
            FANDOM_BASE + "Instar_emergence_(mp3_and_hidden_poem)",
            "mp3_instar_or_interconnectedness_candidate",
            "historical_context_only",
            "source_lock_expected_output_pair",
            True,
        ),
        _candidate(
            "stage5bi-c09-761-mp3",
            "761.MP3 audio candidate",
            "https://github.com/cicada-solvers/The-Complete-Cicada3301-Archive",
            "mp3_instar_candidate",
            "historical_context_only",
            "source_lock_original_artifact_in_archive",
            False,
        ),
        _candidate(
            "stage5bi-c10-outguessed-rune-table-message",
            "OutGuessed message from rune table",
            FANDOM_BASE + "OutGuessed_message_from_rune_table",
            "outguess_expected_output_candidate",
            "positive_control_only",
            "source_lock_expected_output_pair",
            True,
        ),
        _candidate(
            "stage5bi-c11-4gq25-image",
            "4gq25.jpg image fixture candidate",
            FANDOM_BASE + "2016_Puzzle",
            "image_fixture_candidate",
            "historical_context_only",
            "source_lock_original_artifact_in_archive",
            True,
        ),
        _candidate(
            "stage5bi-c12-magicsquares-text",
            "2014 magicsquares.txt",
            "https://github.com/cicada-solvers/The-Complete-Cicada3301-Archive/blob/main/assets/2014/stage07/magicsquares.txt",
            "magic_square_artifact",
            "historical_context_only",
            "source_lock_original_artifact_in_archive",
            False,
        ),
        _candidate(
            "stage5bi-c13-pgp-signature-live-cd",
            "PGP Signature from live cd",
            FANDOM_BASE + "PGP_Signature_from_live_cd",
            "pgp_weak_or_quarantined_candidate",
            "historical_context_only",
            "record_negative_control",
            True,
        ),
        _candidate(
            "stage5bi-c14-message-from-first-onion",
            "Message from first onion",
            FANDOM_BASE + "Message_from_first_onion",
            "onion_route_message_candidate",
            "historical_context_only",
            "source_lock_original_artifact_in_archive",
            False,
        ),
        _candidate(
            "stage5bi-c15-telnet-output-recreation",
            "2013 telnet output/recreation",
            FANDOM_BASE + "The_Telnet_Prompt",
            "recreation_not_reference_implementation",
            "historical_context_only",
            "record_negative_control",
            True,
        ),
        _candidate(
            "stage5bi-c16-liber-primus-keys-frequency",
            "Liber Primus keys and frequency pages",
            FANDOM_BASE + "Liber_Primus_keys",
            "liber_primus_reference_candidate",
            "positive_control_only",
            "source_lock_individually_now",
            False,
        ),
        _candidate(
            "stage5bi-c17-p56-p57-route-context",
            "Page 56/Page 57 route context",
            FANDOM_BASE + "PAGE_56",
            "liber_primus_page_context",
            "historical_context_only",
            "defer_until_archive_crosswalk",
            False,
        ),
        _candidate(
            "stage5bi-c18-fandom-what-happened-2014-page-body",
            "What Happened Part 1 (2014) page body",
            FANDOM_BASE + "What_Happened_Part_1_(2014)",
            "fandom_page_body_snapshot_candidate",
            "blocks_execution_until_source_locked",
            "source_lock_page_body",
            True,
            crosswalk_required=False,
            notes=["Stage 4K fetch failed with 403; metadata planning continues without raw page body."],
        ),
    ]
    return {
        **_base_record(
            "stage5bi_fandom_item_source_lock_candidate_set",
            _schema("schemas/historical-route/stage5bi-fandom-item-source-lock-candidate-v0.schema.json"),
        ),
        "candidate_count": len(records),
        "highest_priority_candidate_count": sum(1 for record in records if record["highest_priority_candidate"]),
        "execution_allowed": False,
        "solve_claim": False,
        "records": records,
    }


def _load_stage5bf_artifacts() -> list[dict[str, Any]]:
    index = _read_yaml(Path("data/historical-route/stage5bf-high-priority-artifact-index.yaml"))
    return list(index.get("artifacts", []))


def _find_artifact(artifacts: list[dict[str, Any]], all_terms: list[str], preferred_terms: list[str] | None = None) -> dict[str, Any] | None:
    preferred_terms = preferred_terms or []
    matches: list[dict[str, Any]] = []
    for artifact in artifacts:
        rel = str(artifact.get("relative_path", "")).lower()
        if all(term.lower() in rel for term in all_terms):
            matches.append(artifact)
    if preferred_terms:
        for artifact in matches:
            rel = str(artifact.get("relative_path", "")).lower()
            if all(term.lower() in rel for term in preferred_terms):
                return artifact
    return matches[0] if matches else None


def _crosswalk(
    candidate_id: str,
    candidate_label: str,
    artifact: dict[str, Any] | None,
    status_when_missing: str = "pending_manual_review",
    status_when_found: str = "original_archive_equivalent_found",
    notes: list[str] | None = None,
) -> dict[str, Any]:
    status = status_when_found if artifact else status_when_missing
    return {
        "crosswalk_id": f"{candidate_id}-archive-crosswalk",
        "candidate_id": candidate_id,
        "candidate_label": candidate_label,
        "preferred_archive_root": LOCAL_ARCHIVE.as_posix(),
        "upstream_archive_url": UPSTREAM_ARCHIVE_URL,
        "archive_crosswalk_status": status,
        "archive_relative_path": artifact.get("relative_path") if artifact else None,
        "archive_sha256": artifact.get("sha256") if artifact else None,
        "archive_size_bytes": artifact.get("size_bytes") if artifact else None,
        "archive_file_type_guess": artifact.get("file_type_guess") if artifact else None,
        "source_stage5bf_artifact_id": artifact.get("artifact_id") if artifact else None,
        "crosswalk_verified_by": "stage5bf_metadata_hash_path_match" if artifact else "not_verified",
        "raw_archive_files_committed": False,
        "execution_allowed": False,
        "solve_claim": False,
        "notes": notes or [],
    }


def build_original_archive_crosswalk() -> dict[str, Any]:
    artifacts = _load_stage5bf_artifacts()
    rows = [
        _crosswalk(
            "stage5bi-c01-2014-growing-hex-surface",
            "2014 growing string 256-byte surface",
            _find_artifact(artifacts, ["2014", "growing"], ["growing string"]),
            status_when_found="probable_archive_path_candidate",
            notes=["The archive metadata contains growing-string candidates; exact 512-hex extraction remains future review."],
        ),
        _crosswalk(
            "stage5bi-c02-2014-1033-hex-surface",
            "2014 <!--1033--> 256-byte surface",
            _find_artifact(artifacts, ["1033"]),
            status_when_found="probable_archive_path_candidate",
            notes=["Archive has 1033 image/outguess candidates; exact hex-surface source remains unresolved."],
        ),
        _crosswalk(
            "stage5bi-c03-2014-3301-hex-surface",
            "2014 <!--3301--> 256-byte surface",
            _find_artifact(artifacts, ["2014", "3301"]),
            status_when_missing="probable_archive_path_candidate",
            status_when_found="probable_archive_path_candidate",
            notes=["3301-labelled 2014 paths exist but need manual verification against exact surface claims."],
        ),
        _crosswalk(
            "stage5bi-c04-page49-51-256-token-surface",
            "Page 49-51 256-position token surface",
            _find_artifact(artifacts, ["assets/2014/liber-primus-complete/49.jpg"]),
            notes=["Stage 5AP/5AR already source-lock the three original-image page anchors; this row is context only."],
        ),
        _crosswalk(
            "stage5bi-c07-interconnectedness-mp3",
            "2014 Interconnectedness MP3",
            _find_artifact(artifacts, ["interconnectedness", ".mp3"]),
        ),
        _crosswalk(
            "stage5bi-c09-761-mp3",
            "761.MP3 audio candidate",
            _find_artifact(artifacts, ["761.mp3"]),
        ),
        _crosswalk(
            "stage5bi-c10-outguessed-rune-table-message",
            "OutGuessed message from rune table",
            _find_artifact(artifacts, ["outguessed message from rune table"]),
        ),
        _crosswalk(
            "stage5bi-c11-4gq25-image",
            "4gq25.jpg image fixture candidate",
            _find_artifact(artifacts, ["4gq25.jpg"], ["assets/2016"]),
        ),
        _crosswalk(
            "stage5bi-c12-magicsquares-text",
            "2014 magicsquares.txt",
            _find_artifact(artifacts, ["magicsquares.txt"]),
        ),
        _crosswalk(
            "stage5bi-c15-telnet-output-recreation",
            "2013 telnet output/recreation",
            _find_artifact(artifacts, ["telnet"]),
            notes=["Archive presence does not make the community Telnet prompt page a reference implementation."],
        ),
        _crosswalk(
            "stage5bi-c13-pgp-signature-live-cd",
            "PGP Signature from live cd",
            _find_artifact(artifacts, ["live cd", "pgp"]),
            status_when_missing="local_archive_search_no_match",
            notes=["Fandom page remains weak/quarantined until stronger original material is found."],
        ),
        _crosswalk(
            "stage5bi-c06-boards-page49-51-256-byte-thread",
            "Boards.net page 49-51 256 byte strings thread",
            None,
            status_when_missing="blocked_archive_absent",
            notes=["Thread may be inaccessible; future Wayback/local archive source-lock required."],
        ),
    ]
    counter = Counter(row["archive_crosswalk_status"] for row in rows)
    return {
        **_base_record(
            "stage5bi_original_archive_crosswalk_candidate_set",
            _schema("schemas/historical-route/stage5bi-original-archive-crosswalk-candidate-v0.schema.json"),
        ),
        "archive_available": _resolve(LOCAL_ARCHIVE).is_dir(),
        "candidate_count": len(rows),
        "verified_archive_crosswalk_count": counter["original_archive_equivalent_found"],
        "probable_archive_path_candidate_count": counter["probable_archive_path_candidate"],
        "raw_archive_files_committed": False,
        "execution_allowed": False,
        "solve_claim": False,
        "records": rows,
    }


def build_media_policy() -> dict[str, Any]:
    return {
        **_base_record(
            "stage5bi_fandom_media_non_original_policy",
            _schema("schemas/historical-route/stage5bi-fandom-media-non-original-policy-v0.schema.json"),
        ),
        "default_media_status": "fandom_copy_reference_only",
        "fandom_images_are_original_source_truth": False,
        "screenshots_are_original_source_truth": False,
        "fandom_image_urls_can_be_route_pointers": True,
        "original_archive_preference": [
            "third_party/CicadaSolversIddqd",
            "Stage 5BF archive/hash metadata",
            "commit-addressed public repository references where applicable",
        ],
        "policy_rules": [
            "Fandom-hosted images are secondary copies by default.",
            "Do not use Fandom image URLs as original source truth.",
            "Crosswalk high-value media to local archive originals or hash-locked equivalents.",
            "Screenshots, resized images, and hosted derivatives remain reference-only.",
            "Original archive equivalence must be recorded before any future positive-control or source-truth use.",
        ],
        "media_status_values": [
            "fandom_copy_reference_only",
            "needs_archive_original_crosswalk",
            "original_archive_equivalent_found",
            "do_not_use_as_source_truth",
        ],
        "raw_images_committed": False,
        "solve_claim": False,
    }


def build_2014_surface_context(crosswalk: dict[str, Any]) -> dict[str, Any]:
    by_candidate = {record["candidate_id"]: record for record in crosswalk.get("records", [])}
    surfaces = [
        {
            "surface_id": "stage5bi-c01-2014-growing-hex-surface",
            "surface_label": "2014 growing string",
            "historical_marker": "Patience is a virtue",
            "claimed_hex_character_count": 512,
            "claimed_byte_count": 256,
            "source_url": FANDOM_BASE + "What_Happened_Part_1_(2014)",
            "archive_crosswalk_status": by_candidate.get("stage5bi-c01-2014-growing-hex-surface", {}).get(
                "archive_crosswalk_status", "pending_manual_review"
            ),
            "execution_allowed": False,
            "combination_with_page49_51_allowed": False,
            "raw_surface_body_committed": False,
            "solve_claim": False,
        },
        {
            "surface_id": "stage5bi-c02-2014-1033-hex-surface",
            "surface_label": "<!--1033-->",
            "historical_marker": "1033",
            "claimed_hex_character_count": 512,
            "claimed_byte_count": 256,
            "source_url": FANDOM_BASE + "What_Happened_Part_1_(2014)",
            "archive_crosswalk_status": by_candidate.get("stage5bi-c02-2014-1033-hex-surface", {}).get(
                "archive_crosswalk_status", "pending_manual_review"
            ),
            "execution_allowed": False,
            "combination_with_page49_51_allowed": False,
            "raw_surface_body_committed": False,
            "solve_claim": False,
        },
        {
            "surface_id": "stage5bi-c03-2014-3301-hex-surface",
            "surface_label": "<!--3301-->",
            "historical_marker": "3301",
            "claimed_hex_character_count": 512,
            "claimed_byte_count": 256,
            "source_url": FANDOM_BASE + "What_Happened_Part_1_(2014)",
            "archive_crosswalk_status": by_candidate.get("stage5bi-c03-2014-3301-hex-surface", {}).get(
                "archive_crosswalk_status", "pending_manual_review"
            ),
            "execution_allowed": False,
            "combination_with_page49_51_allowed": False,
            "raw_surface_body_committed": False,
            "solve_claim": False,
        },
    ]
    return {
        **_base_record(
            "stage5bi_2014_256_byte_surface_context",
            _schema("schemas/historical-route/stage5bi-2014-256-byte-surface-context-v0.schema.json"),
        ),
        "surface_count": len(surfaces),
        "surfaces": surfaces,
        "page49_51_context": {
            "surface_id": "stage5bi-c04-page49-51-256-token-surface",
            "surface_label": "page 49-51 256-position token surface",
            "token_count": 256,
            "source_records": [
                "data/token-block/stage5ap-token-block-canonical-transcription.yaml",
                "data/token-block/stage5ar-token-case-policy.yaml",
                "data/token-block/stage5aw-repaired-token-variant-branch-manifest.yaml",
            ],
            "fandom_context_url": FANDOM_BASE + "Page_49-51",
            "canonical_transcription_changed": False,
            "active_token_block_manifest_changed": False,
            "execution_allowed": False,
            "combination_with_2014_surfaces_allowed": False,
            "solve_claim": False,
        },
        "surface_combination_performed": False,
        "xor_attempt_performed": False,
        "transposition_attempt_performed": False,
        "decode_attempt_performed": False,
        "hash_search_performed": False,
        "solve_claim": False,
    }


def build_negative_controls() -> dict[str, Any]:
    records = [
        {
            "negative_control_id": "stage5bi-infamous-warning",
            "source_page": FANDOM_BASE + "The_Infamous_Warning",
            "classification": "negative_control_or_false_path",
            "reason": [
                "Deep Research report says page documents unconfirmed claims.",
                "Do not treat as positive source evidence.",
            ],
            "planning_effect": ["false_path_warning_should_be_explicit", "pgp_authenticity_gate_required"],
            "execution_allowed": False,
            "solve_claim": False,
        },
        {
            "negative_control_id": "stage5bi-fake-puzzles",
            "source_page": FANDOM_BASE + "Fake_puzzles",
            "classification": "negative_control_or_false_path",
            "reason": ["Page documents fake or unverifiable puzzles."],
            "planning_effect": ["future_false_positive_filter"],
            "execution_allowed": False,
            "solve_claim": False,
        },
        {
            "negative_control_id": "stage5bi-telnet-recreation",
            "source_page": FANDOM_BASE + "The_Telnet_Prompt",
            "classification": "recreation_not_reference_implementation",
            "reason": ["Community recreation should not be treated as exact original service behavior."],
            "planning_effect": ["requires_original_service_or_archive_crosswalk"],
            "execution_allowed": False,
            "solve_claim": False,
        },
        {
            "negative_control_id": "stage5bi-pgp-signature-live-cd-deletion-candidate",
            "source_page": FANDOM_BASE + "PGP_Signature_from_live_cd",
            "classification": "weak_or_quarantined_source",
            "reason": ["Report flags candidate-for-deletion / weak provenance."],
            "planning_effect": ["pgp_truth_requires_stronger_source_lock"],
            "execution_allowed": False,
            "solve_claim": False,
        },
    ]
    return {
        **_base_record(
            "stage5bi_negative_control_quarantine_set",
            _schema("schemas/historical-route/stage5bi-negative-control-quarantine-v0.schema.json"),
        ),
        "negative_control_count": len(records),
        "records": records,
    }


def build_source_gaps() -> dict[str, Any]:
    gaps = [
        {
            "gap_id": "stage5bi-fandom-2014-page-body-not-hash-locked",
            "description": "Stage 4K attempted Fandom 2014 page snapshot but fetch failed 403; Stage 5AL source card is metadata only.",
            "affected_source": FANDOM_BASE + "What_Happened_Part_1_(2014)",
            "blocks_execution": True,
            "blocks_metadata_planning": False,
            "recommended_resolution": [
                "Prefer original archive crosswalk under third_party/CicadaSolversIddqd.",
                "Record page body snapshot only if future permitted fetch/manual source pack exists.",
            ],
        },
        {
            "gap_id": "stage5bi-boards-page49-51-thread-inaccessible",
            "affected_source": "https://cicada3301.boards.net/thread/41/pages-49-256-byte-strings",
            "blocks_execution": True,
            "blocks_metadata_planning": False,
            "recommended_resolution": [
                "Search local archive or Wayback in future source-lock stage.",
                "Do not use community statement as execution evidence.",
            ],
        },
        {
            "gap_id": "stage5bi-2014-256-surfaces-need-original-archive-crosswalk",
            "affected_items": [
                "stage5bi-c01-2014-growing-hex-surface",
                "stage5bi-c02-2014-1033-hex-surface",
                "stage5bi-c03-2014-3301-hex-surface",
            ],
            "blocks_execution": True,
            "blocks_metadata_planning": False,
            "recommended_resolution": ["Verify exact 512-character hex surfaces against original/archive-equivalent sources."],
        },
        {
            "gap_id": "stage5bi-fandom-media-original-equivalence-not-assumed",
            "description": "Fandom-hosted media are secondary copies by default.",
            "blocks_execution": True,
            "blocks_metadata_planning": False,
            "recommended_resolution": ["Crosswalk high-value media to local archive originals or commit-addressed source metadata."],
        },
        {
            "gap_id": "stage5bi-spreadsheet-not-canonical",
            "affected_source": LOCAL_SPREADSHEET.as_posix(),
            "blocks_execution": True,
            "blocks_metadata_planning": False,
            "recommended_resolution": ["Treat spreadsheet as local analysis metadata only; future token options require explicit review provenance."],
        },
    ]
    return {
        **_base_record(
            "stage5bi_source_gap_register",
            _schema("schemas/historical-route/stage5bi-source-gap-register-v0.schema.json"),
        ),
        "source_gap_count": len(gaps),
        "gaps": gaps,
        "execution_allowed": False,
        "solve_claim": False,
    }


def build_guardrail() -> dict[str, Any]:
    return {
        **_base_record("stage5bi_guardrail", _schema("schemas/historical-route/stage5bi-guardrail-v0.schema.json")),
        "metadata_only": True,
        "historical_source_lock_triage_only": True,
        "fandom_page_bodies_committed": False,
        "fandom_images_committed": False,
        "raw_archive_files_committed": False,
        "spreadsheet_committed": False,
        "generated_outputs_committed": False,
        "canonical_transcription_changed": False,
        "active_token_block_manifest_changed": False,
        "token_experiments_executed": False,
        "real_token_block_byte_streams_generated": False,
        "variant_byte_streams_generated": False,
        "variant_branches_enumerated": False,
        "real_variant_branches_materialised": False,
        "full_cartesian_product_enumerated": False,
        "sampled_real_variants_generated": False,
        "fandom_surface_combination_performed": False,
        "xor_attempt_performed": False,
        "transposition_attempt_performed": False,
        "outguess_execution_performed": False,
        "openpuff_execution_performed": False,
        "mp3stego_execution_performed": False,
        "stego_tool_execution_performed": False,
        "pgp_network_key_fetch_performed": False,
        "pgp_verification_performed_as_project_truth": False,
        "hash_search_performed": False,
        "hash_preimage_claim": False,
        "hash_comparison_performed": False,
        "decode_attempt_performed": False,
        "scored_experiments_executed": False,
        "benchmark_performed": False,
        "cryptanalytic_benchmark_performed": False,
        "cuda_execution_performed": False,
        "cuda_source_modified": False,
        "new_cuda_kernels_added": 0,
        "ocr_performed": False,
        "ai_ml_interpretation_performed": False,
        "llm_vision_token_reading_performed": False,
        "semantic_image_interpretation_performed": False,
        "hidden_content_image_forensics_performed": False,
        "audio_analysis_performed": False,
        "canonical_corpus_active": False,
        "page_boundaries_final": False,
        "method_status_upgraded": False,
        "public_website_publication_performed": False,
        "website_expansion_performed": False,
        "solve_claim": False,
    }


def build_token_block_external_context() -> dict[str, Any]:
    return {
        **_base_record(
            "stage5bi_token_block_external_context",
            _schema("schemas/token-block/stage5bi-token-block-external-context-v0.schema.json"),
        ),
        "token_block_id": "stage5ap-page49-51-token-block",
        "context_records": [
            {
                "context_id": "stage5bi-page49-51-fandom-context",
                "source_url": FANDOM_BASE + "Page_49-51",
                "context_role": "secondary_route_context",
                "token_block_relevance": "direct_token_block_relevance",
                "execution_allowed": False,
            },
            {
                "context_id": "stage5bi-page49-51-base-analysis-context",
                "source_url": FANDOM_BASE + "Liber_Primus_pp49-51_data_interpreted_as_base_59,_60,_62_and_64.",
                "context_role": "community_base_analysis_context",
                "token_block_relevance": "direct_token_block_relevance",
                "execution_allowed": False,
            },
            {
                "context_id": "stage5bi-boards-page49-51-thread-context",
                "source_url": "https://cicada3301.boards.net/thread/41/pages-49-256-byte-strings",
                "context_role": "source_gap_context",
                "token_block_relevance": "blocks_execution_until_source_locked",
                "execution_allowed": False,
            },
        ],
        "lineage_preserved": [
            "data/token-block/stage5ap-token-block-canonical-transcription.yaml",
            "data/token-block/stage5aw-repaired-token-variant-branch-manifest.yaml",
            "data/token-block/stage5az-repaired-bounded-variant-family-manifest.yaml",
            "data/token-block/stage5bd-active-manifest-lock.yaml",
        ],
        "canonical_transcription_changed": False,
        "active_token_block_manifest_changed": False,
        "stage5bd_dry_run_records_remain_valid": True,
        "future_token_block_execution_remains_blocked": True,
        "variant_byte_streams_generated": False,
        "execution_allowed": False,
        "solve_claim": False,
    }


def build_surface_token_block_context() -> dict[str, Any]:
    return {
        **_base_record(
            "stage5bi_2014_surface_token_block_context",
            _schema("schemas/token-block/stage5bi-2014-surface-token-block-context-v0.schema.json"),
        ),
        "context_id": "stage5bi-2014-surfaces-page49-51-context-only",
        "surface_ids": [
            "stage5bi-c01-2014-growing-hex-surface",
            "stage5bi-c02-2014-1033-hex-surface",
            "stage5bi-c03-2014-3301-hex-surface",
        ],
        "page49_51_surface_id": "stage5bi-c04-page49-51-256-token-surface",
        "context_only": True,
        "combination_allowed": False,
        "xor_attempt_performed": False,
        "transposition_attempt_performed": False,
        "hash_search_performed": False,
        "decode_attempt_performed": False,
        "execution_allowed": False,
        "canonical_transcription_changed": False,
        "active_token_block_manifest_changed": False,
        "solve_claim": False,
    }


def _workbook_metadata(path: Path) -> dict[str, Any]:
    resolved = _resolve(path)
    if not resolved.is_file():
        return {
            "spreadsheet_found": False,
            "source_path": path.as_posix(),
            "sha256": None,
            "file_size_bytes": None,
            "sheet_names": [],
            "sheet_dimensions": [],
            "parser": "not_present",
        }
    metadata: dict[str, Any] = {
        "spreadsheet_found": True,
        "source_path": _repo_relative(resolved),
        "sha256": _sha256_file(resolved),
        "file_size_bytes": resolved.stat().st_size,
        "sheet_names": [],
        "sheet_dimensions": [],
        "parser": "openpyxl_unavailable",
    }
    try:
        from openpyxl import load_workbook
    except ImportError:
        return metadata
    workbook = load_workbook(resolved, read_only=True, data_only=False)
    try:
        metadata["parser"] = "openpyxl_metadata_only"
        metadata["sheet_names"] = list(workbook.sheetnames)
        metadata["sheet_dimensions"] = [
            {
                "sheet_name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
            }
            for worksheet in workbook.worksheets
        ]
    finally:
        workbook.close()
    return metadata


def _stage5aw_possible_token_universe() -> dict[str, Any]:
    unresolved = _read_yaml("data/token-block/stage5aw-repaired-unresolved-token-variant-records.yaml")
    extras = _read_yaml("data/token-block/stage5aw-repaired-reviewer-extra-possible-tokens.yaml")
    tokens: set[str] = set()
    for record in unresolved.get("records", []):
        tokens.update(str(token) for token in record.get("possible_tokens", []))
    for record in extras.get("records", []):
        token = record.get("possible_token") or record.get("token")
        if token:
            tokens.add(str(token))
    return {
        "unresolved_variant_count": unresolved.get("unresolved_token_variant_count", 0),
        "reviewer_extra_record_count": extras.get("reviewer_extra_possible_token_count", len(extras.get("records", []))),
        "possible_token_universe_count": len(tokens),
        "possible_token_universe_hash": _sha256_text("\n".join(sorted(tokens))) if tokens else None,
    }


def build_spreadsheet_records() -> tuple[dict[str, Any], dict[str, Any]]:
    metadata = _workbook_metadata(LOCAL_SPREADSHEET)
    universe = _stage5aw_possible_token_universe()
    source_lock = {
        **_base_record(
            "stage5bi_local_spreadsheet_source_lock",
            _schema("schemas/source-harvester/stage5bi-local-spreadsheet-source-lock-v0.schema.json"),
        ),
        **metadata,
        "source_role": "local_analysis_metadata",
        "trusted_as_canonical": False,
        "spreadsheet_committed": False,
        "cell_body_committed": False,
        "generated_extract_committed": False,
        "canonical_transcription_changed": False,
        "active_token_block_manifest_changed": False,
        "execution_allowed": False,
        "solve_claim": False,
    }
    reconciliation = {
        **_base_record(
            "stage5bi_spreadsheet_stage5aw_reconciliation",
            _schema("schemas/token-block/stage5bi-spreadsheet-stage5aw-reconciliation-v0.schema.json"),
        ),
        "spreadsheet_found": metadata["spreadsheet_found"],
        "spreadsheet_source_path": metadata["source_path"],
        "spreadsheet_sha256": metadata["sha256"],
        "sheet_names": metadata["sheet_names"],
        "stage5aw_repaired_branch_manifest": "data/token-block/stage5aw-repaired-token-variant-branch-manifest.yaml",
        "stage5aw_repaired_unresolved_records": "data/token-block/stage5aw-repaired-unresolved-token-variant-records.yaml",
        "stage5aw_reviewer_extra_possible_tokens": "data/token-block/stage5aw-repaired-reviewer-extra-possible-tokens.yaml",
        "stage5ay_branch_eligibility_policy": "data/token-block/stage5ay-branch-eligibility-policy.yaml",
        "stage5aw_unresolved_variant_count": universe["unresolved_variant_count"],
        "stage5aw_possible_token_universe_count": universe["possible_token_universe_count"],
        "stage5aw_possible_token_universe_hash": universe["possible_token_universe_hash"],
        "spreadsheet_parse_status": "manual_review_required" if metadata["spreadsheet_found"] else "spreadsheet_absent",
        "spreadsheet_primary_token_stream_has_256_entries": None,
        "spreadsheet_alternative_relation_to_stage5aw": "manual_review_required"
        if metadata["spreadsheet_found"]
        else "not_available",
        "missing_stage5aw_alternatives": [],
        "extra_spreadsheet_alternatives": [],
        "ambiguous_i_l_o_zero_variants_status": "stage5aw_representation_present_but_spreadsheet_manual_review_required"
        if metadata["spreadsheet_found"]
        else "spreadsheet_absent",
        "spreadsheet_is_local_analysis_metadata": True,
        "trusted_as_canonical": False,
        "canonical_transcription_changed": False,
        "branch_manifest_changed": False,
        "active_token_block_manifest_changed": False,
        "execution_allowed": False,
        "variant_byte_streams_generated": False,
        "solve_claim": False,
    }
    return source_lock, reconciliation


def build_crosswalk_summary(
    page_triage: dict[str, Any],
    item_candidates: dict[str, Any],
    crosswalk: dict[str, Any],
    spreadsheet_source_lock: dict[str, Any],
) -> dict[str, Any]:
    status_counter = Counter(record["archive_crosswalk_status"] for record in crosswalk.get("records", []))
    return {
        **_base_record(
            "stage5bi_fandom_crosswalk_source_summary",
            _schema("schemas/source-harvester/stage5bi-fandom-crosswalk-source-summary-v0.schema.json"),
        ),
        "fandom_page_triage_count": page_triage["page_count"],
        "item_source_lock_candidate_count": item_candidates["candidate_count"],
        "highest_priority_candidate_count": item_candidates["highest_priority_candidate_count"],
        "original_archive_crosswalk_candidate_count": crosswalk["candidate_count"],
        "archive_crosswalk_status_counts": dict(sorted(status_counter.items())),
        "local_archive_present": _resolve(LOCAL_ARCHIVE).is_dir(),
        "local_spreadsheet_present": spreadsheet_source_lock["spreadsheet_found"],
        "fandom_media_policy": "data/historical-route/stage5bi-fandom-media-non-original-policy.yaml",
        "raw_archive_files_committed": False,
        "fandom_page_bodies_committed": False,
        "spreadsheet_committed": False,
        "execution_allowed": False,
        "solve_claim": False,
    }


def build_next_stage_decision(summary_preview: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        **_base_record(
            "stage5bi_next_stage_decision",
            _schema("schemas/project-state/stage5bi-next-stage-decision-v0.schema.json"),
        ),
        "selected_next_stage_id": "stage-5bj",
        "selected_next_prompt_type": "codex_metadata_implementation",
        "selected_next_stage_title": (
            "Stage 5BJ - Original-archive crosswalk closure for high-priority Fandom-derived candidates, without execution"
        ),
        "selected_next_stage_reason": (
            "Stage 5BI leaves high-value Fandom-derived surfaces and page bodies blocked until exact original/archive-equivalent "
            "crosswalks close; the next stage should close those metadata gaps before broader historical-route planning."
        ),
        "token_block_execution_selected": False,
        "dwh_hash_search_selected": False,
        "scored_experiments_selected": False,
        "benchmark_selected": False,
        "cuda_selected": False,
        "public_website_expansion_selected": False,
        "stego_execution_selected": False,
        "pgp_verification_selected": False,
        "metadata_source_lock_only": True,
        "solve_claim": False,
        "summary_counts_preview": summary_preview or {},
    }


def build_summary(
    page_triage: dict[str, Any],
    item_candidates: dict[str, Any],
    crosswalk: dict[str, Any],
    source_gaps: dict[str, Any],
    negative_controls: dict[str, Any],
    spreadsheet_source_lock: dict[str, Any],
    spreadsheet_reconciliation: dict[str, Any],
    guardrail: dict[str, Any],
    next_stage: dict[str, Any],
) -> dict[str, Any]:
    return {
        **_base_record("stage5bi_summary", _schema("schemas/project-state/stage5bi-summary-v0.schema.json")),
        "status": "complete",
        "source_stage_ids": ["stage-5bf", "stage-5bg", "stage-5bh-dr"],
        "fandom_page_triage_count": page_triage["page_count"],
        "item_source_lock_candidate_count": item_candidates["candidate_count"],
        "highest_priority_candidate_count": item_candidates["highest_priority_candidate_count"],
        "original_archive_crosswalk_candidate_count": crosswalk["candidate_count"],
        "verified_archive_crosswalk_count": crosswalk["verified_archive_crosswalk_count"],
        "probable_archive_path_candidate_count": crosswalk["probable_archive_path_candidate_count"],
        "source_gap_count": source_gaps["source_gap_count"],
        "negative_control_count": negative_controls["negative_control_count"],
        "fandom_media_policy_created": True,
        "surface_2014_context_created": True,
        "surface_2014_count": 3,
        "page49_51_surface_context_created": True,
        "spreadsheet_source_lock_created": True,
        "spreadsheet_reconciliation_created": True,
        "spreadsheet_found": spreadsheet_source_lock["spreadsheet_found"],
        "local_archive_present": _resolve(LOCAL_ARCHIVE).is_dir(),
        "canonical_transcription_changed": False,
        "active_token_block_manifest_changed": False,
        "stage5bd_dry_run_records_remain_valid": True,
        "future_token_block_execution_remains_blocked": True,
        "recommended_next_prompt_type": next_stage["selected_next_prompt_type"],
        "recommended_next_stage_title": next_stage["selected_next_stage_title"],
        "recommended_next_stage_reason": next_stage["selected_next_stage_reason"],
        **{
            key: value
            for key, value in guardrail.items()
            if key not in {"record_type", "schema", "stage_id", "stage_title"}
        },
    }


def build_stage5bi_records() -> dict[str, Any]:
    page_triage = build_fandom_page_triage()
    item_candidates = build_item_source_lock_candidates()
    crosswalk = build_original_archive_crosswalk()
    media_policy = build_media_policy()
    surface_context = build_2014_surface_context(crosswalk)
    negative_controls = build_negative_controls()
    source_gaps = build_source_gaps()
    guardrail = build_guardrail()
    token_block_context = build_token_block_external_context()
    surface_token_block_context = build_surface_token_block_context()
    spreadsheet_source_lock, spreadsheet_reconciliation = build_spreadsheet_records()
    crosswalk_summary = build_crosswalk_summary(page_triage, item_candidates, crosswalk, spreadsheet_source_lock)
    next_stage = build_next_stage_decision(
        {
            "fandom_page_triage_count": page_triage["page_count"],
            "item_source_lock_candidate_count": item_candidates["candidate_count"],
            "source_gap_count": source_gaps["source_gap_count"],
        }
    )
    summary = build_summary(
        page_triage,
        item_candidates,
        crosswalk,
        source_gaps,
        negative_controls,
        spreadsheet_source_lock,
        spreadsheet_reconciliation,
        guardrail,
        next_stage,
    )
    payloads = {
        "page_triage": page_triage,
        "item_candidates": item_candidates,
        "archive_crosswalk": crosswalk,
        "media_policy": media_policy,
        "surface_context": surface_context,
        "negative_controls": negative_controls,
        "source_gaps": source_gaps,
        "guardrail": guardrail,
        "token_block_context": token_block_context,
        "surface_token_block_context": surface_token_block_context,
        "spreadsheet_reconciliation": spreadsheet_reconciliation,
        "spreadsheet_source_lock": spreadsheet_source_lock,
        "crosswalk_summary": crosswalk_summary,
        "summary": summary,
        "next_stage": next_stage,
    }
    for key, payload in payloads.items():
        _write_yaml(DATA_PATHS[key], payload)
    return payloads


def _load_stage5bi_payloads(paths: dict[str, Path] | None = None) -> dict[str, dict[str, Any]]:
    selected = paths or DATA_PATHS
    return {key: _read_yaml(path) for key, path in selected.items()}


def validate_stage5bi(paths: dict[str, Path] | None = None) -> dict[str, Any]:
    selected = paths or DATA_PATHS
    payloads = _load_stage5bi_payloads(selected)
    errors: list[str] = []
    for key, path in selected.items():
        if not _resolve(path).is_file():
            errors.append(f"missing Stage 5BI file for {key}: {path}")
    page_triage = payloads.get("page_triage", {})
    item_candidates = payloads.get("item_candidates", {})
    crosswalk = payloads.get("archive_crosswalk", {})
    surface_context = payloads.get("surface_context", {})
    negative_controls = payloads.get("negative_controls", {})
    source_gaps = payloads.get("source_gaps", {})
    guardrail = payloads.get("guardrail", {})
    token_block_context = payloads.get("token_block_context", {})
    spreadsheet_source_lock = payloads.get("spreadsheet_source_lock", {})
    spreadsheet_reconciliation = payloads.get("spreadsheet_reconciliation", {})
    summary = payloads.get("summary", {})
    next_stage = payloads.get("next_stage", {})

    if page_triage.get("page_count") != len(page_triage.get("records", [])):
        errors.append("page triage count does not match records")
    if item_candidates.get("candidate_count") != len(item_candidates.get("records", [])):
        errors.append("item source-lock candidate count does not match records")
    if crosswalk.get("candidate_count") != len(crosswalk.get("records", [])):
        errors.append("archive crosswalk candidate count does not match records")
    if source_gaps.get("source_gap_count") != len(source_gaps.get("gaps", [])):
        errors.append("source gap count does not match gaps")
    if negative_controls.get("negative_control_count") != len(negative_controls.get("records", [])):
        errors.append("negative-control count does not match records")
    if surface_context.get("surface_count") != 3:
        errors.append("2014 surface context must contain exactly three surfaces")
    for surface in surface_context.get("surfaces", []):
        if surface.get("execution_allowed") is not False or surface.get("combination_with_page49_51_allowed") is not False:
            errors.append(f"surface {surface.get('surface_id')} must block execution and page49-51 combination")

    page_urls = {record.get("source_url") for record in page_triage.get("records", [])}
    if FANDOM_BASE + "What_Happened_Part_1_(2014)" not in page_urls:
        errors.append("What Happened Part 1 (2014) missing from Fandom page triage")
    if FANDOM_BASE + "Page_49-51" not in page_urls:
        errors.append("Page 49-51 missing from Fandom page triage")
    context_urls = {record.get("source_url") for record in token_block_context.get("context_records", [])}
    if FANDOM_BASE + "Page_49-51" not in context_urls:
        errors.append("Page 49-51 missing from token-block external context")

    negative_ids = {record.get("negative_control_id") for record in negative_controls.get("records", [])}
    for required_id in {
        "stage5bi-infamous-warning",
        "stage5bi-fake-puzzles",
        "stage5bi-telnet-recreation",
        "stage5bi-pgp-signature-live-cd-deletion-candidate",
    }:
        if required_id not in negative_ids:
            errors.append(f"missing negative-control record {required_id}")
    gap_ids = {gap.get("gap_id") for gap in source_gaps.get("gaps", [])}
    for required_gap in {
        "stage5bi-fandom-2014-page-body-not-hash-locked",
        "stage5bi-boards-page49-51-thread-inaccessible",
        "stage5bi-2014-256-surfaces-need-original-archive-crosswalk",
        "stage5bi-fandom-media-original-equivalence-not-assumed",
        "stage5bi-spreadsheet-not-canonical",
    }:
        if required_gap not in gap_ids:
            errors.append(f"missing source gap {required_gap}")

    false_guardrail_keys = [
        "fandom_page_bodies_committed",
        "fandom_images_committed",
        "raw_archive_files_committed",
        "spreadsheet_committed",
        "generated_outputs_committed",
        "canonical_transcription_changed",
        "active_token_block_manifest_changed",
        "token_experiments_executed",
        "real_token_block_byte_streams_generated",
        "variant_byte_streams_generated",
        "variant_branches_enumerated",
        "real_variant_branches_materialised",
        "full_cartesian_product_enumerated",
        "sampled_real_variants_generated",
        "fandom_surface_combination_performed",
        "xor_attempt_performed",
        "transposition_attempt_performed",
        "outguess_execution_performed",
        "openpuff_execution_performed",
        "mp3stego_execution_performed",
        "stego_tool_execution_performed",
        "pgp_network_key_fetch_performed",
        "pgp_verification_performed_as_project_truth",
        "hash_search_performed",
        "hash_preimage_claim",
        "hash_comparison_performed",
        "decode_attempt_performed",
        "scored_experiments_executed",
        "benchmark_performed",
        "cryptanalytic_benchmark_performed",
        "cuda_execution_performed",
        "cuda_source_modified",
        "ocr_performed",
        "ai_ml_interpretation_performed",
        "llm_vision_token_reading_performed",
        "semantic_image_interpretation_performed",
        "hidden_content_image_forensics_performed",
        "audio_analysis_performed",
        "canonical_corpus_active",
        "page_boundaries_final",
        "method_status_upgraded",
        "public_website_publication_performed",
        "website_expansion_performed",
        "solve_claim",
    ]
    for key in false_guardrail_keys:
        if guardrail.get(key) is not False:
            errors.append(f"guardrail {key} must be false")
        if key in summary and summary.get(key) is not False:
            errors.append(f"summary {key} must be false")
    if guardrail.get("metadata_only") is not True:
        errors.append("guardrail metadata_only must be true")
    if guardrail.get("new_cuda_kernels_added") != 0:
        errors.append("new_cuda_kernels_added must be zero")
    if spreadsheet_source_lock.get("trusted_as_canonical") is not False:
        errors.append("spreadsheet source lock must not be canonical")
    if spreadsheet_reconciliation.get("canonical_transcription_changed") is not False:
        errors.append("spreadsheet reconciliation must not change canonical transcription")
    if summary.get("stage5bd_dry_run_records_remain_valid") is not True:
        errors.append("Stage 5BD dry-run records must remain valid")
    if summary.get("active_token_block_manifest_changed") is not False:
        errors.append("active token-block manifest must not change")
    if next_stage.get("selected_next_stage_id") != "stage-5bj":
        errors.append("Stage 5BI next stage must be Stage 5BJ")
    for key in [
        "token_block_execution_selected",
        "dwh_hash_search_selected",
        "scored_experiments_selected",
        "benchmark_selected",
        "cuda_selected",
        "public_website_expansion_selected",
        "stego_execution_selected",
        "pgp_verification_selected",
    ]:
        if next_stage.get(key) is not False:
            errors.append(f"next-stage decision {key} must be false")
    bad_staged_prefixes = (
        "third_party/CicadaSolversIddqd/",
        "third_party/SourceSnapshots/",
        "data/raw/",
        "experiments/results/",
    )
    bad_staged = [path for path in _git_staged_paths() if path.startswith(bad_staged_prefixes)]
    if bad_staged:
        errors.append(f"raw/generated paths staged: {bad_staged}")

    result = {
        "stage5bi_valid": not errors,
        "fandom_page_triage_count": summary.get("fandom_page_triage_count", 0),
        "item_source_lock_candidate_count": summary.get("item_source_lock_candidate_count", 0),
        "original_archive_crosswalk_candidate_count": summary.get("original_archive_crosswalk_candidate_count", 0),
        "verified_archive_crosswalk_count": summary.get("verified_archive_crosswalk_count", 0),
        "source_gap_count": summary.get("source_gap_count", 0),
        "negative_control_count": summary.get("negative_control_count", 0),
        "spreadsheet_found": summary.get("spreadsheet_found"),
        "local_archive_present": summary.get("local_archive_present"),
        "selected_next_stage_id": next_stage.get("selected_next_stage_id"),
        "validation_error_count": len(errors),
        "validation_errors": errors,
    }
    if errors:
        raise ValueError(json.dumps(result, indent=2, sort_keys=True))
    return result


def summarize_stage5bi(path: Path = DATA_PATHS["summary"]) -> dict[str, Any]:
    return _read_yaml(path)
