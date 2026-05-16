"""Workbook loading and checksum helpers."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


@dataclass(frozen=True)
class LoadedLegacyWorkbook:
    path: Path
    sha256: str
    formulas: Workbook
    values: Workbook


def compute_sha256(path: Path) -> str:
    """Compute SHA-256 for a workbook or other local file."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_legacy_workbook(path: Path) -> LoadedLegacyWorkbook:
    """Load workbook twice: formula text and cached/display values."""
    resolved = path.resolve()
    sha256 = compute_sha256(resolved)
    formulas = load_workbook(resolved, data_only=False, read_only=False)
    values = load_workbook(resolved, data_only=True, read_only=False)
    return LoadedLegacyWorkbook(path=resolved, sha256=sha256, formulas=formulas, values=values)
