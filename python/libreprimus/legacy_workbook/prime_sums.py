"""Prime Sums extraction for the legacy workbook."""

from __future__ import annotations

import unicodedata
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from libreprimus.legacy_workbook.loader import LoadedLegacyWorkbook
from libreprimus.legacy_workbook.models import SOURCE_ID, PrimeSumRecord, WarningRecord

SEPARATOR_TOKENS = {"", ".", ",", ";", ":", "·", "·", "\u0387"}
TRUE_VALUES = {"true", "prawda", "1", "yes"}
FALSE_VALUES = {"false", "fałsz", "falsz", "0", "no"}


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    try:
        numeric = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if numeric.is_integer():
        return int(numeric)
    return None


def _clean_token(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return None
    text = str(value).strip()
    if text in SEPARATOR_TOKENS:
        return None
    return text or None


def parse_bool_like(value: Any) -> tuple[bool | None, str | None]:
    """Parse boolean-like workbook values, preserving raw text."""
    if value is None or value == "":
        return None, None
    if isinstance(value, bool):
        return value, str(value)
    if isinstance(value, (int, float)) and value in {0, 1}:
        return bool(value), str(int(value))

    raw = str(value).strip()
    key = unicodedata.normalize("NFKD", raw).casefold()
    key_without_marks = "".join(ch for ch in key if not unicodedata.combining(ch))
    if key in TRUE_VALUES or key_without_marks in TRUE_VALUES:
        return True, raw
    if key in FALSE_VALUES or key_without_marks in FALSE_VALUES:
        return False, raw
    return None, raw


def _row_has_tokens(worksheet: Worksheet, row: int) -> bool:
    for column in range(3, int(worksheet.max_column or 0) + 1):
        if _clean_token(worksheet.cell(row, column).value) is not None:
            return True
    return False


def _row_has_values(worksheet: Worksheet, row: int) -> bool:
    if _to_int(worksheet.cell(row, 1).value) is not None:
        return True
    for column in range(3, int(worksheet.max_column or 0) + 1):
        if _to_int(worksheet.cell(row, column).value) is not None:
            return True
    return False


def extract_prime_sums(
    loaded: LoadedLegacyWorkbook,
) -> tuple[list[PrimeSumRecord], list[WarningRecord]]:
    """Extract alternating token/value rows from the Prime Sums sheet."""
    if "Prime Sums" not in loaded.formulas.sheetnames:
        return [], [
            WarningRecord(
                record_type="legacy_workbook_warning",
                source_id=SOURCE_ID,
                workbook_sha256=loaded.sha256,
                sheet_name=None,
                message="Prime Sums sheet not found.",
            )
        ]

    worksheet = loaded.formulas["Prime Sums"]
    records: list[PrimeSumRecord] = []
    warnings: list[WarningRecord] = []
    sequence_index = 0
    row = 1

    while row < int(worksheet.max_row or 0):
        text_row = row
        value_row = row + 1
        if not (_row_has_tokens(worksheet, text_row) and _row_has_values(worksheet, value_row)):
            row += 1
            continue

        tokens: list[str] = []
        values: list[int] = []
        missing_values = 0
        extra_values = 0

        for column in range(3, int(worksheet.max_column or 0) + 1):
            token = _clean_token(worksheet.cell(text_row, column).value)
            value = _to_int(worksheet.cell(value_row, column).value)
            if token is not None and value is not None:
                tokens.append(token)
                values.append(value)
            elif token is not None and value is None:
                missing_values += 1
            elif token is None and value is not None:
                extra_values += 1

        if not tokens and not values:
            row += 1
            continue

        is_prime, raw_is_prime = parse_bool_like(worksheet.cell(value_row, 2).value)
        if raw_is_prime is not None and is_prime is None:
            warnings.append(
                WarningRecord(
                    record_type="legacy_workbook_warning",
                    source_id=SOURCE_ID,
                    workbook_sha256=loaded.sha256,
                    sheet_name=worksheet.title,
                    message=f"Unrecognized Prime Sums boolean value: {raw_is_prime}.",
                    cell=worksheet.cell(value_row, 2).coordinate,
                    row=value_row,
                )
            )

        if missing_values or extra_values:
            warnings.append(
                WarningRecord(
                    record_type="legacy_workbook_warning",
                    source_id=SOURCE_ID,
                    workbook_sha256=loaded.sha256,
                    sheet_name=worksheet.title,
                    message=(
                        "Prime Sums token/value alignment warning: "
                        f"missing_values={missing_values}, extra_values={extra_values}."
                    ),
                    row=value_row,
                )
            )

        records.append(
            PrimeSumRecord(
                record_type="legacy_prime_sum",
                source_id=SOURCE_ID,
                workbook_sha256=loaded.sha256,
                sheet_name=worksheet.title,
                text_row=text_row,
                value_row=value_row,
                sequence_index=sequence_index,
                tokens=tokens,
                prime_values=values,
                sum=_to_int(worksheet.cell(value_row, 1).value),
                is_prime=is_prime,
                trusted_as_canonical=False,
                raw_is_prime=raw_is_prime,
            )
        )
        sequence_index += 1
        row += 2

    return records, warnings
