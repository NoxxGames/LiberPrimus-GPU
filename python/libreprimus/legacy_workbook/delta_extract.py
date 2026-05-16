"""Solved-delta extraction for non-canonical legacy workbook sheets."""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from libreprimus.legacy_workbook.inventory import classify_sheet
from libreprimus.legacy_workbook.loader import LoadedLegacyWorkbook
from libreprimus.legacy_workbook.models import SOURCE_ID, SolvedDeltaRecord, WarningRecord

LINE_RE = re.compile(r"^line\s+(\d+)\b", re.IGNORECASE)


@dataclass(frozen=True)
class _LineBlock:
    label: str
    number: int | None
    label_row: int
    label_column: int
    row_by_label: dict[str, int]


def _norm(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _label_key(value: Any) -> str:
    return _norm(value).casefold().replace(" ", "").replace("-", "")


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    try:
        numeric = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if numeric.is_integer():
        return int(numeric)
    return None


def _find_line_blocks(worksheet: Worksheet) -> list[_LineBlock]:
    blocks: list[_LineBlock] = []
    for row in worksheet.iter_rows():
        for cell in row:
            text = _norm(cell.value)
            match = LINE_RE.match(text)
            if not match:
                continue

            row_by_label: dict[str, int] = {}
            for offset in range(1, 8):
                label_cell = worksheet.cell(cell.row + offset, cell.column)
                key = _label_key(label_cell.value)
                if key in {"cipherrune", "gematriaposition", "messagerune", "ciphermessage", "messagecipher"}:
                    if key == "gematriaposition" and "cipher_index" not in row_by_label:
                        row_by_label["cipher_index"] = label_cell.row
                    elif key == "gematriaposition":
                        row_by_label["message_index"] = label_cell.row
                    elif key == "cipherrune":
                        row_by_label["cipher_rune"] = label_cell.row
                    elif key == "messagerune":
                        row_by_label["message_token"] = label_cell.row
                    elif key == "ciphermessage":
                        row_by_label["cipher_minus_message"] = label_cell.row
                    elif key == "messagecipher":
                        row_by_label["message_minus_cipher"] = label_cell.row

            blocks.append(
                _LineBlock(
                    label=text,
                    number=int(match.group(1)),
                    label_row=cell.row,
                    label_column=cell.column,
                    row_by_label=row_by_label,
                )
            )
    return blocks


def _cell_value(worksheet: Worksheet, row: int | None, column: int) -> Any:
    if row is None:
        return None
    return worksheet.cell(row, column).value


def _string_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def extract_solved_deltas(
    loaded: LoadedLegacyWorkbook,
) -> tuple[list[SolvedDeltaRecord], list[WarningRecord]]:
    """Extract aligned solved-delta records from known solved-delta sheets."""
    records: list[SolvedDeltaRecord] = []
    warnings: list[WarningRecord] = []

    for sheet_index, worksheet in enumerate(loaded.formulas.worksheets):
        if classify_sheet(worksheet.title) != "solved_delta_sheet":
            continue

        for block in _find_line_blocks(worksheet):
            required = {
                "cipher_rune",
                "cipher_index",
                "message_token",
                "message_index",
                "cipher_minus_message",
                "message_minus_cipher",
            }
            missing = sorted(required.difference(block.row_by_label))
            if missing:
                warnings.append(
                    WarningRecord(
                        record_type="legacy_workbook_warning",
                        source_id=SOURCE_ID,
                        workbook_sha256=loaded.sha256,
                        sheet_name=worksheet.title,
                        message=f"Line block is missing expected rows: {', '.join(missing)}.",
                        line_label=block.label,
                        row=block.label_row,
                    )
                )

            position = 0
            seen_aligned_data = False
            for column in range(block.label_column + 1, int(worksheet.max_column or 0) + 1):
                cipher_rune = _string_or_none(
                    _cell_value(worksheet, block.row_by_label.get("cipher_rune"), column)
                )
                cipher_index = _to_int(_cell_value(worksheet, block.row_by_label.get("cipher_index"), column))
                message_token = _string_or_none(
                    _cell_value(worksheet, block.row_by_label.get("message_token"), column)
                )
                message_index = _to_int(_cell_value(worksheet, block.row_by_label.get("message_index"), column))
                cipher_minus = _to_int(
                    _cell_value(worksheet, block.row_by_label.get("cipher_minus_message"), column)
                )
                message_minus = _to_int(
                    _cell_value(worksheet, block.row_by_label.get("message_minus_cipher"), column)
                )

                values = (cipher_rune, cipher_index, message_token, message_index, cipher_minus, message_minus)
                if all(value is None for value in values):
                    if seen_aligned_data:
                        break
                    continue
                seen_aligned_data = True

                record = SolvedDeltaRecord(
                    record_type="legacy_solved_delta",
                    source_id=SOURCE_ID,
                    workbook_sha256=loaded.sha256,
                    sheet_name=worksheet.title,
                    sheet_index=sheet_index,
                    line_label=block.label,
                    line_number=block.number,
                    position_in_line=position,
                    source_row=block.row_by_label.get("cipher_rune", block.label_row),
                    source_column=column,
                    cipher_rune=cipher_rune,
                    cipher_index=cipher_index,
                    message_token=message_token,
                    message_index=message_index,
                    cipher_minus_message_mod29=cipher_minus,
                    message_minus_cipher_mod29=message_minus,
                    trusted_as_canonical=False,
                    trusted_as_solved_fixture_hint=True,
                )
                records.append(record)

                if any(value is None for value in values):
                    warnings.append(
                        WarningRecord(
                            record_type="legacy_workbook_warning",
                            source_id=SOURCE_ID,
                            workbook_sha256=loaded.sha256,
                            sheet_name=worksheet.title,
                            message="Aligned solved-delta position has missing values.",
                            cell=worksheet.cell(record.source_row, column).coordinate,
                            line_label=block.label,
                            row=record.source_row,
                        )
                    )
                elif (cipher_index - message_index) % 29 != cipher_minus:
                    warnings.append(
                        WarningRecord(
                            record_type="legacy_workbook_warning",
                            source_id=SOURCE_ID,
                            workbook_sha256=loaded.sha256,
                            sheet_name=worksheet.title,
                            message="Cipher-minus-message modulo 29 validation failed.",
                            cell=worksheet.cell(record.source_row, column).coordinate,
                            line_label=block.label,
                            row=record.source_row,
                        )
                    )
                elif (message_index - cipher_index) % 29 != message_minus:
                    warnings.append(
                        WarningRecord(
                            record_type="legacy_workbook_warning",
                            source_id=SOURCE_ID,
                            workbook_sha256=loaded.sha256,
                            sheet_name=worksheet.title,
                            message="Message-minus-cipher modulo 29 validation failed.",
                            cell=worksheet.cell(record.source_row, column).coordinate,
                            line_label=block.label,
                            row=record.source_row,
                        )
                    )

                position += 1

    return records, warnings
