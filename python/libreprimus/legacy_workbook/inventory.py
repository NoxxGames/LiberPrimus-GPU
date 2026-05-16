"""Sheet inventory for legacy workbooks."""

from __future__ import annotations

from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from libreprimus.legacy_workbook.loader import LoadedLegacyWorkbook
from libreprimus.legacy_workbook.models import (
    SOURCE_ID,
    SOLVED_DELTA_SHEETS,
    SheetRecord,
    WarningRecord,
)


def classify_sheet(sheet_name: str) -> str:
    """Classify a sheet by its known Stage 0B role."""
    if sheet_name == "README":
        return "readme"
    if sheet_name == "Prime Sums":
        return "prime_sums"
    if sheet_name in SOLVED_DELTA_SHEETS:
        return "solved_delta_sheet"
    return "unknown"


def _is_formula(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _stringify(value: object) -> str:
    return str(value)


def inventory_sheets(loaded: LoadedLegacyWorkbook) -> tuple[list[SheetRecord], list[WarningRecord]]:
    """Build sheet inventory records and warnings."""
    records: list[SheetRecord] = []
    warnings: list[WarningRecord] = []

    for index, worksheet in enumerate(loaded.formulas.worksheets):
        non_empty_count = 0
        formula_count = 0
        first_cell: Cell | None = None

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                non_empty_count += 1
                if first_cell is None:
                    first_cell = cell
                if _is_formula(cell):
                    formula_count += 1

        classification = classify_sheet(worksheet.title)
        if classification == "unknown":
            warnings.append(
                WarningRecord(
                    record_type="legacy_workbook_warning",
                    source_id=SOURCE_ID,
                    workbook_sha256=loaded.sha256,
                    sheet_name=worksheet.title,
                    message="Unknown sheet classification.",
                )
            )

        records.append(
            SheetRecord(
                record_type="legacy_workbook_sheet",
                source_id=SOURCE_ID,
                workbook_sha256=loaded.sha256,
                sheet_index=index,
                sheet_name=worksheet.title,
                max_row=int(worksheet.max_row or 0),
                max_column=int(worksheet.max_column or 0),
                non_empty_cell_count=non_empty_count,
                formula_cell_count=formula_count,
                classification=classification,
                trusted_as_canonical=False,
                trusted_as_solved_fixture_hint=classification == "solved_delta_sheet",
                first_non_empty_cell=first_cell.coordinate if first_cell is not None else None,
                first_non_empty_row=first_cell.row if first_cell is not None else None,
                first_non_empty_value=_stringify(first_cell.value) if first_cell is not None else None,
            )
        )

    return records, warnings


def worksheet_by_name(loaded: LoadedLegacyWorkbook, sheet_name: str) -> Worksheet:
    """Return a formula worksheet by name."""
    return loaded.formulas[sheet_name]
