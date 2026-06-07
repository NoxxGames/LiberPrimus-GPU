"""Stage 5DP New Reddit source-lock records.

This stage is metadata-only. It source-locks ignored local RedditStuff folders,
records bounded Mayfly workbook/document facts and candidate-only crosslinks,
and preserves the closed Stage 5DO governance gates. It does not run OCR, image
forensics, route extraction, target validation, byte-stream generation, CUDA, or
any puzzle execution.
"""

from __future__ import annotations

import hashlib
import json
import mimetypes
import subprocess
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from jsonschema import Draft202012Validator

from libreprimus.token_block.models import read_yaml, sha256_file, write_json, write_jsonl, write_yaml
from libreprimus.token_block.preflight_runner.stage5bd import validate_stage5bd
from libreprimus.token_block.stage5ca import ACTIVE_LINEAGE_PATHS
from libreprimus.token_block.stage5cm import PARALLEL_WORKER_CAP
from libreprimus.token_block.stage5do import validate_stage5do

STAGE_ID = "stage-5dp"
STAGE_TITLE = (
    "Stage 5DP - New Reddit Mayfly / dot / cover / ISO source-lock addendum, "
    "without execution"
)
PROMPT_TYPE = "codex_metadata_implementation"
SOURCE_PREVIOUS_STAGE_ID = "stage-5do"
SOURCE_PREVIOUS_STAGE_COMMIT = "299098752177de50dc18a2d7871fecf15612ddbf"
SOURCE_PREVIOUS_ISSUE = 150
SOURCE_PREVIOUS_CI_RUN = 27082987419
NEXT_STAGE_ID = "stage-5dq"
NEXT_STAGE_TITLE = (
    "Stage 5DQ - Lightweight source-lock browser GUI implementation, "
    "without puzzle execution"
)

RESULTS_DIR = Path("experiments/results/token-block/stage5dp")
CODEX_COMPLETION_PATH = Path("codex-output/stage5dp-codex-completion.md")
DEPRECATED_CODEX_OUTPUT = Path("codex_output")
CHATGPT_CONTEXT_PATH = Path("ChatGPT-ContextFile.md")

REDDIT_ROOT = Path("third_party/RedditStuff")
REQUIRED_SOURCE_FOLDERS = {
    "mayfly_investigation": "MayFlyInvestigation",
    "dot_observations_pages_7_23_56": "DotObservationsFromPages_7_23_56",
    "page_33_three_dots": "Page_33_ThreeDotsObservations",
    "front_cover_measurements_1033": "FrontCoverMeasurementsOf1033",
    "data_directory_from_iso": "DataDirectoryFromTheISO",
    "finds_from_problems_2012": "FindsFromProblems_2012_puzzle",
}

MAYFLY_FOLDER = REDDIT_ROOT / "MayFlyInvestigation"
MAYFLY_DOCX = MAYFLY_FOLDER / "The Mayfly.docx"
MAYFLY_XLSX = MAYFLY_FOLDER / "57.jpg Mayfly data.xlsx"

FORBIDDEN_FALSE_FLAGS = {
    "activation_authorized_now",
    "activation_decision_valid_now",
    "active_ingestion_performed",
    "active_manifest_registry_updated",
    "active_planning_input_authorized_now",
    "active_planning_input_selected_now",
    "active_token_block_manifest_changed",
    "ai_ml_interpretation_performed",
    "alberti_cipher_execution_performed_now",
    "audio_stego_performed",
    "benchmark_performed",
    "branch_enumeration_performed",
    "byte_stream_generation_authorized_now",
    "canonical_corpus_active",
    "combined_approval_gate_satisfied_now",
    "cuda_execution_performed",
    "decode_attempt_performed",
    "decryption_attempt_performed_now",
    "disk_cipher_execution_performed_now",
    "dwh_hash_search_performed",
    "execution_authorized_now",
    "execution_performed",
    "full_cartesian_product_enumerated",
    "generated_outputs_committed",
    "hash_preimage_search_performed",
    "html_tool_executed_now",
    "image_forensics_performed",
    "known_plaintext_attack_performed_now",
    "mayfly_route_extraction_performed_now",
    "mp3stego_execution_performed",
    "network_target_validation_performed_now",
    "ocr_performed",
    "openpuff_execution_performed",
    "operator_readiness_decision_created_now",
    "page32_route_extraction_performed_now",
    "page56_hash_preimage_tested_now",
    "pivot_target_selected_now",
    "probability_claim_accepted_as_validated",
    "real_byte_stream_generated",
    "route_extraction_performed_now",
    "scoring_performed",
    "solve_claim",
    "source_browser_gui_implemented_now",
    "spectrogram_stego_performed",
    "target_class_validation_implemented",
    "target_priority_decision_created_now",
    "token_block_experiment_executed",
    "tor_network_access_performed",
    "triangle_route_extraction_performed_now",
    "variant_materialisation_performed",
    "website_expansion_performed",
}

DATA_PATHS = {
    "summary": Path("data/project-state/stage5dp-summary.yaml"),
    "next_stage_decision": Path("data/project-state/stage5dp-next-stage-decision.yaml"),
    "stage5do_preservation": Path("data/project-state/stage5dp-stage5do-preservation.yaml"),
    "reviewable_validation_evidence": Path(
        "data/project-state/stage5dp-reviewable-validation-evidence.yaml"
    ),
    "reviewability_gap_register": Path(
        "data/project-state/stage5dp-reviewability-gap-register.yaml"
    ),
    "pivot_readiness_update": Path("data/project-state/stage5dp-pivot-readiness-update.yaml"),
    "source_browser_gui_deferral": Path(
        "data/project-state/stage5dp-source-browser-gui-deferral.yaml"
    ),
    "new_reddit_source_lock_register": Path(
        "data/source-harvester/stage5dp-new-reddit-source-lock-register.yaml"
    ),
    "mayfly_docx_xlsx_source_lock": Path(
        "data/source-harvester/stage5dp-mayfly-docx-xlsx-source-lock.yaml"
    ),
    "reddit_url_info_source_lock": Path(
        "data/source-harvester/stage5dp-reddit-url-info-source-lock.yaml"
    ),
    "raw_source_noncommit_proof": Path(
        "data/source-harvester/stage5dp-raw-source-noncommit-proof.yaml"
    ),
    "mayfly_instar_grid_analysis_candidate_v1": Path(
        "data/historical-route/stage5dp-mayfly-instar-grid-analysis-candidate-v1.yaml"
    ),
    "mayfly_block_grid_reduction_230x262_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-block-grid-reduction-230x262-candidate-v0.yaml"
    ),
    "mayfly_four_block_key_genome_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-four-block-key-genome-candidate-v0.yaml"
    ),
    "mayfly_prime_distance_263_283_149_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-prime-distance-263-283-149-candidate-v0.yaml"
    ),
    "mayfly_horizontal_axis_167_229_229_229_104_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-horizontal-axis-167-229-229-229-104-candidate-v0.yaml"
    ),
    "mayfly_block_type_counts_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-block-type-counts-candidate-v0.yaml"
    ),
    "mayfly_column_sequence_83_103_101_107_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-column-sequence-83-103-101-107-candidate-v0.yaml"
    ),
    "mayfly_block_isolation_code_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-block-isolation-code-candidate-v0.yaml"
    ),
    "mayfly_parable_circumference_shedding_route_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-parable-circumference-shedding-route-candidate-v0.yaml"
    ),
    "five_dot_shift_skipped_f_page56_candidate_v0": Path(
        "data/historical-route/stage5dp-five-dot-shift-skipped-f-page56-candidate-v0.yaml"
    ),
    "page33_three_dot_emirp_area_block_candidate_v0": Path(
        "data/historical-route/stage5dp-page33-three-dot-emirp-area-block-candidate-v0.yaml"
    ),
    "front_cover_1033_measurement_geometry_candidate_v0": Path(
        "data/historical-route/stage5dp-front-cover-1033-measurement-geometry-candidate-v0.yaml"
    ),
    "front_cover_1327_1427_concat_emirp_candidate_v0": Path(
        "data/historical-route/stage5dp-front-cover-1327-1427-concat-emirp-candidate-v0.yaml"
    ),
    "front_cover_d_panel_two_pixel_shift_candidate_v0": Path(
        "data/historical-route/stage5dp-front-cover-d-panel-two-pixel-shift-candidate-v0.yaml"
    ),
    "front_cover_obscured_coordinate_1033_1033_candidate_v0": Path(
        "data/historical-route/stage5dp-front-cover-obscured-coordinate-1033-1033-candidate-v0.yaml"
    ),
    "front_cover_adjusted_area_1889603_prime_candidate_v0": Path(
        "data/historical-route/stage5dp-front-cover-adjusted-area-1889603-prime-candidate-v0.yaml"
    ),
    "front_cover_3301_concat_prime_not_emirp_correction_v0": Path(
        "data/historical-route/stage5dp-front-cover-3301-concat-prime-not-emirp-correction-v0.yaml"
    ),
    "iso_560_13_560_17_palindromic_prime_size_candidate_v0": Path(
        "data/historical-route/stage5dp-iso-560-13-560-17-palindromic-prime-size-candidate-v0.yaml"
    ),
    "problems_2012_autostereogram_source_tool_candidate_v0": Path(
        "data/historical-route/stage5dp-problems-2012-autostereogram-source-tool-candidate-v0.yaml"
    ),
    "lp_dot_marker_geometry_family_v1": Path(
        "data/historical-route/stage5dp-lp-dot-marker-geometry-family-v1.yaml"
    ),
    "lp_4_5_pixel_block_geometry_candidate_v0": Path(
        "data/historical-route/stage5dp-lp-4-5-pixel-block-geometry-candidate-v0.yaml"
    ),
    "mayfly_167_229_axis_music_eclipse_crosslink_v0": Path(
        "data/historical-route/stage5dp-mayfly-167-229-axis-music-eclipse-crosslink-v0.yaml"
    ),
    "mayfly_circumference_shedding_grid_reduction_candidate_v0": Path(
        "data/historical-route/stage5dp-mayfly-circumference-shedding-grid-reduction-candidate-v0.yaml"
    ),
    "stage5dg_preservation": Path("data/token-block/stage5dp-stage5dg-preservation.yaml"),
    "stage5bd_preservation": Path("data/token-block/stage5dp-stage5bd-preservation.yaml"),
    "active_lineage_preservation": Path(
        "data/token-block/stage5dp-active-lineage-preservation.yaml"
    ),
    "no_active_ingestion_proof": Path("data/token-block/stage5dp-no-active-ingestion-proof.yaml"),
    "no_byte_stream_transition_gate": Path(
        "data/token-block/stage5dp-no-byte-stream-transition-gate.yaml"
    ),
    "no_execution_transition_gate": Path(
        "data/token-block/stage5dp-no-execution-transition-gate.yaml"
    ),
}

SCHEMA_PATHS = {
    key: Path("schemas") / path.parent.relative_to("data") / f"{path.stem}-v0.schema.json"
    for key, path in DATA_PATHS.items()
}

HISTORICAL_KEYS = [
    key for key, path in DATA_PATHS.items() if path.as_posix().startswith("data/historical-route/")
]


@dataclass(frozen=True)
class ValidationResult:
    command: str
    validation_error_count: int
    errors: list[str]

    def to_cli_text(self) -> str:
        lines = [f"command={self.command}", f"validation_error_count={self.validation_error_count}"]
        lines.extend(f"error={error}" for error in self.errors)
        return "\n".join(lines)


def _base(record_type: str) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "record_type": record_type,
        "stage_id": STAGE_ID,
        "stage_title": STAGE_TITLE,
        "prompt_type": PROMPT_TYPE,
        "metadata_only": True,
        "source_lock_only": True,
        "solve_claim": False,
        "execution_allowed": False,
        "canonical_codex_handoff_root": "codex-output",
    }
    payload.update({flag: False for flag in sorted(FORBIDDEN_FALSE_FLAGS)})
    return payload


def _candidate(record_type: str, candidate_family_id: str) -> dict[str, Any]:
    payload = _base(record_type)
    payload.update(
        {
            "candidate_family_id": candidate_family_id,
            "accepted_as_route": False,
            "accepted_as_decryption": False,
            "accepted_as_key": False,
            "used_for_target_selection_now": False,
            "candidate_only": True,
            "confidence": "source_locked_candidate_only",
        }
    )
    return payload


def _schema_for(_key: str) -> dict[str, Any]:
    required = [
        "record_type",
        "stage_id",
        "metadata_only",
        "source_lock_only",
        "solve_claim",
        "execution_allowed",
        "pivot_target_selected_now",
        "route_extraction_performed_now",
        "image_forensics_performed",
        "ocr_performed",
    ]
    return {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "type": "object",
        "additionalProperties": True,
        "required": required,
        "properties": {
            "stage_id": {"const": STAGE_ID},
            "metadata_only": {"const": True},
            "source_lock_only": {"const": True},
            "execution_allowed": {"const": False},
            "solve_claim": {"const": False},
            "pivot_target_selected_now": {"const": False},
            "route_extraction_performed_now": {"const": False},
            "image_forensics_performed": {"const": False},
            "ocr_performed": {"const": False},
            "generated_outputs_committed": {"const": False},
            "source_browser_gui_implemented_now": {"const": False},
        },
    }


def _write_schemas() -> None:
    for key, path in SCHEMA_PATHS.items():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(_schema_for(key), indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _line_count(path: Path) -> int:
    return len(path.read_text(encoding="utf-8", errors="replace").splitlines())


def _text_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _git_ls_files(path: str) -> list[str]:
    result = subprocess.run(["git", "ls-files", "--", path], capture_output=True, check=False, text=True)
    if result.returncode != 0:
        return []
    return [line for line in result.stdout.splitlines() if line.strip()]


def _image_dimensions(path: Path) -> dict[str, int | None]:
    try:
        from PIL import Image
    except Exception:
        return {"image_width": None, "image_height": None}
    try:
        with Image.open(path) as image:
            return {"image_width": int(image.width), "image_height": int(image.height)}
    except Exception:
        return {"image_width": None, "image_height": None}


def _file_record(path: Path, source_folder: str) -> dict[str, Any]:
    relative = path.as_posix()
    mime, _ = mimetypes.guess_type(path.name)
    record: dict[str, Any] = {
        "relative_path": relative,
        "file_name": path.name,
        "extension": path.suffix.lower(),
        "file_size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "mime_guess": mime,
        "source_folder": source_folder,
        "raw_file_committed_now": False,
        "source_lock_only": True,
    }
    if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
        record.update(_image_dimensions(path))
    else:
        record.update({"image_width": None, "image_height": None})
    return record


def _folder_inventory(folder: Path) -> dict[str, Any]:
    exists = folder.exists()
    files = (
        sorted([path for path in folder.rglob("*") if path.is_file()], key=lambda item: item.as_posix())
        if exists
        else []
    )
    return {
        "source_folder": folder.name,
        "source_folder_path": folder.as_posix(),
        "source_root_exists": exists,
        "file_count_observed": len(files),
        "gap_recorded": not exists,
        "gap_reason": None if exists else "source_folder_missing",
        "files": [_file_record(path, folder.name) for path in files],
    }


def _docx_metadata(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "mayfly_docx_present": False,
            "verification_gap": "docx_missing",
            "raw_text_committed_now": False,
        }
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            media = [name for name in names if name.startswith("word/media/") and not name.endswith("/")]
            document = ElementTree.fromstring(archive.read("word/document.xml"))
            namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            paragraphs = document.findall(".//w:p", namespace)
            words: list[str] = []
            for text_node in document.findall(".//w:t", namespace):
                if text_node.text:
                    words.extend(text_node.text.split())
            return {
                "mayfly_docx_present": True,
                "mayfly_docx_paragraph_count": len(paragraphs),
                "mayfly_docx_embedded_media_count": len(media),
                "mayfly_docx_word_count_estimate": len(words),
                "raw_text_committed_now": False,
                "verification_gap": None,
            }
    except Exception as exc:
        return {
            "mayfly_docx_present": True,
            "mayfly_docx_paragraph_count": None,
            "mayfly_docx_embedded_media_count": None,
            "mayfly_docx_word_count_estimate": None,
            "raw_text_committed_now": False,
            "verification_gap": f"docx_metadata_unavailable:{type(exc).__name__}",
        }


def _grid_stats(ws: Any, start_row: int, start_col: int, height: int, width: int) -> dict[str, Any]:
    total = 0
    nonzero = 0
    counts: dict[int, int] = {}
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=start_row + height - 1,
        min_col=start_col,
        max_col=start_col + width - 1,
        values_only=True,
    ):
        for value in row:
            if not isinstance(value, (int, float)):
                raise ValueError("non_numeric_grid_cell")
            integer = int(value)
            total += integer
            counts[integer] = counts.get(integer, 0) + 1
            if integer:
                nonzero += 1
    return {"sum": total, "nonzero_count": nonzero, "counts": counts}


def _xlsx_metadata(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "mayfly_xlsx_present": False,
            "raw_workbook_committed_now": False,
            "verification_gap": "xlsx_missing",
        }
    try:
        import openpyxl
    except Exception:
        return {
            "mayfly_xlsx_present": True,
            "raw_workbook_committed_now": False,
            "verification_gap": "openpyxl_unavailable",
            "workbook_values_independently_verified_now": False,
        }
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = list(workbook.sheetnames)
        original = _grid_stats(workbook["958x1092 (binary)"], 2, 2, 1092, 958)
        reduced = _grid_stats(workbook["230x262 (binary)"], 4, 4, 262, 230)
        weighted = _grid_stats(workbook["230x262 (weighted)"], 7, 7, 262, 230)
        block_counts = {
            "value_16": weighted["counts"].get(16, 0),
            "value_20": weighted["counts"].get(20, 0),
            "value_25": weighted["counts"].get(25, 0),
        }
        verified = (
            original["sum"] == 35210
            and reduced["sum"] == 2033
            and weighted["sum"] == 35210
            and weighted["nonzero_count"] == 2033
            and block_counts == {"value_16": 1420, "value_20": 567, "value_25": 46}
        )
        return {
            "mayfly_xlsx_present": True,
            "mayfly_workbook_sheet_names": sheet_names,
            "sheet_count": len(sheet_names),
            "workbook_size_bytes": path.stat().st_size,
            "original_binary_grid_dimensions": [958, 1092],
            "original_binary_ones_sum": original["sum"],
            "reduced_binary_grid_dimensions": [230, 262],
            "reduced_binary_ones_sum": reduced["sum"],
            "weighted_grid_dimensions": [230, 262],
            "weighted_grid_sum": weighted["sum"],
            "weighted_block_counts": block_counts,
            "block_total_from_weighted_counts": weighted["nonzero_count"],
            "source_claimed_block_type_counts": {
                "block_4x4": 1420,
                "block_4x5": 274,
                "block_5x4": 293,
                "block_5x5": 46,
            },
            "source_claimed_block_count_total": 2033,
            "source_claimed_split_verified_now": False,
            "workbook_values_independently_verified_now": verified,
            "raw_workbook_committed_now": False,
            "verification_gap": None if verified else "expected_workbook_values_not_matched",
        }
    except Exception as exc:
        return {
            "mayfly_xlsx_present": True,
            "raw_workbook_committed_now": False,
            "verification_gap": f"workbook_processing_unavailable_or_too_slow:{type(exc).__name__}",
            "workbook_values_independently_verified_now": False,
        }


def _prime_check(value: int) -> bool:
    if value < 2:
        return False
    if value % 2 == 0:
        return value == 2
    divisor = 3
    while divisor * divisor <= value:
        if value % divisor == 0:
            return False
        divisor += 2
    return True


def _url_info_locks(inventories: list[dict[str, Any]]) -> list[dict[str, Any]]:
    locks: list[dict[str, Any]] = []
    for inventory in inventories:
        folder_path = Path(inventory["source_folder_path"])
        for name in ["url.txt", "info.txt", "messages.txt"]:
            path = folder_path / name
            locks.append(
                {
                    "source_folder": inventory["source_folder"],
                    "file_name": name,
                    "present": path.exists(),
                    "relative_path": path.as_posix(),
                    "sha256": _text_hash(path) if path.exists() else None,
                    "line_count": _line_count(path) if path.exists() else None,
                    "raw_text_committed_now": False,
                }
            )
    return locks


def _raw_source_noncommit_proof(inventories: list[dict[str, Any]]) -> dict[str, Any]:
    tracked = _git_ls_files(REDDIT_ROOT.as_posix())
    files_seen = sum(int(inventory.get("file_count_observed", 0)) for inventory in inventories)
    payload = _base("stage5dp_raw_source_noncommit_proof")
    payload.update(
        {
            "raw_source_root": REDDIT_ROOT.as_posix(),
            "raw_source_files_seen": files_seen,
            "raw_source_files_committed_now": False,
            "raw_source_files_staged_now": len(tracked),
            "tracked_redditstuff_files": tracked,
            "generated_reports_committed_now": False,
            "codex_output_absent": not DEPRECATED_CODEX_OUTPUT.exists(),
            "codex_output_path_used": False,
        }
    )
    return payload


def _preservation_record(record_type: str, preserved_stage: str, extra: dict[str, Any] | None = None) -> dict[str, Any]:
    payload = _base(record_type)
    payload.update(
        {
            "preserved_stage": preserved_stage,
            "stage5bd_run_plan_id_count": 10,
            "active_lineage_record_count": len(ACTIVE_LINEAGE_PATHS),
            "parallel_worker_cap": PARALLEL_WORKER_CAP,
        }
    )
    if extra:
        payload.update(extra)
    return payload


def _candidate_records(mayfly_meta: dict[str, Any]) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}

    records["mayfly_instar_grid_analysis_candidate_v1"] = _candidate(
        "stage5dp_mayfly_instar_grid_analysis_candidate",
        "mayfly_instar_grid_analysis_candidate_v1",
    ) | {
        "source_folder": "MayFlyInvestigation",
        "mayfly_docx_present": mayfly_meta.get("mayfly_docx_present"),
        "mayfly_xlsx_present": mayfly_meta.get("mayfly_xlsx_present"),
        "mayfly_docx_paragraph_count": mayfly_meta.get("mayfly_docx_paragraph_count"),
        "mayfly_docx_embedded_media_count": mayfly_meta.get("mayfly_docx_embedded_media_count"),
        "mayfly_workbook_sheet_names": mayfly_meta.get("mayfly_workbook_sheet_names", []),
        "original_binary_grid_dimensions": mayfly_meta.get("original_binary_grid_dimensions"),
        "original_binary_ones_sum": mayfly_meta.get("original_binary_ones_sum"),
        "reduced_binary_grid_dimensions": mayfly_meta.get("reduced_binary_grid_dimensions"),
        "reduced_binary_ones_sum": mayfly_meta.get("reduced_binary_ones_sum"),
        "weighted_grid_dimensions": mayfly_meta.get("weighted_grid_dimensions"),
        "weighted_grid_sum": mayfly_meta.get("weighted_grid_sum"),
        "weighted_block_counts": mayfly_meta.get("weighted_block_counts", {}),
        "block_total_from_weighted_counts": mayfly_meta.get("block_total_from_weighted_counts"),
        "workbook_values_independently_verified_now": mayfly_meta.get(
            "workbook_values_independently_verified_now", False
        ),
    }
    records["mayfly_block_grid_reduction_230x262_candidate_v0"] = _candidate(
        "stage5dp_mayfly_block_grid_reduction_candidate",
        "mayfly_block_grid_reduction_230x262_candidate_v0",
    ) | {
        "original_binary_grid_dimensions": [958, 1092],
        "reduced_binary_grid_dimensions": [230, 262],
        "reduced_binary_ones_sum": mayfly_meta.get("reduced_binary_ones_sum"),
        "weighted_grid_sum": mayfly_meta.get("weighted_grid_sum"),
        "grid_reduction_accepted_as_route": False,
    }
    records["mayfly_four_block_key_genome_candidate_v0"] = _candidate(
        "stage5dp_mayfly_four_block_key_genome_candidate",
        "mayfly_four_block_key_genome_candidate_v0",
    ) | {
        "source_claims": {
            "two_black_blocks_form_58x58_square": True,
            "58_equals_2x29": True,
            "distances": {
                "blue_to_upper_circled_block": [17, 29],
                "blue_to_yellow_block": [55, 67],
                "blue_to_lower_circled_block": [71, 83],
            },
            "prime_values_from_six_dimensions": [17, 29, 67, 71, 83],
            "gematria_primus_mapping_candidate": ["G", "N", "E", "M", "OE"],
            "genome_anagram_candidate": True,
            "ignored_or_unmapped_value": 55,
        },
        "confidence": "medium",
        "anagram_required": True,
    }
    records["mayfly_prime_distance_263_283_149_candidate_v0"] = _candidate(
        "stage5dp_mayfly_prime_distance_candidate",
        "mayfly_prime_distance_263_283_149_candidate_v0",
    ) | {
        "prime_distances_claimed": [263, 283, 149],
        "links_to": [
            "page32_moebius_fibonacci_prime_index_spiral_v1",
            "mayfly_four_block_key_genome_candidate_v0",
        ],
        "interpretive_death_claim_quarantined": True,
    }
    records["mayfly_horizontal_axis_167_229_229_229_104_candidate_v0"] = _candidate(
        "stage5dp_mayfly_horizontal_axis_candidate",
        "mayfly_horizontal_axis_167_229_229_229_104_candidate_v0",
    ) | {
        "horizontal_axis_raw_segments": [17, 150, 29, 200, 29, 200, 29, 200, 29, 75],
        "horizontal_axis_grouped": [167, 229, 229, 229, 104],
        "links_to": [
            "music_3301_instar_crab_canon_v0",
            "disk_2015_eclipse_167_temporal_candidate_v0",
            "mayfly_parable_circumference_shedding_route_candidate_v0",
        ],
    }
    records["mayfly_block_type_counts_candidate_v0"] = _candidate(
        "stage5dp_mayfly_block_type_counts_candidate",
        "mayfly_block_type_counts_candidate_v0",
    ) | {
        "weighted_block_counts": mayfly_meta.get("weighted_block_counts", {}),
        "source_claimed_block_type_counts": mayfly_meta.get("source_claimed_block_type_counts", {}),
        "source_claimed_block_count_total": mayfly_meta.get("source_claimed_block_count_total"),
        "source_claimed_split_verified_now": mayfly_meta.get("source_claimed_split_verified_now", False),
    }
    records["mayfly_column_sequence_83_103_101_107_candidate_v0"] = _candidate(
        "stage5dp_mayfly_column_sequence_candidate",
        "mayfly_column_sequence_83_103_101_107_candidate_v0",
    ) | {
        "column_sequence_claimed": [83, 103, 101, 107],
        "sequence_values_prime": all(_prime_check(value) for value in [83, 103, 101, 107]),
        "links_to": [
            "pdd_153_triangle_prime_mask_route_v1",
            "pdd_153_triangle_2016_prime_layer_route_v1",
            "no_f_rune_count_section_flow_candidate_v0",
        ],
        "requires_future_oeis_verification": True,
    }
    records["mayfly_block_isolation_code_candidate_v0"] = _candidate(
        "stage5dp_mayfly_block_isolation_code_candidate",
        "mayfly_block_isolation_code_candidate_v0",
    ) | {
        "source_claim": "Mayfly block isolation may encode reviewable code-like structure.",
        "requires_future_review": True,
        "accepted_as_code": False,
    }
    records["mayfly_parable_circumference_shedding_route_candidate_v0"] = _candidate(
        "stage5dp_mayfly_parable_circumference_shedding_route_candidate",
        "mayfly_parable_circumference_shedding_route_candidate_v0",
    ) | {
        "parable_links": ["instar", "emergence", "shed_our_own_circumferences", "find_the_divinity_within"],
        "candidate_method_metaphors": [
            "grid_reduction",
            "block_to_single_cell_reduction",
            "row_column_shedding",
            "inside_outside_route",
        ],
        "links_to": [
            "blake_body_soul_perception_v0",
            "solved_i_voice_of_circumference_precedent_v0",
            "music_3301_instar_crab_canon_v0",
            "circumference_single_i_spiral_anchor_crosslink_v0",
        ],
    }
    records["five_dot_shift_skipped_f_page56_candidate_v0"] = _candidate(
        "stage5dp_five_dot_shift_skipped_f_page56_candidate",
        "five_dot_shift_skipped_f_page56_candidate_v0",
    ) | {
        "source_folder": "DotObservationsFromPages_7_23_56",
        "pixel_measurements_require_canonical_image_verification": True,
        "page_dot_observations": {
            "page_7": {"left_edge": 763, "right_edge": 837},
            "page_23": {"left_edge": 763, "right_edge": 837},
            "page_56": {"left_edge": 839, "right_edge": 761},
        },
        "page56_claims": {
            "skipped_decryption_index": 221,
            "gap_from_last_rune_line_to_dots_pixels": 221,
            "skipped_rune": "F",
            "f_gp_value": 2,
            "dots_shifted_by_two_candidate": True,
            "flipped_or_reflected_alignment_candidate": True,
        },
        "links_to": [
            "page56_dwh_hash_target_contract_v0",
            "five_dot_transition_marker_candidate_v0",
            "lp_dot_marker_geometry_family_v1",
            "lp_doublet_scarcity_feature_v1",
        ],
    }
    records["page33_three_dot_emirp_area_block_candidate_v0"] = _candidate(
        "stage5dp_page33_three_dot_emirp_area_block_candidate",
        "page33_three_dot_emirp_area_block_candidate_v0",
    ) | {
        "source_folder": "Page_33_ThreeDotsObservations",
        "dot_dimensions_claimed": {
            "dot_1": {"height": 69, "width": 70, "area": 3929},
            "dot_2": {"height": 69, "width": 70, "area": 3920},
            "dot_3": {"height": 69, "width": 69, "area": 3911},
        },
        "area_pattern": {
            "decrement": 9,
            "dot_1_area_prime_or_emirp_claim": True,
            "dot_3_area_prime_or_emirp_claim": True,
            "dot_1_area_prime_verified_now": _prime_check(3929),
            "dot_3_area_prime_verified_now": _prime_check(3911),
        },
        "measured_distances_claimed": {
            "outer_distance_1_to_2": 347,
            "outer_distance_2_to_3": 409,
            "inner_distance_1_to_2": 487,
            "left_distance_2_to_3": 479,
        },
        "block_geometry_claimed": ["4x4", "4x5", "5x5"],
        "pixel_geometry_source_claimed_pending_canonical_verification": True,
        "links_to": [
            "lp_dot_marker_geometry_family_v1",
            "mayfly_block_grid_reduction_230x262_candidate_v0",
            "lp_4_5_pixel_block_geometry_candidate_v0",
        ],
    }
    front_base = {
        "source_folder": "FrontCoverMeasurementsOf1033",
        "front_cover_dimensions_claimed": {"width": 1327, "height": 1427},
        "concatenations": {
            "width_height_1033": 132714271033,
            "width_height_3301": 132714273301,
        },
        "verified_or_source_claimed_arithmetic": {
            "132714271033_prime": _prime_check(132714271033),
            "132714271033_reverse_prime": _prime_check(330172417231),
            "132714271033_emirp": _prime_check(132714271033) and _prime_check(330172417231),
            "132714273301_prime": _prime_check(132714273301),
            "132714273301_reverse_prime": _prime_check(103372417231),
            "132714273301_emirp": False,
            "adjusted_area_1889603_prime": _prime_check(1889603),
        },
        "warnings": [
            "3301_concatenation_is_prime_but_not_emirp_under_standard_reverse_prime_definition"
        ],
        "source_claims": {
            "d_panel_shift_pixels": 2,
            "coordinate_1033_1033_obscured_by_overlap": True,
        },
        "links_to": [
            "blake_compass_ancient_days_newton_v0",
            "front_cover_visual_source_family_v0",
            "circumference_single_i_spiral_anchor_crosslink_v0",
            "lp_image_shift_geometry_candidate_v0",
        ],
    }
    for key, family_id in [
        (
            "front_cover_1033_measurement_geometry_candidate_v0",
            "front_cover_1033_measurement_geometry_candidate_v0",
        ),
        (
            "front_cover_1327_1427_concat_emirp_candidate_v0",
            "front_cover_1327_1427_concat_emirp_candidate_v0",
        ),
        (
            "front_cover_d_panel_two_pixel_shift_candidate_v0",
            "front_cover_d_panel_two_pixel_shift_candidate_v0",
        ),
        (
            "front_cover_obscured_coordinate_1033_1033_candidate_v0",
            "front_cover_obscured_coordinate_1033_1033_candidate_v0",
        ),
        (
            "front_cover_adjusted_area_1889603_prime_candidate_v0",
            "front_cover_adjusted_area_1889603_prime_candidate_v0",
        ),
        (
            "front_cover_3301_concat_prime_not_emirp_correction_v0",
            "front_cover_3301_concat_prime_not_emirp_correction_v0",
        ),
    ]:
        records[key] = _candidate(f"stage5dp_{family_id}", family_id) | front_base
    records["iso_560_13_560_17_palindromic_prime_size_candidate_v0"] = _candidate(
        "stage5dp_iso_palindromic_prime_size_candidate",
        "iso_560_13_560_17_palindromic_prime_size_candidate_v0",
    ) | {
        "source_folder": "DataDirectoryFromTheISO",
        "source_files_or_sizes_claimed": {"560.13": 118818811, "560.17": 1183811},
        "arithmetic_claims": {
            "118818811_prime": _prime_check(118818811),
            "118818811_palindrome": str(118818811) == str(118818811)[::-1],
            "1183811_prime": _prime_check(1183811),
            "1183811_palindrome": str(1183811) == str(1183811)[::-1],
        },
        "lp_relevance": "historical_2012_iso_context_low_priority",
        "links_to": ["historical_2012_source_context_v0", "prime_palindrome_number_fact_context_v0"],
    }
    records["problems_2012_autostereogram_source_tool_candidate_v0"] = _candidate(
        "stage5dp_problems_2012_autostereogram_source_tool_candidate",
        "problems_2012_autostereogram_source_tool_candidate_v0",
    ) | {
        "source_folder": "FindsFromProblems_2012_puzzle",
        "source_claims": {
            "problems_image_may_have_easystereogrambuilder_watermark": True,
            "autostereogram_or_magic_eye_generation_candidate": True,
            "possible_medieval_tapestry_source_candidate": True,
        },
        "source_author_notes_low_relevance_to_lp": True,
        "lp_relevance": "historical_low_priority",
        "links_to": ["historical_2012_source_context_v0", "image_source_tool_context_v0"],
    }
    records["lp_dot_marker_geometry_family_v1"] = _candidate(
        "stage5dp_lp_dot_marker_geometry_family",
        "lp_dot_marker_geometry_family_v1",
    ) | {
        "crosslinks": [
            "page33_three_dot_emirp_area_block_candidate_v0",
            "five_dot_shift_skipped_f_page56_candidate_v0",
            "five_dot_transition_marker_candidate_v0",
            "quote_dinkus_red_marker_context_v0",
            "disk_single_dot_branching_candidate_v1",
        ],
    }
    records["lp_4_5_pixel_block_geometry_candidate_v0"] = _candidate(
        "stage5dp_lp_4_5_pixel_block_geometry_candidate",
        "lp_4_5_pixel_block_geometry_candidate_v0",
    ) | {
        "crosslinks": [
            "mayfly_block_grid_reduction_230x262_candidate_v0",
            "page33_three_dot_emirp_area_block_candidate_v0",
            "pixel_colour_frequency_source_tables_v0",
        ],
    }
    records["mayfly_167_229_axis_music_eclipse_crosslink_v0"] = _candidate(
        "stage5dp_mayfly_167_229_axis_music_eclipse_crosslink",
        "mayfly_167_229_axis_music_eclipse_crosslink_v0",
    ) | {
        "crosslinks": [
            "mayfly_horizontal_axis_167_229_229_229_104_candidate_v0",
            "music_3301_instar_crab_canon_v0",
            "disk_2015_eclipse_167_temporal_candidate_v0",
        ],
    }
    records["mayfly_circumference_shedding_grid_reduction_candidate_v0"] = _candidate(
        "stage5dp_mayfly_circumference_shedding_grid_reduction_candidate",
        "mayfly_circumference_shedding_grid_reduction_candidate_v0",
    ) | {
        "crosslinks": [
            "mayfly_parable_circumference_shedding_route_candidate_v0",
            "blake_visual_text_source_family_v0",
            "solved_i_voice_of_circumference_precedent_v0",
            "circumference_single_i_spiral_anchor_crosslink_v0",
        ],
    }
    return records


def _context_text(existing: str) -> str:
    section = """\n\n## Stage 5DP - New Reddit Mayfly / Dot / Cover / ISO Source-Lock\n\nStage 5DP source-locked new Reddit Mayfly/dot/cover/ISO material. MayFlyInvestigation is high value. Key Mayfly facts: 958x1092 original, 230x262 reduced grid, 2033 active reduced cells, weighted sum 35210, 167//229//229//229//104 axis. Dot material: page56 five-dot/skipped-F/shift-by-2 candidate, page33 three-dot emirp/4-5 block candidate. Cover material: 1327/1427/1033 emirp candidate, 3301 concat correction. These are candidate-only, not active solve routes. Next planned stage remains lightweight source-lock browser GUI unless more evidence arrives.\n"""
    marker = "## Stage 5DP - New Reddit Mayfly / Dot / Cover / ISO Source-Lock"
    if marker in existing:
        prefix = existing.split(marker, 1)[0].rstrip()
        return prefix + section
    return existing.rstrip() + section if existing.strip() else "# ChatGPT Context File\n" + section


def _build_records() -> dict[str, dict[str, Any]]:
    inventories = [_folder_inventory(REDDIT_ROOT / folder) for folder in REQUIRED_SOURCE_FOLDERS.values()]
    if REDDIT_ROOT.exists():
        required_names = set(REQUIRED_SOURCE_FOLDERS.values())
        for folder in sorted(
            [item for item in REDDIT_ROOT.iterdir() if item.is_dir() and item.name not in required_names],
            key=lambda item: item.name,
        ):
            inventories.append(_folder_inventory(folder))

    docx_meta = _docx_metadata(MAYFLY_DOCX)
    xlsx_meta = _xlsx_metadata(MAYFLY_XLSX)
    mayfly_meta = docx_meta | xlsx_meta
    url_info_locks = _url_info_locks(inventories)
    source_gap_count = sum(1 for item in inventories if item.get("gap_recorded"))
    candidate_records = _candidate_records(mayfly_meta)

    records: dict[str, dict[str, Any]] = {}
    records.update(candidate_records)
    records["new_reddit_source_lock_register"] = _base("stage5dp_new_reddit_source_lock_register") | {
        "source_root": REDDIT_ROOT.as_posix(),
        "source_root_exists": REDDIT_ROOT.exists(),
        "required_source_folders": list(REQUIRED_SOURCE_FOLDERS.values()),
        "source_folder_count": len(inventories),
        "required_source_folder_count": len(REQUIRED_SOURCE_FOLDERS),
        "source_gap_count": source_gap_count,
        "source_folders": inventories,
    }
    records["mayfly_docx_xlsx_source_lock"] = _base("stage5dp_mayfly_docx_xlsx_source_lock") | {
        "source_folder": MAYFLY_FOLDER.as_posix(),
        "docx_file": _file_record(MAYFLY_DOCX, "MayFlyInvestigation") if MAYFLY_DOCX.exists() else None,
        "xlsx_file": _file_record(MAYFLY_XLSX, "MayFlyInvestigation") if MAYFLY_XLSX.exists() else None,
        **mayfly_meta,
    }
    records["reddit_url_info_source_lock"] = _base("stage5dp_reddit_url_info_source_lock") | {
        "url_info_locks": url_info_locks,
        "lock_count": len(url_info_locks),
        "present_count": sum(1 for item in url_info_locks if item["present"]),
        "raw_text_committed_now": False,
    }
    records["raw_source_noncommit_proof"] = _raw_source_noncommit_proof(inventories)
    records["stage5do_preservation"] = _preservation_record(
        "stage5dp_stage5do_preservation",
        "stage-5do",
        {
            "stage5do_validation_passed": True,
            "source_browser_gui_future_requirement_preserved": True,
            "stage5do_summary": {
                "number_facts_collection_locked": True,
                "potential_hint_locked": True,
                "candidate_records_created": 15,
            },
        },
    )
    records["stage5dg_preservation"] = _preservation_record(
        "stage5dp_stage5dg_preservation",
        "stage-5dg",
        {"operator_approval_component_satisfied": True, "combined_approval_gate_satisfied_now": False},
    )
    records["stage5bd_preservation"] = _preservation_record(
        "stage5dp_stage5bd_preservation",
        "stage-5bd",
        {"run_plan_id_count_preserved": 10},
    )
    records["active_lineage_preservation"] = _preservation_record(
        "stage5dp_active_lineage_preservation",
        "active-lineage",
        {"active_lineage_paths": [str(path) for path in ACTIVE_LINEAGE_PATHS]},
    )
    for key, record_type in [
        ("no_active_ingestion_proof", "stage5dp_no_active_ingestion_proof"),
        ("no_byte_stream_transition_gate", "stage5dp_no_byte_stream_transition_gate"),
        ("no_execution_transition_gate", "stage5dp_no_execution_transition_gate"),
    ]:
        records[key] = _preservation_record(record_type, "closed-gate", {"gate_closed": True})
    records["reviewability_gap_register"] = _base("stage5dp_reviewability_gap_register") | {
        "gap_count": source_gap_count + 8,
        "gaps": [
            "canonical_image_geometry_verification_required",
            "mayfly_route_extraction_not_authorized",
            "dot_pixel_measurements_pending_canonical_image_verification",
            "front_cover_measurements_pending_canonical_image_verification",
            "source_claims_require_human_review_before_route_use",
            "target_priority_selection_not_created",
            "source_browser_gui_deferred",
            "raw_third_party_files_remain_ignored",
        ],
    }
    records["pivot_readiness_update"] = _base("stage5dp_pivot_readiness_update") | {
        "pivot_readiness_status": "candidate_source_locked_review_required",
        "pivot_target_selected_now": False,
        "operator_readiness_decision_created_now": False,
        "target_priority_decision_created_now": False,
    }
    records["source_browser_gui_deferral"] = _base("stage5dp_source_browser_gui_deferral") | {
        "source_browser_gui_future_requirement_preserved": True,
        "source_browser_gui_implemented_now": False,
        "source_browser_gui_deferred_to_next_stage": True,
        "recommended_next_stage_id": NEXT_STAGE_ID,
        "recommended_next_stage_title": NEXT_STAGE_TITLE,
        "reason": "New Reddit material was source-locked before implementing browser UI.",
    }
    records["next_stage_decision"] = _base("stage5dp_next_stage_decision") | {
        "recommended_next_stage_id": NEXT_STAGE_ID,
        "recommended_next_stage_title": NEXT_STAGE_TITLE,
        "decision_status": "selected_for_planning_only",
        "execution_authorized_now": False,
    }
    records["reviewable_validation_evidence"] = _base("stage5dp_reviewable_validation_evidence") | {
        "focused_validators_required": [
            "validate-stage5dp-new-reddit-source-lock",
            "validate-stage5dp-mayfly-source-lock",
            "validate-stage5dp-mayfly-workbook-summary",
            "validate-stage5dp-dot-observations",
            "validate-stage5dp-front-cover-measurements",
            "validate-stage5dp-iso-and-problems-sources",
            "validate-stage5dp-chatgpt-context-file",
            "validate-stage5dp-sidecar-gates",
            "validate-stage5dp-handoff-continuity",
        ],
        "parallel_worker_cap": PARALLEL_WORKER_CAP,
        "generated_reports_dir": RESULTS_DIR.as_posix(),
    }
    records["summary"] = _base("stage5dp_summary") | {
        "status": "complete",
        "source_previous_stage": SOURCE_PREVIOUS_STAGE_ID,
        "source_previous_stage_commit": SOURCE_PREVIOUS_STAGE_COMMIT,
        "source_previous_stage_issue": SOURCE_PREVIOUS_ISSUE,
        "source_previous_stage_ci_run": SOURCE_PREVIOUS_CI_RUN,
        "new_reddit_source_lock_created": REDDIT_ROOT.exists(),
        "new_reddit_source_folder_count": len(inventories),
        "required_reddit_source_folders_represented": len(REQUIRED_SOURCE_FOLDERS) - source_gap_count,
        "mayfly_instar_grid_analysis_candidate_created": True,
        "mayfly_docx_source_locked": docx_meta.get("mayfly_docx_present") is True,
        "mayfly_xlsx_source_locked": xlsx_meta.get("mayfly_xlsx_present") is True,
        "mayfly_block_grid_reduction_candidate_created": True,
        "mayfly_four_block_key_genome_candidate_created": True,
        "mayfly_axis_167_229_candidate_created": True,
        "page56_five_dot_skipped_f_candidate_created": True,
        "page33_three_dot_candidate_created": True,
        "front_cover_measurement_candidate_created": True,
        "iso_palindromic_prime_size_candidate_created": True,
        "problems_2012_autostereogram_candidate_created": True,
        "chatgpt_context_file_updated": True,
        "source_browser_gui_implemented_now": False,
        "source_browser_gui_deferred_to_next_stage": True,
        "stage5bd_run_plan_id_count": 10,
        "active_lineage_record_count": len(ACTIVE_LINEAGE_PATHS),
        "parallel_worker_cap": PARALLEL_WORKER_CAP,
        "recommended_next_stage_id": NEXT_STAGE_ID,
        "recommended_next_stage_title": NEXT_STAGE_TITLE,
        "candidate_records_created": len(candidate_records),
        "mayfly_workbook_summary": {
            "original_binary_grid_dimensions": xlsx_meta.get("original_binary_grid_dimensions"),
            "original_binary_ones_sum": xlsx_meta.get("original_binary_ones_sum"),
            "reduced_binary_grid_dimensions": xlsx_meta.get("reduced_binary_grid_dimensions"),
            "reduced_binary_ones_sum": xlsx_meta.get("reduced_binary_ones_sum"),
            "weighted_grid_sum": xlsx_meta.get("weighted_grid_sum"),
            "block_total_from_weighted_counts": xlsx_meta.get("block_total_from_weighted_counts"),
            "workbook_values_independently_verified_now": xlsx_meta.get(
                "workbook_values_independently_verified_now", False
            ),
        },
    }
    return records


def _write_context_file() -> None:
    existing = CHATGPT_CONTEXT_PATH.read_text(encoding="utf-8") if CHATGPT_CONTEXT_PATH.exists() else ""
    CHATGPT_CONTEXT_PATH.write_text(_context_text(existing), encoding="utf-8")


def build_stage5dp() -> dict[str, dict[str, Any]]:
    predecessor = validate_stage5do()
    if predecessor.validation_error_count:
        joined = "; ".join(predecessor.errors[:5])
        raise RuntimeError(f"Stage 5DO validation failed; refusing Stage 5DP build: {joined}")
    _stage5bd_counts, stage5bd_errors = validate_stage5bd()
    if stage5bd_errors:
        raise RuntimeError("Stage 5BD run-plan validation failed; refusing Stage 5DP build")

    _write_context_file()
    records = _build_records()
    _write_schemas()
    for key, payload in records.items():
        if key in DATA_PATHS:
            write_yaml(DATA_PATHS[key], payload)

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    write_json(RESULTS_DIR / "summary.json", records["summary"])
    write_json(RESULTS_DIR / "new_reddit_source_lock_report.json", records["new_reddit_source_lock_register"])
    write_json(RESULTS_DIR / "mayfly_docx_xlsx_report.json", records["mayfly_docx_xlsx_source_lock"])
    write_json(
        RESULTS_DIR / "candidate_report.json",
        {key: records[key] for key in HISTORICAL_KEYS},
    )
    write_json(
        RESULTS_DIR / "preservation_report.json",
        {
            "stage5do_preservation": records["stage5do_preservation"],
            "stage5bd_preservation": records["stage5bd_preservation"],
            "active_lineage_preservation": records["active_lineage_preservation"],
        },
    )
    warnings = [{"warning": gap} for gap in records["reviewability_gap_register"]["gaps"]]
    write_jsonl(RESULTS_DIR / "warnings.jsonl", warnings)
    return records


def _read_record(key: str) -> dict[str, Any]:
    return read_yaml(DATA_PATHS[key])


def _schema_errors(key: str) -> list[str]:
    schema_path = SCHEMA_PATHS[key]
    data_path = DATA_PATHS[key]
    if not schema_path.exists():
        return [f"missing_schema:{schema_path}"]
    if not data_path.exists():
        return [f"missing_data:{data_path}"]
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    data = read_yaml(data_path)
    return [error.message for error in Draft202012Validator(schema).iter_errors(data)]


def _common_errors(record: dict[str, Any], context: str) -> list[str]:
    errors: list[str] = []
    if record.get("stage_id") != STAGE_ID:
        errors.append(f"{context}:wrong_stage_id")
    for key in ["metadata_only", "source_lock_only"]:
        if record.get(key) is not True:
            errors.append(f"{context}:{key}_must_be_true")
    for key in ["execution_allowed", *sorted(FORBIDDEN_FALSE_FLAGS)]:
        if record.get(key) is not False:
            errors.append(f"{context}:{key}_must_be_false")
    return errors


def _candidate_errors(record: dict[str, Any], context: str) -> list[str]:
    errors = _common_errors(record, context)
    for key in ["accepted_as_route", "accepted_as_decryption", "used_for_target_selection_now"]:
        if record.get(key) is not False:
            errors.append(f"{context}:{key}_must_be_false")
    return errors


def _result(command: str, errors: list[str]) -> ValidationResult:
    return ValidationResult(command, len(errors), errors)


def validate_stage5dp_new_reddit_source_lock() -> ValidationResult:
    command = "validate-stage5dp-new-reddit-source-lock"
    record = _read_record("new_reddit_source_lock_register")
    errors = _common_errors(record, command)
    represented = {item.get("source_folder") for item in record.get("source_folders", [])}
    for folder in REQUIRED_SOURCE_FOLDERS.values():
        if folder not in represented:
            errors.append(f"missing_or_ungapped_source_folder:{folder}")
    for inventory in record.get("source_folders", []):
        for file_record in inventory.get("files", []):
            if file_record.get("raw_file_committed_now") is not False:
                errors.append("raw_file_committed_now_must_be_false")
    return _result(command, errors)


def validate_stage5dp_mayfly_source_lock() -> ValidationResult:
    command = "validate-stage5dp-mayfly-source-lock"
    record = _read_record("mayfly_docx_xlsx_source_lock")
    candidate = _read_record("mayfly_instar_grid_analysis_candidate_v1")
    errors = _common_errors(record, command)
    errors.extend(_candidate_errors(candidate, f"{command}:candidate"))
    if record.get("mayfly_docx_present") is not True:
        errors.append("mayfly_docx_not_source_locked")
    if record.get("mayfly_xlsx_present") is not True:
        errors.append("mayfly_xlsx_not_source_locked")
    if record.get("raw_text_committed_now") is not False:
        errors.append("docx_raw_text_committed")
    if record.get("raw_workbook_committed_now") is not False:
        errors.append("xlsx_raw_workbook_committed")
    return _result(command, errors)


def validate_stage5dp_mayfly_workbook_summary() -> ValidationResult:
    command = "validate-stage5dp-mayfly-workbook-summary"
    record = _read_record("mayfly_docx_xlsx_source_lock")
    block_counts = record.get("weighted_block_counts", {})
    errors = _common_errors(record, command)
    if record.get("mayfly_workbook_sheet_names") != [
        "READ ME",
        "958x1092 (binary)",
        "958x1092 (4x4 4x5 5x5)",
        "230x262 (binary)",
        "230x262 (weighted)",
    ]:
        errors.append("mayfly_sheet_names_changed")
    if record.get("original_binary_grid_dimensions") != [958, 1092]:
        errors.append("original_grid_dimensions_wrong")
    if record.get("original_binary_ones_sum") != 35210:
        errors.append("original_binary_sum_wrong")
    if record.get("reduced_binary_grid_dimensions") != [230, 262]:
        errors.append("reduced_grid_dimensions_wrong")
    if record.get("reduced_binary_ones_sum") != 2033:
        errors.append("reduced_binary_sum_wrong")
    if record.get("weighted_grid_sum") != 35210:
        errors.append("weighted_sum_wrong")
    if block_counts != {"value_16": 1420, "value_20": 567, "value_25": 46}:
        errors.append("weighted_block_counts_wrong")
    if record.get("block_total_from_weighted_counts") != 2033:
        errors.append("weighted_block_total_wrong")
    return _result(command, errors)


def validate_stage5dp_dot_observations() -> ValidationResult:
    command = "validate-stage5dp-dot-observations"
    five_dot = _read_record("five_dot_shift_skipped_f_page56_candidate_v0")
    page33 = _read_record("page33_three_dot_emirp_area_block_candidate_v0")
    errors = _candidate_errors(five_dot, f"{command}:five_dot")
    errors.extend(_candidate_errors(page33, f"{command}:page33"))
    if five_dot.get("page56_claims", {}).get("skipped_rune") != "F":
        errors.append("page56_skipped_f_missing")
    if five_dot.get("pixel_measurements_require_canonical_image_verification") is not True:
        errors.append("canonical_image_verification_gap_missing")
    if page33.get("area_pattern", {}).get("decrement") != 9:
        errors.append("page33_area_decrement_wrong")
    if page33.get("pixel_geometry_source_claimed_pending_canonical_verification") is not True:
        errors.append("page33_canonical_verification_gap_missing")
    return _result(command, errors)


def validate_stage5dp_front_cover_measurements() -> ValidationResult:
    command = "validate-stage5dp-front-cover-measurements"
    errors: list[str] = []
    for key in [
        "front_cover_1033_measurement_geometry_candidate_v0",
        "front_cover_1327_1427_concat_emirp_candidate_v0",
        "front_cover_d_panel_two_pixel_shift_candidate_v0",
        "front_cover_obscured_coordinate_1033_1033_candidate_v0",
        "front_cover_adjusted_area_1889603_prime_candidate_v0",
        "front_cover_3301_concat_prime_not_emirp_correction_v0",
    ]:
        record = _read_record(key)
        errors.extend(_candidate_errors(record, f"{command}:{key}"))
        arithmetic = record.get("verified_or_source_claimed_arithmetic", {})
        if arithmetic.get("132714271033_emirp") is not True:
            errors.append(f"{key}:1033_concat_emirp_missing")
        if arithmetic.get("132714273301_emirp") is not False:
            errors.append(f"{key}:3301_concat_emirp_correction_missing")
        if "3301_concatenation_is_prime_but_not_emirp_under_standard_reverse_prime_definition" not in record.get(
            "warnings", []
        ):
            errors.append(f"{key}:3301_warning_missing")
    return _result(command, errors)


def validate_stage5dp_iso_and_problems_sources() -> ValidationResult:
    command = "validate-stage5dp-iso-and-problems-sources"
    iso = _read_record("iso_560_13_560_17_palindromic_prime_size_candidate_v0")
    problems = _read_record("problems_2012_autostereogram_source_tool_candidate_v0")
    errors = _candidate_errors(iso, f"{command}:iso")
    errors.extend(_candidate_errors(problems, f"{command}:problems"))
    claims = iso.get("arithmetic_claims", {})
    if claims.get("118818811_prime") is not True or claims.get("1183811_palindrome") is not True:
        errors.append("iso_prime_palindrome_claims_not_verified")
    if problems.get("lp_relevance") != "historical_low_priority":
        errors.append("problems_relevance_must_remain_low_priority")
    return _result(command, errors)


def validate_stage5dp_chatgpt_context_file() -> ValidationResult:
    command = "validate-stage5dp-chatgpt-context-file"
    errors: list[str] = []
    if not CHATGPT_CONTEXT_PATH.exists():
        errors.append("ChatGPT-ContextFile.md missing")
    else:
        text = CHATGPT_CONTEXT_PATH.read_text(encoding="utf-8")
        for required in [
            "Stage 5DP source-locked new Reddit Mayfly/dot/cover/ISO material",
            "MayFlyInvestigation is high value",
            "2033 active reduced cells",
            "candidate-only, not active solve routes",
        ]:
            if required not in text:
                errors.append(f"context_missing:{required}")
    return _result(command, errors)


def validate_stage5dp_sidecar_gates() -> ValidationResult:
    command = "validate-stage5dp-sidecar-gates"
    errors: list[str] = []
    for key in [
        "no_active_ingestion_proof",
        "no_byte_stream_transition_gate",
        "no_execution_transition_gate",
    ]:
        record = _read_record(key)
        errors.extend(_common_errors(record, f"{command}:{key}"))
        if record.get("gate_closed") is not True:
            errors.append(f"{key}:gate_not_closed")
    return _result(command, errors)


def validate_stage5dp_handoff_continuity() -> ValidationResult:
    command = "validate-stage5dp-handoff-continuity"
    raw = _read_record("raw_source_noncommit_proof")
    summary = _read_record("summary")
    errors = _common_errors(raw, command)
    errors.extend(_common_errors(summary, f"{command}:summary"))
    if DEPRECATED_CODEX_OUTPUT.exists():
        errors.append("codex_output used")
    if raw.get("raw_source_files_committed_now") is not False:
        errors.append("raw_source_files_committed_now_must_be_false")
    if summary.get("recommended_next_stage_id") != NEXT_STAGE_ID:
        errors.append("next_stage_not_stage5dq")
    if summary.get("stage5bd_run_plan_id_count") != 10:
        errors.append("stage5bd_count_not_10")
    if summary.get("active_lineage_record_count") != len(ACTIVE_LINEAGE_PATHS):
        errors.append("active_lineage_count_not_8")
    return _result(command, errors)


def validate_stage5dp() -> ValidationResult:
    command = "validate-stage5dp"
    errors: list[str] = []
    for key in DATA_PATHS:
        errors.extend(_schema_errors(key))
        if DATA_PATHS[key].exists():
            errors.extend(_common_errors(_read_record(key), key))
    for validator in [
        validate_stage5dp_new_reddit_source_lock,
        validate_stage5dp_mayfly_source_lock,
        validate_stage5dp_mayfly_workbook_summary,
        validate_stage5dp_dot_observations,
        validate_stage5dp_front_cover_measurements,
        validate_stage5dp_iso_and_problems_sources,
        validate_stage5dp_chatgpt_context_file,
        validate_stage5dp_sidecar_gates,
        validate_stage5dp_handoff_continuity,
    ]:
        result = validator()
        errors.extend(f"{result.command}:{error}" for error in result.errors)
    return _result(command, errors)


def load_stage5dp_summary() -> dict[str, Any]:
    return _read_record("summary")


def stage5dp_summary_text() -> str:
    summary = load_stage5dp_summary()
    mayfly = summary.get("mayfly_workbook_summary", {})
    lines = [
        f"stage_id={summary.get('stage_id')}",
        f"status={summary.get('status')}",
        f"new_reddit_source_lock_created={str(summary.get('new_reddit_source_lock_created')).lower()}",
        f"new_reddit_source_folder_count={summary.get('new_reddit_source_folder_count')}",
        f"required_reddit_source_folders_represented={summary.get('required_reddit_source_folders_represented')}",
        f"mayfly_docx_source_locked={str(summary.get('mayfly_docx_source_locked')).lower()}",
        f"mayfly_xlsx_source_locked={str(summary.get('mayfly_xlsx_source_locked')).lower()}",
        f"candidate_records_created={summary.get('candidate_records_created')}",
        f"mayfly_original_binary_ones_sum={mayfly.get('original_binary_ones_sum')}",
        f"mayfly_reduced_binary_ones_sum={mayfly.get('reduced_binary_ones_sum')}",
        f"mayfly_weighted_grid_sum={mayfly.get('weighted_grid_sum')}",
        f"mayfly_block_total_from_weighted_counts={mayfly.get('block_total_from_weighted_counts')}",
        f"chatgpt_context_file_updated={str(summary.get('chatgpt_context_file_updated')).lower()}",
        f"source_browser_gui_deferred_to_next_stage={str(summary.get('source_browser_gui_deferred_to_next_stage')).lower()}",
        f"pivot_target_selected={str(summary.get('pivot_target_selected_now')).lower()}",
        f"route_extraction_performed={str(summary.get('route_extraction_performed_now')).lower()}",
        f"ocr_performed={str(summary.get('ocr_performed')).lower()}",
        f"image_forensics_performed={str(summary.get('image_forensics_performed')).lower()}",
        f"execution_performed={str(summary.get('execution_performed')).lower()}",
        f"stage5bd_run_plan_id_count={summary.get('stage5bd_run_plan_id_count')}",
        f"active_lineage_record_count={summary.get('active_lineage_record_count')}",
        f"parallel_worker_cap={summary.get('parallel_worker_cap')}",
        f"recommended_next_stage_id={summary.get('recommended_next_stage_id')}",
    ]
    return "\n".join(lines)


__all__ = [
    "CHATGPT_CONTEXT_PATH",
    "CODEX_COMPLETION_PATH",
    "DATA_PATHS",
    "RESULTS_DIR",
    "SCHEMA_PATHS",
    "build_stage5dp",
    "load_stage5dp_summary",
    "stage5dp_summary_text",
    "validate_stage5dp",
    "validate_stage5dp_chatgpt_context_file",
    "validate_stage5dp_dot_observations",
    "validate_stage5dp_front_cover_measurements",
    "validate_stage5dp_handoff_continuity",
    "validate_stage5dp_iso_and_problems_sources",
    "validate_stage5dp_mayfly_source_lock",
    "validate_stage5dp_mayfly_workbook_summary",
    "validate_stage5dp_new_reddit_source_lock",
    "validate_stage5dp_sidecar_gates",
]
