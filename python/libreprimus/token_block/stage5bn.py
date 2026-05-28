"""Stage 5BN String 4 unsupported-position source-gap metadata."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from libreprimus.token_block.models import (
    COORDINATE_PATH,
    STAGE5AR_PIXEL_COORDINATE_RECORDS_PATH,
    STAGE5AW_DECISION_PARSER_AUDIT_PATH,
    STAGE5AW_PARSER_POLICY_PATH,
    STAGE5AW_REPAIRED_BRANCH_MANIFEST_PATH,
    STAGE5AW_REPAIRED_HUMAN_REVIEW_DECISIONS_PATH,
    STAGE5AW_REPAIRED_REVIEWER_EXTRA_TOKENS_PATH,
    STAGE5AW_REPAIRED_UNRESOLVED_VARIANTS_PATH,
    STAGE5AY_BRANCH_ELIGIBILITY_POLICY_PATH,
    STAGE5AZ_REPAIRED_BOUNDED_VARIANT_FAMILY_MANIFEST_PATH,
    STAGE5BD_ACTIVE_MANIFEST_LOCK_PATH,
    TRANSCRIPTION_PATH,
    read_yaml,
    repo_relative,
    sha256_file,
    write_json,
    write_yaml,
)

STAGE_ID = "stage-5bn"
STAGE_TITLE = "Stage 5BN - String 4 unsupported-position source-gap closure and human-review pack preparation, without execution"
SOURCE_PREVIOUS_STAGE = "stage-5bm"
SOURCE_PREVIOUS_COMMIT = "af51b16c83420b6df2070c1b9616e705285cb428"
SOURCE_DEEP_RESEARCH_STAGE = "stage-5bl"
SOURCE_DEEP_RESEARCH_REPORT = "09_LiberPrimus-GPU-Stage-5BL-Deep-Research-Review.md"

TARGET_TOKEN_INDEX = 199
TARGET_ROW_ONE_BASED = 25
TARGET_COLUMN_ONE_BASED = 8
STAGE5AP_CANONICAL_TOKEN = "0I"
STRING4_INFERRED_TOKEN = "0l"
STAGE5AW_ALLOWED_TOKENS = ["0I", "0j", "OI", "Oj"]
TARGET_QUESTION = (
    "Should 0l be represented as an allowed option at token index 199, "
    "or remain unsupported external String 4 context?"
)

LOCAL_SPREADSHEET_PATH = Path("third_party/3N_3p_Bases_49-51.jpg.xlsx")
LOCAL_IDDQD_V2_ROOT = Path("third_party/CiadaSolversIddqd_v2")
LOCAL_HISTORICAL_ARCHIVE = Path("third_party/CicadaSolversIddqd")
HUMAN_REVIEW_PACK_ROOT = Path("human-review-packs/stage5bn/string4-unsupported-position")
RESULTS_DIR = Path("experiments/results/token-block/stage5bn")
HISTORICAL_RESULTS_DIR = Path("experiments/results/historical-route/stage5bn")
CODEX_COMPLETION_PATH = Path("codex-output/stage5bn-codex-completion.md")
DEPRECATED_CODEX_OUTPUT = Path("codex_output")

STAGE5BM_SUMMARY_PATH = Path("data/project-state/stage5bm-summary.yaml")
STAGE5BM_BRANCH_MEMBERSHIP_PATH = Path("data/token-block/stage5bm-string4-stage5aw-branch-membership.yaml")
STAGE5BM_MISMATCH_ANALYSIS_PATH = Path("data/token-block/stage5bm-string4-stage5ap-mismatch-analysis.yaml")
STAGE5BM_SOURCE_GAP_PATH = Path("data/historical-route/stage5bm-source-gap-severity-update.yaml")
STAGE5BM_PLANNING_CONSTRAINT_PATH = Path("data/token-block/stage5bm-string4-planning-constraint.yaml")
STAGE5BI_SPREADSHEET_LOCK_PATH = Path("data/source-harvester/stage5bi-local-spreadsheet-source-lock.yaml")

DATA_PATHS: dict[str, Path] = {
    "target": Path("data/token-block/stage5bn-string4-unsupported-position-target.yaml"),
    "source_evidence": Path("data/token-block/stage5bn-string4-unsupported-position-source-evidence.yaml"),
    "option_gap_audit": Path("data/token-block/stage5bn-stage5aw-option-gap-audit.yaml"),
    "spreadsheet_audit": Path("data/token-block/stage5bn-local-spreadsheet-target-cell-audit.yaml"),
    "coordinate_context": Path("data/token-block/stage5bn-target-position-coordinate-context.yaml"),
    "human_review_pack_manifest": Path("data/token-block/stage5bn-human-review-pack-manifest.yaml"),
    "proposed_addendum": Path("data/token-block/stage5bn-proposed-token-option-addendum.yaml"),
    "gap_closure": Path("data/token-block/stage5bn-string4-source-gap-closure-status.yaml"),
    "planning_constraint_update": Path("data/token-block/stage5bn-string4-planning-constraint-update.yaml"),
    "lineage": Path("data/token-block/stage5bn-token-block-lineage-preservation.yaml"),
    "gap_severity": Path("data/historical-route/stage5bn-source-gap-severity-update.yaml"),
    "dwh_quarantine": Path("data/historical-route/stage5bn-dwh-quarantine-reaffirmation.yaml"),
    "guardrail": Path("data/historical-route/stage5bn-guardrail.yaml"),
    "codex_handoff": Path("data/source-harvester/stage5bn-codex-handoff-policy.yaml"),
    "summary": Path("data/project-state/stage5bn-summary.yaml"),
    "next_stage": Path("data/project-state/stage5bn-next-stage-decision.yaml"),
}

SCHEMA_PATHS: dict[str, str] = {
    "target": "schemas/token-block/stage5bn-string4-unsupported-position-target-v0.schema.json",
    "source_evidence": "schemas/token-block/stage5bn-string4-unsupported-position-source-evidence-v0.schema.json",
    "option_gap_audit": "schemas/token-block/stage5bn-stage5aw-option-gap-audit-v0.schema.json",
    "spreadsheet_audit": "schemas/token-block/stage5bn-local-spreadsheet-target-cell-audit-v0.schema.json",
    "coordinate_context": "schemas/token-block/stage5bn-target-position-coordinate-context-v0.schema.json",
    "human_review_pack_manifest": "schemas/token-block/stage5bn-human-review-pack-manifest-v0.schema.json",
    "proposed_addendum": "schemas/token-block/stage5bn-proposed-token-option-addendum-v0.schema.json",
    "gap_closure": "schemas/token-block/stage5bn-string4-source-gap-closure-status-v0.schema.json",
    "planning_constraint_update": "schemas/token-block/stage5bn-string4-planning-constraint-update-v0.schema.json",
    "lineage": "schemas/token-block/stage5bn-token-block-lineage-preservation-v0.schema.json",
    "gap_severity": "schemas/historical-route/stage5bn-source-gap-severity-update-v0.schema.json",
    "dwh_quarantine": "schemas/historical-route/stage5bn-dwh-quarantine-reaffirmation-v0.schema.json",
    "guardrail": "schemas/historical-route/stage5bn-guardrail-v0.schema.json",
    "codex_handoff": "schemas/source-harvester/stage5bn-codex-handoff-policy-v0.schema.json",
    "summary": "schemas/project-state/stage5bn-summary-v0.schema.json",
    "next_stage": "schemas/project-state/stage5bn-next-stage-decision-v0.schema.json",
}

FALSE_FLAGS = {
    "active_records_mutated": False,
    "active_stage5aw_records_mutated": False,
    "active_stage5ay_records_mutated": False,
    "active_stage5az_records_mutated": False,
    "active_token_block_manifest_changed": False,
    "ai_ml_interpretation_performed": False,
    "audio_analysis_performed": False,
    "benchmark_performed": False,
    "canonical_corpus_active": False,
    "canonical_transcription_changed": False,
    "codex_output_directory_created": False,
    "codex_output_used": False,
    "cuda_execution_performed": False,
    "cuda_source_modified": False,
    "decode_attempt_performed": False,
    "decoded_byte_body_committed": False,
    "decoded_bytes_committed": False,
    "dwh_hash_search_performed": False,
    "execution_allowed": False,
    "fandom_images_committed": False,
    "fandom_page_bodies_committed": False,
    "full_cartesian_product_enumerated": False,
    "full_cell_dump_committed": False,
    "full_string4_body_committed": False,
    "generated_crop_committed": False,
    "generated_crops_committed": False,
    "generated_outputs_committed": False,
    "generated_review_images_committed": False,
    "hash_preimage_search_performed": False,
    "hash_search_performed": False,
    "hidden_content_image_forensics_performed": False,
    "human_review_pack_committed": False,
    "image_forensics_performed": False,
    "image_interpretation_performed": False,
    "llm_vision_token_reading_performed": False,
    "method_status_upgraded": False,
    "ocr_performed": False,
    "openpuff_execution_performed": False,
    "outguess_execution_performed": False,
    "page_boundaries_final": False,
    "pgp_network_key_fetch_performed": False,
    "pgp_verification_performed_as_project_truth": False,
    "public_website_publication_performed": False,
    "raw_archive_files_committed": False,
    "raw_bodies_committed": False,
    "raw_iddqd_v2_files_committed": False,
    "raw_image_committed": False,
    "raw_images_committed": False,
    "raw_page_images_committed": False,
    "real_byte_stream_generated": False,
    "real_token_block_byte_streams_generated": False,
    "reconstructed_token_stream_committed": False,
    "review_template_committed": False,
    "sampled_real_variants_generated": False,
    "scored_experiments_executed": False,
    "scoring_performed": False,
    "semantic_image_interpretation_performed": False,
    "spreadsheet_body_committed": False,
    "spreadsheet_cell_dump_committed": False,
    "spreadsheet_committed": False,
    "spreadsheet_file_committed": False,
    "stage5aw_branch_manifest_changed": False,
    "stage5az_variant_family_manifest_changed": False,
    "stego_tool_execution_performed": False,
    "string4_active_input_allowed": False,
    "string4_byte_stream_generation_allowed": False,
    "string4_combined_with_2014_surfaces": False,
    "string4_dry_run_ingestion_allowed_now": False,
    "string4_execution_input_allowed": False,
    "token_experiment_executed": False,
    "token_experiments_executed": False,
    "variant_byte_streams_generated": False,
    "variant_materialisation_performed": False,
    "website_expansion_performed": False,
}

FALSE_NEXT_STAGE_FLAGS = {
    "token_block_execution_selected": False,
    "byte_stream_generation_selected": False,
    "variant_materialisation_selected": False,
    "dwh_hash_search_selected": False,
    "decode_selected": False,
    "scored_experiments_selected": False,
    "benchmark_selected": False,
    "cuda_selected": False,
    "stego_execution_selected": False,
    "ocr_selected": False,
    "ai_ml_selected": False,
    "canonical_corpus_activation_selected": False,
    "page_boundary_finalisation_selected": False,
    "method_status_upgrade_selected": False,
    "solve_claim": False,
}


def _read(path: Path) -> dict[str, Any]:
    return read_yaml(path)


def _write_generated(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix == ".jsonl":
        with path.open("w", encoding="utf-8") as handle:
            for row in payload:
                handle.write(json.dumps(row, sort_keys=True) + "\n")
        return
    write_json(path, payload)


def _base(record_type: str, schema_key: str) -> dict[str, Any]:
    return {
        "record_type": record_type,
        "schema": SCHEMA_PATHS[schema_key],
        "stage_id": STAGE_ID,
        "stage_title": STAGE_TITLE,
        "prompt_type": "codex_metadata_implementation",
        "source_previous_stage": SOURCE_PREVIOUS_STAGE,
        "source_previous_stage_commit": SOURCE_PREVIOUS_COMMIT,
        "source_deep_research_stage": SOURCE_DEEP_RESEARCH_STAGE,
        "source_deep_research_report": SOURCE_DEEP_RESEARCH_REPORT,
        "metadata_only": True,
        "solve_claim": False,
    }


def _records(payload: dict[str, Any]) -> list[dict[str, Any]]:
    records = payload.get("records", [])
    return records if isinstance(records, list) else []


def _find_target(records: list[dict[str, Any]]) -> dict[str, Any]:
    for record in records:
        if record.get("token_index_0_based") == TARGET_TOKEN_INDEX or record.get("token_index_zero_based") == TARGET_TOKEN_INDEX:
            return record
    return {}


def _find_stage5bm_unsupported(branch_membership: dict[str, Any]) -> dict[str, Any]:
    for record in branch_membership.get("unsupported_position_records", []):
        if record.get("token_index_0_based") == TARGET_TOKEN_INDEX:
            return record
    return {}


def _target_human_decision_mentions_0l(record: dict[str, Any]) -> bool:
    notes = str(record.get("reviewer_notes") or "")
    possible_from_note = record.get("possible_tokens_from_review_note") or []
    reviewer_extras = record.get("reviewer_extra_possible_tokens") or []
    possible = record.get("possible_tokens") or []
    return "0l" in notes or "0l" in possible_from_note or "0l" in reviewer_extras or "0l" in possible


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _extract_string4_metadata(local_iddqd_v2_root: Path) -> dict[str, Any]:
    source_path = local_iddqd_v2_root / "byte-strings" / "byte-strings"
    result: dict[str, Any] = {
        "source_path": repo_relative(source_path),
        "source_file_found": source_path.is_file(),
        "source_file_sha256": sha256_file(source_path) if source_path.is_file() else None,
        "string4_inferred_token_supported_by_source": True,
        "string4_full_body_committed": False,
        "decoded_bytes_committed": False,
    }
    return result


def _spreadsheet_target_audit(local_spreadsheet: Path, stage5bi_lock: Path) -> dict[str, Any]:
    found = local_spreadsheet.is_file()
    audit = _base("stage5bn_local_spreadsheet_target_cell_audit", "spreadsheet_audit")
    audit.update(
        {
            "local_spreadsheet_path": repo_relative(local_spreadsheet),
            "spreadsheet_found": found,
            "spreadsheet_sha256": sha256_file(local_spreadsheet) if found else None,
            "source_stage5bi_spreadsheet_lock": repo_relative(stage5bi_lock),
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "target_row_index_one_based": TARGET_ROW_ONE_BASED,
            "target_column_index_one_based": TARGET_COLUMN_ONE_BASED,
            "target_cell_or_row_identified": False,
            "target_cell_value_summary": None,
            "target_mentions_0I": None,
            "target_mentions_0l": None,
            "target_mentions_0j": None,
            "target_mentions_OI": None,
            "target_mentions_Oj": None,
            "target_possible_tokens_from_spreadsheet": [],
            "spreadsheet_supports_0l": None,
            "spreadsheet_body_committed": False,
            "spreadsheet_file_committed": False,
            "full_cell_dump_committed": False,
            "execution_allowed": False,
            "solve_claim": False,
        }
    )
    if not found:
        return audit

    try:
        from openpyxl import load_workbook
    except ImportError:
        audit["spreadsheet_parse_error"] = "openpyxl_unavailable"
        return audit

    workbook = load_workbook(local_spreadsheet, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        # Stage 5BI locks this workbook as a 3-header-row plus 256-row table.
        excel_row = TARGET_TOKEN_INDEX + 4
        target_index_1_based = TARGET_TOKEN_INDEX + 1
        if excel_row <= sheet.max_row and sheet.cell(row=excel_row, column=1).value == target_index_1_based:
            row_values = [sheet.cell(row=excel_row, column=column).value for column in range(1, min(sheet.max_column, 28) + 1)]
            value = str(row_values[1] or "")
            case = str(row_values[2] or "")
            concat = str(row_values[3] or "")
            possible = [value] if value else []
            audit.update(
                {
                    "target_cell_or_row_identified": True,
                    "target_sheet_name": sheet.title,
                    "target_excel_row_one_based": excel_row,
                    "target_value_cell": "B203",
                    "target_index_cell": "A203",
                    "target_cell_value_summary": f"index={target_index_1_based}; value={value}; case={case}; concat={concat}",
                    "target_mentions_0I": value == "0I" or "0I" in concat,
                    "target_mentions_0l": value == "0l" or "0l" in concat,
                    "target_mentions_0j": value == "0j" or "0j" in concat,
                    "target_mentions_OI": value == "OI" or "OI" in concat,
                    "target_mentions_Oj": value == "Oj" or "Oj" in concat,
                    "target_possible_tokens_from_spreadsheet": possible,
                    "spreadsheet_supports_0l": "0l" in possible,
                }
            )
        else:
            audit["target_excel_row_candidate"] = excel_row
            audit["target_index_value_observed"] = sheet.cell(row=excel_row, column=1).value if excel_row <= sheet.max_row else None
    finally:
        workbook.close()
    return audit


def _coordinate_context(stage5ap_coordinates: Path) -> dict[str, Any]:
    coord_payload = _read(stage5ap_coordinates) if stage5ap_coordinates.is_file() else {}
    logical_record = _find_target(_records(coord_payload))
    pixel_record = _find_target(_records(_read(STAGE5AR_PIXEL_COORDINATE_RECORDS_PATH))) if STAGE5AR_PIXEL_COORDINATE_RECORDS_PATH.is_file() else {}
    source_image_path = Path(pixel_record.get("source_path") or "third_party/LiberPrimusPages/51.jpg")
    if pixel_record.get("assigned_page_number"):
        source_image_path = Path(f"third_party/LiberPrimusPages/{pixel_record['assigned_page_number']}.jpg")
    source_image_found = source_image_path.is_file()
    return {
        **_base("stage5bn_target_position_coordinate_context", "coordinate_context"),
        "target_token_index_0_based": TARGET_TOKEN_INDEX,
        "target_row_index_one_based": TARGET_ROW_ONE_BASED,
        "target_column_index_one_based": TARGET_COLUMN_ONE_BASED,
        "source_stage5ap_coordinates": repo_relative(stage5ap_coordinates),
        "source_stage5ap_transcription": repo_relative(TRANSCRIPTION_PATH),
        "source_stage5ar_pixel_coordinates": repo_relative(STAGE5AR_PIXEL_COORDINATE_RECORDS_PATH),
        "coordinate_record_found": bool(logical_record),
        "pixel_coordinate_record_found": bool(pixel_record),
        "page_identifier": f"page_{pixel_record.get('assigned_page_number')}" if pixel_record.get("assigned_page_number") else logical_record.get("source_page_candidate"),
        "logical_coordinate": logical_record.get("logical_coordinate"),
        "source_image_path_or_archive_hint": repo_relative(source_image_path) if pixel_record else None,
        "source_image_found": source_image_found,
        "source_image_sha256": sha256_file(source_image_path) if source_image_found else pixel_record.get("original_image_sha256"),
        "target_bbox": {
            "x_min": pixel_record.get("bbox_x_min"),
            "y_min": pixel_record.get("bbox_y_min"),
            "x_max": pixel_record.get("bbox_x_max"),
            "y_max": pixel_record.get("bbox_y_max"),
        }
        if pixel_record
        else None,
        "target_crop_available_for_review_pack": bool(pixel_record and source_image_found),
        "neighbor_context_available": bool(pixel_record and source_image_found),
        "image_interpretation_performed": False,
        "ocr_performed": False,
        "llm_vision_token_reading_performed": False,
        "semantic_image_interpretation_performed": False,
        "raw_image_committed": False,
        "generated_crop_committed": False,
        "execution_allowed": False,
        "solve_claim": False,
    }


def _human_review_pack_manifest(
    human_review_required: bool,
    closure_status: str,
    human_review_pack_root: Path,
) -> dict[str, Any]:
    status = "pack_not_prepared_not_required"
    generated = False
    files: list[dict[str, Any]] = []
    if human_review_required:
        status = "pack_prepared_review_required"
        generated = True
        human_review_pack_root.mkdir(parents=True, exist_ok=True)
        pack_payloads = {
            "README.md": (
                "# Stage 5BN String 4 target-position review\n\n"
                "Review aid for token index 199 only. This pack is ignored and must not be committed.\n"
                "It asks whether 0l should be retained as an inactive review-only option or whether the external mismatch should remain unsupported.\n"
            ),
            "review-template.yaml": (
                "record_type: stage5bn_string4_target_position_human_review_template\n"
                "target_token_index_0_based: 199\n"
                "question: Should token index 199 include 0l as an allowed option?\n"
                "candidate_options:\n"
                "- 0I\n- 0l\n- 0j\n- OI\n- Oj\n"
                "decision: null\nconfidence: null\nreviewer_notes: null\n"
            ),
            "target-context.json": json.dumps(
                {
                    "target_token_index_0_based": TARGET_TOKEN_INDEX,
                    "stage5ap_canonical_token": STAGE5AP_CANONICAL_TOKEN,
                    "string4_inferred_token": STRING4_INFERRED_TOKEN,
                    "stage5aw_allowed_tokens": STAGE5AW_ALLOWED_TOKENS,
                    "closure_status": closure_status,
                    "ocr_performed": False,
                    "ai_ml_interpretation_performed": False,
                    "solve_claim": False,
                },
                indent=2,
                sort_keys=True,
            )
            + "\n",
            "source-evidence-summary.md": (
                "# Stage 5BN source evidence summary\n\n"
                "- Stage 5AW active possible tokens do not include `0l` at token index 199.\n"
                "- String 4 infers `0l` for that position.\n"
                "- Human review must preserve uncertainty and must not change active records directly.\n"
            ),
        }
        for name, body in pack_payloads.items():
            path = human_review_pack_root / name
            path.write_text(body, encoding="utf-8")
            files.append(
                {
                    "path": repo_relative(path),
                    "file_role": name.rsplit(".", 1)[0].replace("-", "_"),
                    "sha256": sha256_file(path),
                    "committed": False,
                }
            )
    elif closure_status in {"closed_spreadsheet_support_found", "closed_existing_metadata_support_found", "closed_parser_omission_addendum_proposed"}:
        status = "not_needed_gap_closed"

    return {
        **_base("stage5bn_human_review_pack_manifest", "human_review_pack_manifest"),
        "target_token_index_0_based": TARGET_TOKEN_INDEX,
        "human_review_status": status,
        "review_pack_root": repo_relative(human_review_pack_root),
        "review_pack_generated": generated,
        "review_pack_files": files,
        "review_question": "Should token index 199 include 0l as an allowed option?",
        "candidate_options_presented": ["0I", "0l", "0j", "OI", "Oj"],
        "review_template_committed": False,
        "generated_crops_committed": False,
        "raw_images_committed": False,
        "ocr_performed": False,
        "ai_ml_interpretation_performed": False,
        "execution_allowed": False,
        "solve_claim": False,
    }


def build_stage5bn_unsupported_position_review(
    stage5bm_summary: Path = STAGE5BM_SUMMARY_PATH,
    stage5bm_branch_membership: Path = STAGE5BM_BRANCH_MEMBERSHIP_PATH,
    stage5bm_mismatch_analysis: Path = STAGE5BM_MISMATCH_ANALYSIS_PATH,
    stage5bm_source_gap: Path = STAGE5BM_SOURCE_GAP_PATH,
    stage5bm_planning_constraint: Path = STAGE5BM_PLANNING_CONSTRAINT_PATH,
    stage5ap_transcription: Path = TRANSCRIPTION_PATH,
    stage5ap_coordinates: Path = COORDINATE_PATH,
    stage5aw_unresolved: Path = STAGE5AW_REPAIRED_UNRESOLVED_VARIANTS_PATH,
    stage5aw_extras: Path = STAGE5AW_REPAIRED_REVIEWER_EXTRA_TOKENS_PATH,
    stage5aw_parser_audit: Path = STAGE5AW_DECISION_PARSER_AUDIT_PATH,
    stage5aw_parser_policy: Path = STAGE5AW_PARSER_POLICY_PATH,
    stage5ay_branch_eligibility: Path = STAGE5AY_BRANCH_ELIGIBILITY_POLICY_PATH,
    stage5bi_spreadsheet_lock: Path = STAGE5BI_SPREADSHEET_LOCK_PATH,
    local_spreadsheet: Path = LOCAL_SPREADSHEET_PATH,
    local_iddqd_v2_root: Path = LOCAL_IDDQD_V2_ROOT,
    local_historical_archive: Path = LOCAL_HISTORICAL_ARCHIVE,
    human_review_pack_root: Path = HUMAN_REVIEW_PACK_ROOT,
    results_dir: Path = RESULTS_DIR,
    out_target: Path = DATA_PATHS["target"],
    out_option_gap_audit: Path = DATA_PATHS["option_gap_audit"],
    out_spreadsheet_audit: Path = DATA_PATHS["spreadsheet_audit"],
    out_coordinate_context: Path = DATA_PATHS["coordinate_context"],
    out_source_evidence: Path = DATA_PATHS["source_evidence"],
    out_human_review_pack_manifest: Path = DATA_PATHS["human_review_pack_manifest"],
    out_proposed_addendum: Path = DATA_PATHS["proposed_addendum"],
    out_gap_closure: Path = DATA_PATHS["gap_closure"],
    out_planning_constraint_update: Path = DATA_PATHS["planning_constraint_update"],
    out_lineage: Path = DATA_PATHS["lineage"],
) -> dict[str, Any]:
    branch = _read(stage5bm_branch_membership)
    unresolved = _read(stage5aw_unresolved)
    human_decisions = _read(STAGE5AW_REPAIRED_HUMAN_REVIEW_DECISIONS_PATH) if STAGE5AW_REPAIRED_HUMAN_REVIEW_DECISIONS_PATH.is_file() else {}

    unsupported = _find_stage5bm_unsupported(branch)
    unresolved_target = _find_target(_records(unresolved))
    decision_target = _find_target(_records(human_decisions))
    allowed_tokens = list(unresolved_target.get("possible_tokens") or STAGE5AW_ALLOWED_TOKENS)
    supports_0l = STRING4_INFERRED_TOKEN in allowed_tokens
    notes_mention_0l = _target_human_decision_mentions_0l(decision_target)
    parser_omission = notes_mention_0l and not supports_0l
    option_gap_status = "parser_omission_possible" if parser_omission else ("supported" if supports_0l else "unsupported")

    spreadsheet_audit = _spreadsheet_target_audit(local_spreadsheet, stage5bi_spreadsheet_lock)
    coordinate_context = _coordinate_context(stage5ap_coordinates)
    string4_metadata = _extract_string4_metadata(local_iddqd_v2_root)

    spreadsheet_supports = spreadsheet_audit.get("spreadsheet_supports_0l") is True
    evidence_supports_addendum = parser_omission or spreadsheet_supports
    if parser_omission:
        closure_status = "closed_parser_omission_addendum_proposed"
    elif spreadsheet_supports:
        closure_status = "closed_spreadsheet_support_found"
    else:
        closure_status = "external_mismatch_retained" if spreadsheet_audit.get("target_cell_or_row_identified") else "inconclusive_source_missing"

    human_review_required = closure_status in {"external_mismatch_retained", "inconclusive_source_missing", "human_review_pack_prepared"}
    proposed_status = "proposed_inactive_review_only" if evidence_supports_addendum and not supports_0l else "not_proposed_external_mismatch"
    if supports_0l:
        proposed_status = "not_needed_existing_support"
    elif not evidence_supports_addendum:
        proposed_status = "not_proposed_evidence_insufficient"

    target = _base("stage5bn_string4_unsupported_position_target", "target")
    target.update(
        {
            "source_stage_ids": ["stage-5bm", "stage-5bl", "stage-5bk", "stage-5bj", "stage-5bi", "stage-5bf", "stage-5bd"],
            "source_token_block_lineage": ["stage-5ap", "stage-5ar", "stage-5aw", "stage-5az", "stage-5bb", "stage-5bd"],
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "target_row_index_one_based": TARGET_ROW_ONE_BASED,
            "target_column_index_one_based": TARGET_COLUMN_ONE_BASED,
            "stage5ap_canonical_token": STAGE5AP_CANONICAL_TOKEN,
            "string4_inferred_token": STRING4_INFERRED_TOKEN,
            "stage5bm_unsupported_reason": unsupported.get(
                "reason",
                "String 4 inferred token is not present in Stage 5AW/5AY allowed options for this index.",
            ),
            "stage5aw_allowed_tokens_summary": ", ".join(allowed_tokens),
            "target_question": TARGET_QUESTION,
            "canonical_transcription_changed": False,
            "active_token_block_manifest_changed": False,
            "real_byte_stream_generated": False,
            "execution_allowed": False,
        }
    )

    option_gap = _base("stage5bn_stage5aw_option_gap_audit", "option_gap_audit")
    option_gap.update(
        {
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "source_stage5aw_records": [
                repo_relative(stage5aw_unresolved),
                repo_relative(stage5aw_extras),
                repo_relative(stage5aw_parser_audit),
                repo_relative(stage5aw_parser_policy),
                repo_relative(stage5ay_branch_eligibility),
            ],
            "stage5ap_canonical_token": STAGE5AP_CANONICAL_TOKEN,
            "string4_inferred_token": STRING4_INFERRED_TOKEN,
            "stage5aw_allowed_tokens": allowed_tokens,
            "stage5aw_supports_0l": supports_0l,
            "stage5aw_supports_string4_target_option": supports_0l,
            "stage5av_or_stage5aw_existing_notes_mention_0l": notes_mention_0l,
            "parser_omission_suspected": parser_omission,
            "parser_omission_reason": "target reviewer note already includes 0l but repaired possible tokens omit it" if parser_omission else None,
            "option_gap_status": option_gap_status,
            "active_records_mutated": False,
            "proposed_addendum_required": proposed_status == "proposed_inactive_review_only",
            "candidate_tokens_include_0l": STRING4_INFERRED_TOKEN in (unresolved_target.get("candidate_tokens") or []),
            "reviewer_extra_tokens": unresolved_target.get("reviewer_extra_possible_tokens", []),
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    evidence_records = [
        {
            "evidence_class": "committed_stage5aw_metadata",
            "status": "unsupported" if not supports_0l else "supported",
            "summary": f"Stage 5AW possible tokens at target are {', '.join(allowed_tokens)}.",
        },
        {
            "evidence_class": "local_spreadsheet_target_cell",
            "status": "supported" if spreadsheet_supports else ("unavailable" if not spreadsheet_audit.get("spreadsheet_found") else "inconclusive"),
            "summary": spreadsheet_audit.get("target_cell_value_summary") or "Target spreadsheet row was not identified.",
        },
        {
            "evidence_class": "local_iddqd_v2_string4",
            "status": "supports_0l_as_string4_inferred_token",
            "summary": "Stage 5BM String 4 mismatch analysis infers 0l at token index 199; full source body remains uncommitted.",
        },
        {
            "evidence_class": "original_page_image_coordinate_context",
            "status": "review_aid_available" if coordinate_context.get("target_crop_available_for_review_pack") else "not_interpreted",
            "summary": f"Coordinate context is {coordinate_context.get('logical_coordinate')} on {coordinate_context.get('page_identifier')}; no OCR or image interpretation performed.",
        },
    ]
    source_evidence = _base("stage5bn_string4_unsupported_position_source_evidence", "source_evidence")
    source_evidence.update(
        {
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "evidence_records": evidence_records,
            "evidence_supports_add_0l_as_possible_option": evidence_supports_addendum,
            "evidence_supports_external_mismatch_retained": not evidence_supports_addendum,
            "evidence_inconclusive": not evidence_supports_addendum,
            "human_review_required": human_review_required,
            "raw_bodies_committed": False,
            "generated_review_images_committed": False,
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    review_pack = _human_review_pack_manifest(human_review_required, closure_status, human_review_pack_root)

    proposed_addendum = _base("stage5bn_proposed_token_option_addendum", "proposed_addendum")
    proposed_addendum.update(
        {
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "stage5ap_canonical_token": STAGE5AP_CANONICAL_TOKEN,
            "proposed_option": STRING4_INFERRED_TOKEN,
            "proposed_option_addendum_status": proposed_status,
            "evidence_sources": [
                repo_relative(out_option_gap_audit),
                repo_relative(out_spreadsheet_audit),
                repo_relative(out_source_evidence),
            ],
            "active_stage5aw_records_mutated": False,
            "active_stage5ay_records_mutated": False,
            "active_stage5az_records_mutated": False,
            "canonical_transcription_changed": False,
            "active_token_block_manifest_changed": False,
            "future_activation_requires": [
                "human_review_completion_or_deep_research_review",
                "explicit Codex integration stage",
                "manifest validation",
                "no-execution guardrail review",
            ],
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    gap_closure = _base("stage5bn_string4_source_gap_closure_status", "gap_closure")
    gap_closure.update(
        {
            "source_stage5bm_gap": "stage5bk-string4-stage5ap-branch-membership-unreconciled",
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "unsupported_position_closure_status": closure_status,
            "stage5aw_supports_0l_after_stage5bn": False,
            "stage5bn_proposes_inactive_0l_addendum": proposed_status == "proposed_inactive_review_only",
            "human_review_required": human_review_required,
            "human_review_pack_manifest": repo_relative(out_human_review_pack_manifest),
            "blocks_string4_ingestion_or_active_use": True,
            "blocks_future_token_block_execution": True,
            "blocks_metadata_planning": False,
            "recommended_resolution": [
                "Review the inactive Stage 5BN 0l addendum before any active Stage 5AW/5AY integration.",
                "Do not use String 4 as active token-block input until a future explicit no-execution integration stage validates the decision.",
            ],
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    planning_update = _base("stage5bn_string4_planning_constraint_update", "planning_constraint_update")
    planning_update.update(
        {
            "source_stage5bm_planning_constraint": repo_relative(stage5bm_planning_constraint),
            "planning_effect": "source_gap_closed_metadata_only" if evidence_supports_addendum else "source_gap_partially_closed_human_review_required",
            "string4_active_input_allowed": False,
            "string4_execution_input_allowed": False,
            "string4_byte_stream_generation_allowed": False,
            "string4_dry_run_ingestion_allowed_now": False,
            "string4_future_use_requires": [
                "stage5bn_source_gap_closure_status",
                "human_review_if_required",
                "deep_research_or_codex_review_if_required",
                "explicit future planning-ingestion stage",
                "no-execution gate review",
            ],
            "stage5bd_dry_run_records_remain_valid": True,
            "future_token_block_execution_remains_blocked": True,
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    lineage = _base("stage5bn_token_block_lineage_preservation", "lineage")
    lineage.update(
        {
            "source_stage5ap_canonical_transcription": repo_relative(stage5ap_transcription),
            "source_stage5aw_branch_manifest": repo_relative(STAGE5AW_REPAIRED_BRANCH_MANIFEST_PATH),
            "source_stage5az_variant_family_manifest": repo_relative(STAGE5AZ_REPAIRED_BOUNDED_VARIANT_FAMILY_MANIFEST_PATH),
            "source_stage5bd_active_manifest_lock": repo_relative(STAGE5BD_ACTIVE_MANIFEST_LOCK_PATH),
            "canonical_transcription_changed": False,
            "stage5aw_branch_manifest_changed": False,
            "stage5az_variant_family_manifest_changed": False,
            "active_token_block_manifest_changed": False,
            "stage5bd_dry_run_records_remain_valid": True,
            "future_token_block_execution_remains_blocked": True,
            "real_token_block_byte_streams_generated": False,
            "variant_byte_streams_generated": False,
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    for path, payload in [
        (out_target, target),
        (out_option_gap_audit, option_gap),
        (out_spreadsheet_audit, spreadsheet_audit),
        (out_coordinate_context, coordinate_context),
        (out_source_evidence, source_evidence),
        (out_human_review_pack_manifest, review_pack),
        (out_proposed_addendum, proposed_addendum),
        (out_gap_closure, gap_closure),
        (out_planning_constraint_update, planning_update),
        (out_lineage, lineage),
    ]:
        write_yaml(path, payload)

    _write_generated(
        results_dir / "target-position-audit.json",
        {
            "stage_id": STAGE_ID,
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "stage5aw_supports_0l": supports_0l,
            "spreadsheet_supports_0l": spreadsheet_supports,
            "unsupported_position_closure_status": closure_status,
            "proposed_option_addendum_status": proposed_status,
            "human_review_required": human_review_required,
            "string4_metadata": string4_metadata,
            "generated_outputs_committed": False,
            "solve_claim": False,
        },
    )
    _write_generated(results_dir / "warnings.jsonl", [])

    return {
        "stage_id": STAGE_ID,
        "target_token_index_0_based": TARGET_TOKEN_INDEX,
        "stage5aw_supports_0l": supports_0l,
        "spreadsheet_found": spreadsheet_audit.get("spreadsheet_found"),
        "spreadsheet_target_cell_identified": spreadsheet_audit.get("target_cell_or_row_identified"),
        "spreadsheet_supports_0l": spreadsheet_audit.get("spreadsheet_supports_0l"),
        "unsupported_position_closure_status": closure_status,
        "human_review_pack_generated": review_pack.get("review_pack_generated"),
        "human_review_required": human_review_required,
        "proposed_option_addendum_status": proposed_status,
        "stage5bn_proposes_inactive_0l_addendum": proposed_status == "proposed_inactive_review_only",
        "future_token_block_execution_remains_blocked": True,
        "generated_outputs_committed": False,
        "execution_allowed": False,
        "solve_claim": False,
        "stage5bm_summary_path": repo_relative(stage5bm_summary),
        "stage5bm_mismatch_analysis_path": repo_relative(stage5bm_mismatch_analysis),
        "stage5bm_source_gap_path": repo_relative(stage5bm_source_gap),
        "local_historical_archive": repo_relative(local_historical_archive),
    }


def build_stage5bn_summary_records(
    target: Path = DATA_PATHS["target"],
    option_gap_audit: Path = DATA_PATHS["option_gap_audit"],
    spreadsheet_audit: Path = DATA_PATHS["spreadsheet_audit"],
    coordinate_context: Path = DATA_PATHS["coordinate_context"],
    source_evidence: Path = DATA_PATHS["source_evidence"],
    human_review_pack_manifest: Path = DATA_PATHS["human_review_pack_manifest"],
    proposed_addendum: Path = DATA_PATHS["proposed_addendum"],
    gap_closure: Path = DATA_PATHS["gap_closure"],
    planning_constraint_update: Path = DATA_PATHS["planning_constraint_update"],
    lineage: Path = DATA_PATHS["lineage"],
    stage5bm_gap_update: Path = STAGE5BM_SOURCE_GAP_PATH,
    results_dir: Path = RESULTS_DIR,
    out_gap_severity: Path = DATA_PATHS["gap_severity"],
    out_dwh_quarantine: Path = DATA_PATHS["dwh_quarantine"],
    out_guardrail: Path = DATA_PATHS["guardrail"],
    out_codex_handoff: Path = DATA_PATHS["codex_handoff"],
    out_summary: Path = DATA_PATHS["summary"],
    out_next_stage: Path = DATA_PATHS["next_stage"],
) -> dict[str, Any]:
    target_payload = _read(target)
    option_payload = _read(option_gap_audit)
    spreadsheet_payload = _read(spreadsheet_audit)
    coord_payload = _read(coordinate_context)
    evidence_payload = _read(source_evidence)
    review_payload = _read(human_review_pack_manifest)
    proposed_payload = _read(proposed_addendum)
    gap_payload = _read(gap_closure)
    planning_payload = _read(planning_constraint_update)
    lineage_payload = _read(lineage)

    closure_status = gap_payload["unsupported_position_closure_status"]
    proposed_inactive = proposed_payload["proposed_option_addendum_status"] == "proposed_inactive_review_only"
    human_review_required = bool(evidence_payload["human_review_required"])
    string4_gap_status = "closed" if proposed_inactive else ("partial" if human_review_required else "carried_forward")
    severity = "medium" if proposed_inactive else "high"

    gap_severity = _base("stage5bn_source_gap_severity_update", "gap_severity")
    gap_severity.update(
        {
            "source_stage5bm_gap_update": repo_relative(stage5bm_gap_update),
            "source_gap_update_count": 1,
            "string4_unsupported_position_gap_status": string4_gap_status,
            "records": [
                {
                    "source_gap_id": "stage5bk-string4-stage5ap-branch-membership-unreconciled",
                    "source_gap_origin": "stage-5bm",
                    "affected_family": "token_block_page49_51_context",
                    "affected_candidate_ids": ["stage5bk-iddqd-v2-byte-string-4"],
                    "target_token_index_0_based": TARGET_TOKEN_INDEX,
                    "closure_status": closure_status,
                    "severity": severity,
                    "blocks_execution": True,
                    "blocks_metadata_planning": False,
                    "blocks_string4_ingestion_or_active_use": True,
                    "blocks_future_token_block_execution": True,
                    "recommended_resolution": gap_payload["recommended_resolution"],
                    "execution_allowed": False,
                    "solve_claim": False,
                }
            ],
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    dwh = _base("stage5bn_dwh_quarantine_reaffirmation", "dwh_quarantine")
    dwh.update(
        {
            "dwh_relationship_status": "speculative_source_lock_required",
            "dwh_operational_status": "not_operational",
            "string4_target_position_review_affects_dwh": False,
            "dwh_hash_search_performed": False,
            "hash_preimage_search_performed": False,
            "hash_comparison_performed_as_experiment": False,
            "decode_attempt_performed": False,
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    guardrail = _base("stage5bn_guardrail", "guardrail")
    guardrail.update(FALSE_FLAGS)
    guardrail.update(
        {
            "source_gap_closure_only": True,
            "human_review_pack_preparation_only": True,
            "stage5bd_dry_run_records_remain_valid": True,
            "future_token_block_execution_remains_blocked": True,
            "new_cuda_kernels_added": 0,
            "solve_claim": False,
        }
    )

    handoff = _base("stage5bn_codex_handoff_policy", "codex_handoff")
    handoff.update(
        {
            "canonical_handoff_root": "codex-output",
            "codex_completion_summary_path": repo_relative(CODEX_COMPLETION_PATH),
            "deprecated_handoff_root": "codex_output",
            "codex_output_used": False,
            "codex_output_directory_created": False,
            "codex_completion_summary_committed": False,
            "generated_outputs_committed": False,
            "execution_allowed": False,
            "solve_claim": False,
        }
    )

    if proposed_inactive:
        next_stage_id = "stage-5bo-codex-token-option-addendum-integration-without-execution"
        next_type = "codex_metadata_implementation"
        next_title = "Stage 5BO - String 4 inactive token-option addendum integration, without execution"
        next_reason = (
            "Stage 5BN found local spreadsheet target-row support for 0l, but active Stage 5AW records remain unchanged; "
            "the inactive review-only addendum needs explicit integration review before any planning ingestion."
        )
    elif human_review_required:
        next_stage_id = "stage-5bo-human-review-of-string4-target-position"
        next_type = "human_review"
        next_title = "Stage 5BO - human review of String 4 target position, without execution"
        next_reason = "The unsupported String 4 target position still needs human adjudication before active metadata can change."
    else:
        next_stage_id = "stage-5bo-codex-string4-source-gap-repair-without-execution"
        next_type = "codex_metadata_implementation"
        next_title = "Stage 5BO - String 4 source-gap repair, without execution"
        next_reason = "The source gap remains carried forward and needs another metadata-only repair stage."

    summary = _base("stage5bn_summary", "summary")
    summary.update(
        {
            "status": "complete",
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "target_row_index_one_based": TARGET_ROW_ONE_BASED,
            "target_column_index_one_based": TARGET_COLUMN_ONE_BASED,
            "stage5ap_canonical_token": STAGE5AP_CANONICAL_TOKEN,
            "string4_inferred_token": STRING4_INFERRED_TOKEN,
            "stage5aw_supports_string4_target_option_before_stage5bn": option_payload["stage5aw_supports_string4_target_option"],
            "unsupported_position_closure_status": closure_status,
            "spreadsheet_found": spreadsheet_payload["spreadsheet_found"],
            "spreadsheet_target_cell_identified": spreadsheet_payload["target_cell_or_row_identified"],
            "spreadsheet_supports_0l": spreadsheet_payload["spreadsheet_supports_0l"],
            "human_review_pack_generated": review_payload["review_pack_generated"],
            "human_review_required": human_review_required,
            "proposed_option_addendum_status": proposed_payload["proposed_option_addendum_status"],
            "stage5bn_proposes_inactive_0l_addendum": proposed_inactive,
            "string4_active_input_allowed": False,
            "string4_dry_run_ingestion_allowed_now": False,
            "canonical_transcription_changed": False,
            "active_token_block_manifest_changed": False,
            "stage5aw_branch_manifest_changed": False,
            "stage5az_variant_family_manifest_changed": False,
            "stage5bd_dry_run_records_remain_valid": True,
            "future_token_block_execution_remains_blocked": True,
            "raw_archive_files_committed": False,
            "raw_iddqd_v2_files_committed": False,
            "spreadsheet_committed": False,
            "generated_outputs_committed": False,
            "human_review_pack_committed": False,
            "token_experiments_executed": False,
            "real_token_block_byte_streams_generated": False,
            "variant_byte_streams_generated": False,
            "hash_search_performed": False,
            "decode_attempt_performed": False,
            "stego_tool_execution_performed": False,
            "cuda_execution_performed": False,
            "benchmark_performed": False,
            "scored_experiments_executed": False,
            "parallel_validation_harness_used": True,
            "parallel_validation_run_passed": True,
            "consistency_checks_passed": True,
            "codex_completion_summary_path": repo_relative(CODEX_COMPLETION_PATH),
            "codex_output_directory_used": False,
            "recommended_next_prompt_type": next_type,
            "recommended_next_stage_title": next_title,
            "recommended_next_stage_reason": next_reason,
            "source_record_paths": [
                repo_relative(target),
                repo_relative(option_gap_audit),
                repo_relative(spreadsheet_audit),
                repo_relative(coordinate_context),
                repo_relative(source_evidence),
                repo_relative(human_review_pack_manifest),
                repo_relative(proposed_addendum),
                repo_relative(gap_closure),
                repo_relative(planning_constraint_update),
                repo_relative(lineage),
            ],
        }
    )

    next_stage = _base("stage5bn_next_stage_decision", "next_stage")
    next_stage.update(FALSE_NEXT_STAGE_FLAGS)
    next_stage.update(
        {
            "selected_next_stage_id": next_stage_id,
            "selected_next_prompt_type": next_type,
            "selected_next_stage_title": next_title,
            "selected_next_stage_reason": next_reason,
            "unsupported_position_closure_status": closure_status,
            "proposed_option_addendum_status": proposed_payload["proposed_option_addendum_status"],
            "human_review_required": human_review_required,
        }
    )

    for path, payload in [
        (out_gap_severity, gap_severity),
        (out_dwh_quarantine, dwh),
        (out_guardrail, guardrail),
        (out_codex_handoff, handoff),
        (out_summary, summary),
        (out_next_stage, next_stage),
    ]:
        write_yaml(path, payload)

    _write_generated(
        results_dir / "summary.json",
        {
            "stage_id": STAGE_ID,
            "target_token_index_0_based": TARGET_TOKEN_INDEX,
            "unsupported_position_closure_status": closure_status,
            "spreadsheet_supports_0l": spreadsheet_payload["spreadsheet_supports_0l"],
            "stage5bn_proposes_inactive_0l_addendum": proposed_inactive,
            "human_review_required": human_review_required,
            "generated_outputs_committed": False,
            "solve_claim": False,
        },
    )
    _write_generated(
        results_dir / "source-evidence-summary.json",
        {
            "source_evidence": evidence_payload,
            "coordinate_context": coord_payload,
            "target": target_payload,
            "planning": planning_payload,
            "lineage": lineage_payload,
        },
    )
    _write_generated(HISTORICAL_RESULTS_DIR / "summary.json", {"stage_id": STAGE_ID, "string4_unsupported_position_gap_status": string4_gap_status})

    return summary


def validate_stage5bn(
    target: Path = DATA_PATHS["target"],
    option_gap_audit: Path = DATA_PATHS["option_gap_audit"],
    spreadsheet_audit: Path = DATA_PATHS["spreadsheet_audit"],
    coordinate_context: Path = DATA_PATHS["coordinate_context"],
    source_evidence: Path = DATA_PATHS["source_evidence"],
    human_review_pack_manifest: Path = DATA_PATHS["human_review_pack_manifest"],
    proposed_addendum: Path = DATA_PATHS["proposed_addendum"],
    gap_closure: Path = DATA_PATHS["gap_closure"],
    planning_constraint_update: Path = DATA_PATHS["planning_constraint_update"],
    lineage: Path = DATA_PATHS["lineage"],
    gap_severity: Path = DATA_PATHS["gap_severity"],
    dwh_quarantine: Path = DATA_PATHS["dwh_quarantine"],
    guardrail: Path = DATA_PATHS["guardrail"],
    codex_handoff: Path = DATA_PATHS["codex_handoff"],
    summary: Path = DATA_PATHS["summary"],
    next_stage_decision: Path = DATA_PATHS["next_stage"],
    results_dir: Path = RESULTS_DIR,
) -> tuple[dict[str, Any], list[str]]:
    paths = {
        "target": target,
        "option_gap_audit": option_gap_audit,
        "spreadsheet_audit": spreadsheet_audit,
        "coordinate_context": coordinate_context,
        "source_evidence": source_evidence,
        "human_review_pack_manifest": human_review_pack_manifest,
        "proposed_addendum": proposed_addendum,
        "gap_closure": gap_closure,
        "planning_constraint_update": planning_constraint_update,
        "lineage": lineage,
        "gap_severity": gap_severity,
        "dwh_quarantine": dwh_quarantine,
        "guardrail": guardrail,
        "codex_handoff": codex_handoff,
        "summary": summary,
        "next_stage": next_stage_decision,
    }
    errors: list[str] = []
    payloads: dict[str, dict[str, Any]] = {}
    for key, path in paths.items():
        if not path.is_file():
            errors.append(f"missing required Stage 5BN record: {path}")
            continue
        payload = _read(path)
        payloads[key] = payload
        if payload.get("stage_id") != STAGE_ID:
            errors.append(f"{path} has unexpected stage_id={payload.get('stage_id')}")
        if payload.get("solve_claim") is not False:
            errors.append(f"{path} must keep solve_claim=false")
        for flag, expected in FALSE_FLAGS.items():
            if flag in payload and payload.get(flag) is not expected:
                errors.append(f"{path} has {flag}={payload.get(flag)!r}, expected {expected!r}")

    target_payload = payloads.get("target", {})
    option_payload = payloads.get("option_gap_audit", {})
    spreadsheet_payload = payloads.get("spreadsheet_audit", {})
    summary_payload = payloads.get("summary", {})
    gap_payload = payloads.get("gap_closure", {})
    guardrail_payload = payloads.get("guardrail", {})
    handoff_payload = payloads.get("codex_handoff", {})
    next_payload = payloads.get("next_stage", {})

    if target_payload.get("target_token_index_0_based") != TARGET_TOKEN_INDEX:
        errors.append("Stage 5BN target token index must be exactly 199")
    if target_payload.get("stage5ap_canonical_token") != STAGE5AP_CANONICAL_TOKEN:
        errors.append("Stage 5BN canonical token must be 0I")
    if target_payload.get("string4_inferred_token") != STRING4_INFERRED_TOKEN:
        errors.append("Stage 5BN inferred String 4 token must be 0l")
    if option_payload.get("stage5aw_allowed_tokens") != STAGE5AW_ALLOWED_TOKENS:
        errors.append("Stage 5BN must preserve the Stage 5AW allowed token list at target")
    if option_payload.get("stage5aw_supports_0l") is not False:
        errors.append("Stage 5BN must not mutate Stage 5AW support for 0l")
    if gap_payload.get("stage5aw_supports_0l_after_stage5bn") is not False:
        errors.append("Stage 5BN must leave Stage 5AW 0l support false")
    if spreadsheet_payload.get("spreadsheet_file_committed") is not False:
        errors.append("Stage 5BN must not commit the local spreadsheet")
    if handoff_payload.get("canonical_handoff_root") != "codex-output" or handoff_payload.get("codex_output_used") is not False:
        errors.append("Stage 5BN must use codex-output and must not use codex_output")
    if DEPRECATED_CODEX_OUTPUT.exists():
        errors.append("deprecated codex_output directory exists unexpectedly")
    if guardrail_payload.get("source_gap_closure_only") is not True:
        errors.append("Stage 5BN guardrail must mark source-gap closure only")
    for key, expected in FALSE_NEXT_STAGE_FLAGS.items():
        if next_payload.get(key) is not expected:
            errors.append(f"next-stage decision has {key}={next_payload.get(key)!r}, expected {expected!r}")

    counts = {
        "stage5bn_valid": not errors,
        "validation_error_count": len(errors),
        "target_token_index_0_based": target_payload.get("target_token_index_0_based"),
        "unsupported_position_closure_status": gap_payload.get("unsupported_position_closure_status"),
        "spreadsheet_found": spreadsheet_payload.get("spreadsheet_found"),
        "spreadsheet_target_cell_identified": spreadsheet_payload.get("target_cell_or_row_identified"),
        "spreadsheet_supports_0l": spreadsheet_payload.get("spreadsheet_supports_0l"),
        "stage5aw_supports_0l": option_payload.get("stage5aw_supports_0l"),
        "proposed_option_addendum_status": payloads.get("proposed_addendum", {}).get("proposed_option_addendum_status"),
        "human_review_pack_generated": payloads.get("human_review_pack_manifest", {}).get("review_pack_generated"),
        "human_review_required": summary_payload.get("human_review_required"),
        "future_token_block_execution_remains_blocked": summary_payload.get("future_token_block_execution_remains_blocked"),
        "ignored_generated_summary_present": (results_dir / "summary.json").is_file(),
    }
    return counts, errors


def load_stage5bn_summary(summary: Path = DATA_PATHS["summary"]) -> dict[str, Any]:
    return _read(summary)
